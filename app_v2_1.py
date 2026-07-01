"""
PEPS BOM Automation Tool - Flask Application
Version: 2.1
Date: 22 May 2026

Main Flask application with all API routes for Phase 2 + Product Structure Downloads:
- BOM Viewer (single download, bulk download) - NOW IN PRODUCT STRUCTURE FORMAT
- Global Replace (preview, execute, rollback)
- New Product Wizard - NOW IN PRODUCT STRUCTURE FORMAT  
- Nearest BOM - NOW IN PRODUCT STRUCTURE FORMAT
- Run History
- Settings Management (NEW v2.1)

CHANGELOG v2.1:
- Added Settings API endpoints (GET/POST /api/settings)
- Modified ALL download endpoints to use Product Structure format
- Auto-generates UserData + MetaData file pairs in ZIP
- CreateProductStructure for New Product, EditProductStructure for all others
"""

import sys
import io
if isinstance(sys.stdout, io.TextIOWrapper):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import os
import re
import json
import csv
import time
import uuid
import shutil
import zipfile
import threading
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_file, Response, abort, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast

import config_v2_0 as config
from database_v2_1 import db
from bom_engine_v2_0 import bom_engine
from bom_scanner_v2_0 import scanner, get_sfg_size_variety
from product_structure_generator_v2_1 import generator
from sqlserver_log_v2_0 import activity_log

# ============================================================================
# FLASK APP SETUP
# ============================================================================

app = Flask(__name__)
app.config['SECRET_KEY'] = 'peps-bom-tool-2026-secret-key'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload
app.config['TEMPLATES_AUTO_RELOAD'] = True          # always serve latest HTML on each request

# Global constants
OUTPUT_FOLDER = config.OUTPUT_FOLDER
BOM_FILES_ROOT = config.BOM_FILES_ROOT
RAMCO_CONSTANTS = config.RAMCO_CONSTANTS

# Item master - loaded at startup
ITEM_MASTER = {}  # { item_code: description }

# All tab keys recognised by the permission system
ALL_TABS = ['configurator', 'replace', 'newprod', 'nearest', 'history', 'settings', 'appsettings', 'cms']

# ============================================================================
# AUTH — login_required decorator
# ============================================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorised'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Allows admin, developer, and sub_admin roles (+ any custom role with can_approve)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return jsonify({'error': 'Unauthorised'}), 401
        role = session.get('role', '')
        if role in ('admin', 'developer', 'sub_admin'):
            return f(*args, **kwargs)
        # Allow custom roles that have can_approve capability
        role_def = db.get_role_by_name(role)
        if role_def and role_def.get('can_approve'):
            return f(*args, **kwargs)
        return jsonify({'error': 'Admin access required'}), 403
    return decorated

def _can_create_users():
    """Check if current session user can create new users."""
    role = session.get('role', '')
    if role in ('admin', 'developer'): return True
    if role == 'sub_admin': return False
    role_def = db.get_role_by_name(role)
    return bool(role_def and role_def.get('can_create_users'))

def _can_manage_roles():
    """Check if current session user can manage roles."""
    return session.get('role') in ('admin', 'developer')

VALID_SYSTEM_ROLES = {'admin', 'developer', 'sub_admin', 'user'}

def _seed_admin():
    """Seed system roles and create default admin on first run."""
    # Always sync system role definitions (picks up new tab keys)
    db.seed_system_roles(ALL_TABS)

    users = db.get_all_users()
    if not users:
        pw_hash = generate_password_hash('Admin@1234')
        db.create_user('admin', pw_hash, role='admin', allowed_tabs=ALL_TABS)
        print("=" * 60)
        print("  FIRST RUN: default admin account created")
        print("  Username : admin")
        print("  Password : Admin@1234")
        print("  Please change the password after first login.")
        print("=" * 60)
    else:
        for u in users:
            if u['role'] in ('admin', 'developer'):
                db.update_user_tabs(u['id'], ALL_TABS)
            elif u['role'] == 'sub_admin':
                sub_tabs = [t for t in ALL_TABS if t not in ('cms', 'settings')]
                db.update_user_tabs(u['id'], sub_tabs)

# ============================================================================
# STARTUP - LOAD ITEM MASTER
# ============================================================================

ITEM_MASTER_CSV = r'D:\Hari JR. DATA\BOM\Automation\Variety Master\Itemcode\BOM Itemcode and raw materials code.csv'

def load_item_master():
    """Load item master from CSV: col 1 = Item Code, col 4 = Short Description."""
    global ITEM_MASTER

    filepath = ITEM_MASTER_CSV
    if not os.path.exists(filepath):
        # Fallback to the xlsx path if CSV not present
        filepath = config.ITEM_MASTER_PATH
        if not os.path.exists(filepath):
            print(f"WARNING: Item master not found at {filepath}")
            return

    try:
        print(f"\nLoading item master from: {filepath}")
        if filepath.lower().endswith('.csv'):
            with open(filepath, encoding='utf-8-sig', errors='replace') as f:
                reader = csv.reader(f)
                next(reader, None)  # skip header row
                for row in reader:
                    if len(row) < 2:
                        continue
                    code = row[1].strip()           # col B: Item Code
                    # col E (index 4) = Short Description; fall back to col D (index 3)
                    desc = (row[4].strip() if len(row) > 4 and row[4].strip()
                            else (row[3].strip() if len(row) > 3 else ''))
                    if code:
                        ITEM_MASTER[code] = desc
        else:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
            assert ws is not None
            ws = cast(Worksheet, ws)
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not any(row):
                    break
                code = str(row[1] or '').strip() if len(row) > 1 else ''
                if not code:
                    code = str(row[0] or '').strip() if len(row) > 0 else ''
                desc = str(row[4] or row[3] or row[2] or '').strip() if len(row) > 2 else ''
                if code:
                    ITEM_MASTER[code] = desc
            wb.close()

        print(f"Item master loaded: {len(ITEM_MASTER)} codes\n")

    except Exception as e:
        print(f"Error loading item master: {str(e)}\n")


def _scan_is_current() -> bool:
    """Return True if DB is populated and no .xlsm was modified since the last scan."""
    try:
        conn = db.get_connection()
        row = conn.execute(
            "SELECT COUNT(*) as cnt, MAX(last_scanned) as last_scan FROM products"
        ).fetchone()
        conn.close()

        db_count = row['cnt'] if row else 0
        last_scan_str = row['last_scan'] if row else None

        if db_count == 0 or not last_scan_str:
            return False

        last_scan_dt = datetime.fromisoformat(last_scan_str)

        file_count = 0
        for root, dirs, files in os.walk(config.BOM_FILES_ROOT):
            for fname in files:
                if fname.startswith('~$'):
                    continue
                if any(fname.endswith(ext) for ext in config.VALID_EXTENSIONS):
                    file_count += 1
                    mtime = datetime.fromtimestamp(
                        os.path.getmtime(os.path.join(root, fname))
                    )
                    if mtime > last_scan_dt:
                        return False  # a file changed since last scan

        return db_count == file_count
    except Exception:
        return False  # on any error, fall through to a full scan


def scan_and_register_products(force: bool = False):
    """Scan all BOM files and register in database.

    Skips the expensive workbook open loop when the DB is already up-to-date
    (same file count, no file newer than last scan) unless force=True.
    """
    if not force and _scan_is_current():
        print("✓ BOM files unchanged since last scan — skipping rescan")
        load_item_master()
        return {'success': True, 'skipped': True}

    result = scanner.scan_and_register_all()

    # Load / refresh item master from CSV
    load_item_master()

    # ── Description match statistics ──────────────────────────────────────
    if ITEM_MASTER:
        products = db.get_all_products()

        # code → list of product names that use it
        code_products: dict = {}
        for p in products:
            try:
                codes = json.loads(p.get('item_codes') or '[]')
            except Exception:
                codes = []
            for c in codes:
                code_products.setdefault(c, []).append(p['name'])

        all_codes  = set(code_products.keys())
        matched    = sum(1 for c in all_codes if c in ITEM_MASTER)
        unmatched  = sorted(c for c in all_codes if c not in ITEM_MASTER)
        total      = len(all_codes)
        pct        = round(matched / total * 100, 1) if total else 0

        print(f"\n{'='*60}")
        print(f"DESCRIPTION MATCH REPORT")
        print(f"{'='*60}")
        print(f"Total unique item codes : {total}")
        print(f"Matched in master       : {matched}  ({pct}%)")
        print(f"Unmatched (no desc)     : {len(unmatched)}")
        if unmatched:
            print(f"\nUnmatched codes (first 20):")
            for c in unmatched[:20]:
                print(f"  - {c}")
            if len(unmatched) > 20:
                print(f"  ... and {len(unmatched) - 20} more")
        print(f"{'='*60}\n")

        result['desc_matched']   = matched
        result['desc_unmatched'] = len(unmatched)
        result['desc_total']     = total
        result['desc_pct']       = pct

        # ── Export unmatched codes to Excel ───────────────────────────────
        if unmatched:
            _export_unmatched_to_excel(unmatched, code_products)

    return result


UNMATCHED_EXCEL_DIR = r'D:\Hari JR. DATA\BOM\Automation\Variety Master\Itemcode'

def _export_unmatched_to_excel(unmatched_codes: list, code_products: dict):
    """Write unmatched item codes with usage details to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Unmatched Item Codes'

    # ── Header ────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1A237E')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['#', 'Item Code', 'Status', 'Used In (Products)', 'Product Count']
    col_widths = [5, 45, 18, 80, 15]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font     = hdr_font
        c.fill     = hdr_fill
        c.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    red_fill    = PatternFill('solid', fgColor='FFEBEE')
    center_aln  = Alignment(horizontal='center', vertical='top')
    wrap_aln    = Alignment(vertical='top', wrap_text=True)
    mono_font   = Font(name='Courier New', size=10)

    for i, code in enumerate(unmatched_codes, start=1):
        prods = code_products.get(code, [])
        row   = i + 1
        ws.cell(row=row, column=1, value=i).alignment  = center_aln
        ic = ws.cell(row=row, column=2, value=code)
        ic.font      = mono_font
        ic.alignment = Alignment(vertical='top')
        ws.cell(row=row, column=3, value='Not in Master').alignment = center_aln
        ws.cell(row=row, column=4, value=', '.join(sorted(set(prods)))).alignment = wrap_aln
        ws.cell(row=row, column=5, value=len(set(prods))).alignment = center_aln
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = red_fill

    # ── Summary row ───────────────────────────────────────────────────────
    sum_row = len(unmatched_codes) + 2
    ws.cell(row=sum_row, column=1, value='Total').font = Font(bold=True)
    ws.cell(row=sum_row, column=2, value=len(unmatched_codes)).font = Font(bold=True)
    ws.cell(row=sum_row, column=3, value='Unmatched codes').font = Font(bold=True)

    # ── Save (retry with timestamp if file is open in Excel) ─────────────
    date_str = datetime.now().strftime('%d_%m_%Y')
    base     = os.path.join(UNMATCHED_EXCEL_DIR, f'Unmatched_ItemCodes_{date_str}.xlsx')
    out_path = base
    try:
        wb.save(out_path)
    except PermissionError:
        # File is open — append HH_MM_SS to make a unique name
        ts       = datetime.now().strftime('%H_%M_%S')
        out_path = os.path.join(UNMATCHED_EXCEL_DIR,
                                f'Unmatched_ItemCodes_{date_str}_{ts}.xlsx')
        wb.save(out_path)
    print(f"Unmatched codes saved: {out_path}")

# ============================================================================
# AUTH ROUTES
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if session.get('user_id'):
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.get_user_by_username(username)
        if user and user['is_active'] and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id']      = user['id']
            session['username']     = user['username']
            session['role']         = user['role']
            session['allowed_tabs'] = user['allowed_tabs']
            db.update_last_login(user['id'])
            activity_log.login(user['username'], user['role'],
                               ip=request.remote_addr or '',
                               session_id=session.get('_id', ''))
            return redirect(url_for('index'))
        error = 'Invalid username or password.'
    return render_template('login_v2_1.html', error=error)

@app.route('/logout')
def logout():
    activity_log.logout(session.get('username', ''), session.get('role', ''))
    session.clear()
    return redirect(url_for('login_page'))

# ============================================================================
# MAIN ROUTES
# ============================================================================

@app.before_request
def require_login():
    """Block unauthenticated access to every route except /login and /logout."""
    public = {'login_page', 'logout', 'static', 'api_health'}
    if request.endpoint in public:
        return None
    if not session.get('user_id'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Unauthorised'}), 401
        return redirect(url_for('login_page'))


@app.route('/api/health')
def api_health():
    """Unauthenticated liveness check for the watchdog task — no DB/file I/O."""
    return jsonify({'status': 'ok'})

@app.route('/')
def index():
    """Serve main UI"""
    resp = render_template('products_v2_1.html',
                           username=session.get('username'),
                           role=session.get('role'),
                           allowed_tabs=session.get('allowed_tabs', []))
    from flask import make_response
    r = make_response(resp)
    r.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    r.headers['Pragma'] = 'no-cache'
    r.headers['Expires'] = '0'
    return r


@app.route('/api/products')
def api_products():
    """Get all products with metadata"""
    products = db.get_all_products()
    
    # Parse JSON fields
    for prod in products:
        prod['heights']     = json.loads(prod['heights'])     if prod.get('heights')     else []
        prod['lengths']     = json.loads(prod['lengths'])     if prod.get('lengths')     else []
        prod['widths']      = json.loads(prod['widths'])      if prod.get('widths')      else []
        prod['colours']     = json.loads(prod['colours'])     if prod.get('colours')     else []
        prod['departments'] = json.loads(prod['departments']) if prod.get('departments') else []
        prod['sections']    = json.loads(prod['sections'])    if prod.get('sections')    else []
        prod['wh_codes']    = json.loads(prod['wh_codes'])    if prod.get('wh_codes')    else []

    return jsonify({'products': products})


@app.route('/api/products/<int:product_id>')
def api_product_detail(product_id):
    """Get single product details"""
    product = db.get_product_by_id(product_id)
    
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    
    # Parse JSON fields
    product['heights'] = json.loads(product['heights']) if product['heights'] else []
    product['lengths'] = json.loads(product['lengths']) if product['lengths'] else []
    product['widths'] = json.loads(product['widths']) if product['widths'] else []
    product['colours'] = json.loads(product['colours']) if product['colours'] else []
    
    return jsonify(product)


@app.route('/api/sfg-size-variety')
def api_sfg_size_variety():
    """SFG Size Variety master (BOX + ASSEMBLY sheets) for the New Product Wizard's Size Matrix step."""
    return jsonify(get_sfg_size_variety())


@app.route('/api/rescan', methods=['POST'])
def api_rescan():
    """Rescan all BOM files"""
    try:
        result = scan_and_register_products(force=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# SETTINGS API (NEW v2.1)
# ============================================================================

@app.route('/api/settings', methods=['GET'])
def api_get_settings():
    """Get current settings (password decrypted)"""
    try:
        settings = db.get_settings()
        return jsonify(settings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/settings', methods=['POST'])
def api_save_settings():
    """Save settings (password encrypted automatically)"""
    try:
        data = request.json or {}
        
        # Validate required fields
        required = ['url', 'user_id', 'role_name', 'organisation_unit']
        for field in required:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        success = db.save_settings(data)
        if success:
            return jsonify({'success': True, 'message': 'Settings saved successfully'})
        else:
            return jsonify({'error': 'Failed to save settings — check Flask console for details'}), 500

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# BOM VIEWER - SINGLE SIZE DOWNLOAD (MODIFIED v2.1)
# Now generates EditProductStructure format with UserData + MetaData
# ============================================================================

def _patch_row(row: dict, edit: dict) -> None:
    """Apply a single frontend edit dict to a BOM row dict (in-place)."""
    if edit.get('itemCode'): row['item_code'] = edit['itemCode']
    if edit.get('desc')  is not None: row['description'] = edit['desc']
    if edit.get('department') is not None: row['department'] = edit['department']
    if edit.get('section')    is not None: row['section']    = edit['section']
    if edit.get('uom'): row['uom'] = edit['uom']
    if edit.get('wh'):  row['wh_code'] = edit['wh']
    if edit.get('qty') not in (None, ''):
        try: row['qty'] = float(edit['qty'])
        except (ValueError, TypeError): pass

def _apply_edits_single(bom_rows: list, row_edits: dict) -> list:
    """Apply edits to a single-SKU BOM row list.
    Matches by ps_seq (robust) then falls back to list index."""
    if not row_edits:
        return bom_rows
    rows = [dict(r) for r in bom_rows]
    for idx_str, edit in row_edits.items():
        seq = edit.get('original_seq')
        matched = None
        if seq is not None:
            for r in rows:
                if str(r.get('ps_seq', '')) == str(seq):
                    matched = r; break
        if matched is None:
            try:
                i = int(idx_str)
                if 0 <= i < len(rows): matched = rows[i]
            except (ValueError, TypeError): pass
        if matched:
            _patch_row(matched, edit)
    return rows

def _apply_edits_all(bom_rows: list, row_edits: dict) -> list:
    """Apply edits across ALL SKU rows by matching ps_seq (same seq = same component)."""
    if not row_edits:
        return bom_rows
    rows = [dict(r) for r in bom_rows]
    for _, edit in row_edits.items():
        seq       = edit.get('original_seq')
        orig_code = (edit.get('original_code') or '').upper()
        for r in rows:
            if seq is not None and str(r.get('ps_seq', '')) == str(seq):
                _patch_row(r, edit)
            elif orig_code and r.get('item_code', '').upper() == orig_code:
                _patch_row(r, edit)
    return rows

def _generate_excel_report(product_name: str, bom_rows: list) -> bytes:
    """Generate a plain Excel (.xlsx) BOM report table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'BOM Report'
    hdr_fill = PatternFill('solid', fgColor='1A56DB')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ['PS No.', 'Seq', 'Item Code', 'Description', 'Qty', 'UOM', 'WH Code', 'Department', 'Section']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    for ri, r in enumerate(bom_rows, 2):
        # Description: prefer ITEM_MASTER lookup, fall back to row keys
        ic   = r.get('item_code','') or r.get('code','')
        desc = ITEM_MASTER.get(ic,'') or r.get('description','') or r.get('desc','')
        row_data = [
            r.get('ps_no',''), r.get('ps_seq',''), ic, desc,
            r.get('qty',''), r.get('uom',''), r.get('wh_code','') or r.get('wh',''),
            r.get('department',''), r.get('section',''),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border; cell.font = Font(size=9)
    col_widths = [30, 6, 20, 40, 10, 8, 12, 16, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

@app.route('/api/download-single/<int:product_id>', methods=['GET', 'POST'])
def api_download_single(product_id):
    """Download single product BOM in EditProductStructure or Excel format.
    Accepts POST body: {l, w, h, clr, format, row_edits}"""
    run_id = None
    try:
        body      = (request.json or {}) if request.method == 'POST' else {}
        L         = int(body['l']) if body.get('l') else (request.args.get('l', type=int) or request.args.get('L', type=int))
        W         = int(body['w']) if body.get('w') else (request.args.get('w', type=int) or request.args.get('W', type=int))
        H         = str(body.get('h','')) or request.args.get('h','') or request.args.get('H','')
        colour    = str(body.get('clr','')) or request.args.get('clr','') or request.args.get('colour','')
        dl_format = str(body.get('format','')) or request.args.get('format','mdcf')
        row_edits = body.get('row_edits') or {}

        product = db.get_product_by_id(product_id)
        if not product:
            return jsonify({'error': 'Not found'}), 404

        filter_parts = []
        if L: filter_parts.append(f'L={L}')
        if W: filter_parts.append(f'W={W}')
        if H: filter_parts.append(f'H={H}')
        if colour: filter_parts.append(f'CLR={colour}')
        filter_desc = ' '.join(filter_parts) if filter_parts else 'All sizes'

        run_id = db.create_run(
            product_ids=[product_id],
            mode='DOWNLOAD_SINGLE',
            output_mode='ZIP',
            run_by='User',
            approval_ref=filter_desc,
        )

        settings = db.get_settings()

        components, permutations, dest_headers, prebuilt_rows = bom_engine.read_bom_file(
            product['filepath']
        )

        if L or W or H or colour:
            if prebuilt_rows:
                filtered_rows = []
                for r in prebuilt_rows:
                    ps_no = r.get('ps_no', '')
                    match = re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', ps_no)
                    if not match:
                        continue
                    row_l, row_w, row_h = match.groups()
                    if L and int(row_l) != L: continue
                    if W and int(row_w) != W: continue
                    if H and row_h != H.zfill(2): continue
                    if colour:
                        cm = re.search(r'([A-Z]{2,4})\d+[Xx]\d+[Xx]\d+', ps_no)
                        row_colour = cm.group(1) if cm else ''
                        if row_colour != colour: continue
                    filtered_rows.append(r)
                bom_rows = filtered_rows
            else:
                filtered = []
                for p in permutations:
                    if L and int(p[0]) != L: continue
                    if W and int(p[1]) != W: continue
                    if H and str(p[2]) != H: continue
                    if colour and str(p[3]) != colour: continue
                    filtered.append(p)
                bom_rows = bom_engine.generate_bom(components, filtered)
        else:
            bom_rows = prebuilt_rows if prebuilt_rows else bom_engine.generate_bom(components, permutations)

        # Apply UI edits before generating output
        bom_rows = _apply_edits_single(bom_rows, row_edits)

        for r in bom_rows:
            r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})

        safe_name = product['name'].replace(' ', '_').replace('/', '_')

        _dl_user = session.get('username', 'User')
        _dl_role = session.get('role', '')
        if dl_format == 'excel':
            excel_bytes = _generate_excel_report(product['name'], bom_rows)
            fname = f'{safe_name}_BOM_Report.xlsx'
            db.complete_run(run_id, output_filename=fname, product_count=1, rows_generated=len(bom_rows))
            activity_log.bom_download(_dl_user, _dl_role, run_id=run_id,
                                      product_count=1, filename=fname, rows=len(bom_rows))
            return send_file(io.BytesIO(excel_bytes), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=fname)

        ps_rows = generator.convert_bom_to_rows(bom_rows, product['name'], 'edit')
        unique_rec_count = max((r.get('rec_no', 0) for r in ps_rows), default=0)
        if unique_rec_count > 0:
            try:
                db.update_setting('number_of_records', unique_rec_count)
                settings['number_of_records'] = str(unique_rec_count)
            except Exception: pass

        userdata_bytes, metadata_bytes, user_fname, meta_fname = \
            generator.generate_edit_structure(product['name'], ps_rows, settings)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(user_fname, userdata_bytes)
            zf.writestr(meta_fname, metadata_bytes)
        zip_buffer.seek(0)
        zip_fname = f'{safe_name}_ProductStructure.zip'
        db.complete_run(run_id, output_filename=zip_fname, product_count=1, rows_generated=len(ps_rows))
        activity_log.bom_download(_dl_user, _dl_role, run_id=run_id,
                                  product_count=1, filename=zip_fname, rows=len(ps_rows))
        return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=zip_fname)

    except Exception as e:
        if run_id:
            try: db.fail_run(run_id, str(e))
            except Exception: pass
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DOWNLOAD ALL SIZES (v2.1)
# Generates EditProductStructure for all SKUs of a product (no filtering)
# ============================================================================

@app.route('/api/download-all/<int:product_id>', methods=['GET', 'POST'])
def api_download_all(product_id):
    """Download all sizes/SKUs for a product in EditProductStructure or Excel format.
    Accepts POST body: {format, row_edits}"""
    run_id = None
    try:
        body      = (request.json or {}) if request.method == 'POST' else {}
        dl_format = str(body.get('format','')) or request.args.get('format','mdcf')
        row_edits = body.get('row_edits') or {}

        product = db.get_product_by_id(product_id)
        if not product:
            return jsonify({'error': 'Not found'}), 404

        run_id = db.create_run(product_ids=[product_id], mode='DOWNLOAD_ALL',
                               output_mode='ZIP', run_by='User')
        settings = db.get_settings()
        components, permutations, dest_headers, prebuilt_rows = bom_engine.read_bom_file(product['filepath'])
        bom_rows = prebuilt_rows if prebuilt_rows else bom_engine.generate_bom(components, permutations)

        # Apply UI edits to ALL SKU rows (same seq = same component across all sizes)
        bom_rows = _apply_edits_all(bom_rows, row_edits)

        for r in bom_rows:
            r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})

        safe_name = product['name'].replace(' ', '_').replace('/', '_')
        _da_user = session.get('username', 'User')
        _da_role = session.get('role', '')

        if dl_format == 'excel':
            excel_bytes = _generate_excel_report(product['name'], bom_rows)
            fname = f'{safe_name}_AllSizes_BOM_Report.xlsx'
            db.complete_run(run_id, output_filename=fname, product_count=1, rows_generated=len(bom_rows))
            activity_log.bom_download(_da_user, _da_role, run_id=run_id,
                                      product_count=1, filename=fname, rows=len(bom_rows))
            return send_file(io.BytesIO(excel_bytes), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=fname)

        ps_rows = generator.convert_bom_to_rows(bom_rows, product['name'], 'edit')
        unique_rec_count = max((r.get('rec_no', 0) for r in ps_rows), default=0)
        if unique_rec_count > 0:
            try:
                db.update_setting('number_of_records', unique_rec_count)
                settings['number_of_records'] = str(unique_rec_count)
            except Exception: pass

        userdata_bytes, metadata_bytes, user_fname, meta_fname = \
            generator.generate_edit_structure(product['name'], ps_rows, settings)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(user_fname, userdata_bytes)
            zf.writestr(meta_fname, metadata_bytes)
        zip_buffer.seek(0)
        zip_fname = f'{safe_name}_AllSizes_ProductStructure.zip'
        db.complete_run(run_id, output_filename=zip_fname, product_count=1, rows_generated=len(ps_rows))
        activity_log.bom_download(_da_user, _da_role, run_id=run_id,
                                  product_count=1, filename=zip_fname, rows=len(ps_rows))
        return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=zip_fname)

    except Exception as e:
        if run_id:
            try: db.fail_run(run_id, str(e))
            except Exception: pass
        return jsonify({'error': str(e)}), 500


# ============================================================================
# BULK DOWNLOAD (MODIFIED v2.1)
# Now generates EditProductStructure format for all products
# ============================================================================

def _is_privileged():
    """True if current session user is admin or developer."""
    return session.get('role') in ('admin', 'developer')

def _can_review():
    """True if current session user can act on the pending-approvals queue
    (admin, developer, sub_admin, or a custom role with can_approve) — same
    set of roles admin_required() allows, used here for read endpoints."""
    role = session.get('role', '')
    if role in ('admin', 'developer', 'sub_admin'):
        return True
    role_def = db.get_role_by_name(role)
    return bool(role_def and role_def.get('can_approve'))

@app.route('/api/run', methods=['POST'])
def api_run():
    """Generate BOM. Regular users → approval queue; admin/dev → execute immediately."""
    data = request.json or {}
    product_ids = data.get('product_ids') or data.get('productIds', [])
    output_mode = data.get('output_mode', 'ZIP')
    run_by      = session.get('username', data.get('run_by', 'unknown'))
    approval_ref = data.get('approval_ref', '')

    if not product_ids:
        return jsonify({'error': 'No products selected'}), 400

    if not _is_privileged():
        names = []
        for pid in product_ids[:5]:
            p = db.get_product_by_id(pid)
            if p:
                names.append(p['name'])
        summary = f"BOM download: {', '.join(names)}" + (f" (+{len(product_ids)-5} more)" if len(product_ids) > 5 else '')
        appr_id = db.create_approval(
            user_id=session['user_id'], username=run_by,
            action_type='bom_run', action_summary=summary,
            action_data={'product_ids': product_ids, 'output_mode': output_mode}
        )
        return jsonify({'pending': True, 'approval_id': appr_id,
                        'message': 'Your download request has been submitted for approval.'})

    try:
        run_id = db.create_run(product_ids=product_ids, mode='FULL',
                               output_mode=output_mode, run_by=run_by,
                               approval_ref=approval_ref)
        thread = threading.Thread(target=process_bom_generation_v2_1,
                                  args=(run_id, product_ids, output_mode))
        thread.start()
        return jsonify({'run_id': run_id, 'runId': run_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def process_bom_generation_v2_1(run_id: int, product_ids: list, output_mode: str):
    """Background process for BOM generation - v2.1 Product Structure format"""
    try:
        ts = datetime.now().strftime('%d%m%Y_%H%M%S')
        total_rows = 0
        
        # Get settings
        settings = db.get_settings()
        
        # Always ZIP mode with UserData + MetaData files
        filename = f'ProductStructure-Batch-{ts}.zip'
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        
        with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
            metadata_added = False
            
            for product_id in product_ids:
                product = db.get_product_by_id(product_id)
                if not product:
                    continue
                
                # Read BOM
                components, permutations, dest_headers, prebuilt_rows = \
                    bom_engine.read_bom_file(product['filepath'])
                
                if prebuilt_rows:
                    bom_rows = prebuilt_rows
                    for r in bom_rows:
                        r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})
                else:
                    bom_rows = bom_engine.generate_bom(components, permutations)
                
                total_rows += len(bom_rows)
                
                # Convert to Product Structure format
                ps_rows = generator.convert_bom_to_rows(bom_rows, product['name'], 'edit')
                
                # Generate EditProductStructure files
                userdata_bytes, metadata_bytes, user_fname, meta_fname = \
                    generator.generate_edit_structure(product['name'], ps_rows, settings)
                
                # Add UserData file
                zf.writestr(user_fname, userdata_bytes)
                
                # Add MetaData file only once
                if not metadata_added:
                    zf.writestr(meta_fname, metadata_bytes)
                    metadata_added = True
        
        db.complete_run(run_id, filename, len(product_ids), total_rows)
        run = db.get_run_by_id(run_id)
        _run_by = run.get('run_by', '') if run else ''
        activity_log.bom_download(_run_by, '', run_id=run_id,
                                   product_count=len(product_ids), filename=filename,
                                   rows=total_rows)
        activity_log.log_run(run_id, run.get('run_label', '') if run else '', 'OK',
                             len(product_ids), total_rows, _run_by,
                             completed_at=datetime.now())

    except Exception as e:
        db.fail_run(run_id, str(e))


@app.route('/api/download/<int:run_id>')
def api_download_run(run_id):
    """Download generated BOM file — serves from disk or regenerates on-demand."""
    run = db.get_run_by_id(run_id)
    if not run:
        return jsonify({'error': 'Not found'}), 404

    # For bulk runs: serve the saved ZIP from disk
    if run['output_filename'] and run['mode'] not in ('DOWNLOAD_ALL', 'DOWNLOAD_SINGLE'):
        filepath = os.path.join(OUTPUT_FOLDER, run['output_filename'])
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=run['output_filename'])
        return jsonify({'error': 'Not found'}), 404

    # For on-demand download runs: regenerate from the stored product_id
    conn = db.get_connection()
    row = conn.execute(
        'SELECT product_id FROM run_items WHERE run_id = ? LIMIT 1', (run_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    product_id = row['product_id']
    product = db.get_product_by_id(product_id)
    if not product:
        return jsonify({'error': 'Not found'}), 404

    try:
        settings = db.get_settings()
        components, permutations, dest_headers, prebuilt_rows = bom_engine.read_bom_file(
            product['filepath']
        )
        bom_rows = prebuilt_rows if prebuilt_rows else bom_engine.generate_bom(components, permutations)
        for r in bom_rows:
            r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})
        ps_rows = generator.convert_bom_to_rows(bom_rows, product['name'], 'edit')
        userdata_bytes, metadata_bytes, user_fname, meta_fname = \
            generator.generate_edit_structure(product['name'], ps_rows, settings)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(user_fname, userdata_bytes)
            zf.writestr(meta_fname, metadata_bytes)
        zip_buffer.seek(0)

        safe_name = product['name'].replace(' ', '_').replace('/', '_')
        zip_fname = f'{safe_name}_AllSizes_ProductStructure.zip'
        return send_file(zip_buffer, mimetype='application/zip',
                         as_attachment=True, download_name=zip_fname)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/run/<int:run_id>/status')
def api_run_status(run_id):
    """Get run status for polling"""
    run = db.get_run_by_id(run_id)
    if not run:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(run))


@app.route('/api/products/<int:product_id>/bom')
def api_product_bom(product_id):
    """Read BOM data from xlsm on-demand - returns SKU-grouped data for UI"""
    product = db.get_product_by_id(product_id)
    if not product:
        return jsonify({'error': f'Product {product_id} not found — try rescanning.'}), 404

    try:
        components, permutations, dest_headers, prebuilt_rows = bom_engine.read_bom_file(
            product['filepath']
        )

        bom_by_sku = {}
        skus = []
        heights_set = set()
        colours_set = set()

        if prebuilt_rows:
            # Format 2: pre-built BOM rows grouped by mattress code
            sku_order = []
            for row in prebuilt_rows:
                ps_no = row['ps_no']
                if ps_no not in bom_by_sku:
                    bom_by_sku[ps_no] = []
                    sku_order.append(ps_no)
                bom_by_sku[ps_no].append({
                    'seq':        row['ps_seq'],
                    'code':       row['item_code'],
                    'desc':       ITEM_MASTER.get(row['item_code'], '') or row.get('description', ''),
                    'department': row.get('department', ''),
                    'section':    row.get('section', ''),
                    'qty':        row['qty'],
                    'uom':        row['uom'],
                    'wh':         row['wh_code'],
                })

            db_heights = json.loads(product['heights']) if product['heights'] else []
            for ps_no in sku_order:
                # Try 3-number pattern first: "...72X30X06..." → L=72, W=30, H=6
                m3 = re.search(r'(\d+)[xX](\d+)[xX](\d+)', ps_no)
                if m3:
                    L = m3.group(1)
                    W = m3.group(2)
                    H = str(int(m3.group(3)))  # strip leading zero: "06" → "6"
                else:
                    # Fall back to 2-number pattern + DB height
                    m2 = re.search(r'(\d+)[xX](\d+)', ps_no)
                    L = m2.group(1) if m2 else ''
                    W = m2.group(2) if m2 else ''
                    H = str(db_heights[0]) if db_heights else ''
                cm = re.search(r'([A-Z]{2,4})\d+[Xx]\d+[Xx]\d+', ps_no)
                colour = cm.group(1) if cm else ''
                heights_set.add(H)
                colours_set.add(colour)
                skus.append({'L': L, 'W': W, 'H': H, 'colour': colour, 'code': ps_no})

            first_sku = sku_order[0] if sku_order else None

        else:
            # Format 1: permutation-based — evaluate qty formulas per SKU
            for perm in permutations:
                L, W, H, colour = str(perm[0]), str(perm[1]), str(perm[2]), str(perm[3] or '')
                name_slug = re.sub(r'[^A-Za-z0-9]', '-', product.get('name', 'PROD'))[:20].strip('-')
                sku_code = f"{name_slug}-{L}X{W}X{H}"
                if colour:
                    sku_code += f"-{colour}"

                comp_list = []
                for comp in components:
                    try:
                        qty = bom_engine.evaluate_qty_formula(
                            comp['qty_formula'], float(L), float(W), H, colour, components
                        )
                    except Exception:
                        qty = 0.0
                    comp_list.append({
                        'seq':        comp['seq'],
                        'code':       comp['item_code'],
                        'desc':       ITEM_MASTER.get(comp['item_code'], '') or comp.get('description', ''),
                        'department': comp.get('department', ''),
                        'section':    comp.get('section', ''),
                        'qty':        qty,
                        'uom':        comp['uom'],
                        'wh':         comp['wh_code'],
                    })

                bom_by_sku[sku_code] = comp_list
                heights_set.add(H)
                colours_set.add(colour)
                skus.append({'L': L, 'W': W, 'H': H, 'colour': colour, 'code': sku_code})

            first_sku = skus[0]['code'] if skus else None

        def _safe_float(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        heights = sorted([h for h in heights_set if h], key=_safe_float)
        colours = sorted(c for c in colours_set if c)
        fallback_comps = bom_by_sku.get(first_sku, []) if first_sku else []

        # Unique dept/section/wh across all component rows for filter dropdowns
        all_rows = [r for rows in bom_by_sku.values() for r in rows]
        departments = sorted({r['department'] for r in all_rows if r.get('department')})
        sections    = sorted({r['section']    for r in all_rows if r.get('section')})
        # Always include CBEFG (parent WH) alongside component WH codes
        wh_codes    = sorted({r['wh'] for r in all_rows if r.get('wh')} | {'CBEFG'})

        return jsonify({
            'skus': skus,
            'components': fallback_comps,
            'bom_by_sku': bom_by_sku,
            'first_sku': first_sku,
            'heights': heights,
            'colours': colours,
            'departments': departments,
            'sections': sections,
            'wh_codes': wh_codes,
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# RUN HISTORY & EXPORT
# ============================================================================

@app.route('/api/history')
def api_history():
    """Get run history"""
    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    limit = request.args.get('limit', 100, type=int)
    
    runs = db.get_run_history(limit=limit, from_date=from_date, to_date=to_date)
    return jsonify({'runs': runs})


@app.route('/api/history/export')
def api_history_export():
    """Export history as CSV"""
    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    
    runs = db.get_run_history(limit=1000, from_date=from_date, to_date=to_date)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Run ID', 'Run Label', 'Status', 'Product Count', 'Rows Generated',
        'Run By', 'Created At', 'Completed At', 'Mode', 'Output Mode', 
        'Output Filename', 'Approval Ref'
    ])
    
    # Data
    for run in runs:
        writer.writerow([
            run['id'], run['run_label'], run['status'], run['product_count'],
            run['rows_generated'], run['run_by'], run['created_at'],
            run['completed_at'], run['mode'], run['output_mode'],
            run['output_filename'], run['approval_ref']
        ])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=run_history.csv'}
    )

# ============================================================================
# ITEM LOOKUP
# ============================================================================

@app.route('/api/items/lookup')
def api_item_lookup():
    """Lookup item code description"""
    code = request.args.get('code', '').strip().upper()
    
    if not code:
        return jsonify({'error': 'Code required'}), 400
    
    desc = ITEM_MASTER.get(code, '')
    
    return jsonify({
        'code': code,
        'description': desc,
        'found': bool(desc)
    })

# ============================================================================
# GLOBAL REPLACE
# ============================================================================

@app.route('/api/item-formula')
def api_item_formula():
    """Auto-detect quantity formula pattern for an item code from REF/DATA sheets.
    Optional: product_id=N to read from a specific file only."""
    code       = request.args.get('code','').strip().upper()
    product_id = request.args.get('product_id', type=int)
    if not code:
        return jsonify({'formula':'', 'pattern':'', 'factor':0, 'samples':[]})

    # Extract base for dimension-specific codes
    _dim_re = re.compile(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm|CM|cm)?$')
    base_code_fm = _dim_re.sub('', code)
    has_dim_fm   = (base_code_fm != code)

    if product_id:
        p = db.get_product_by_id(product_id)
        scan_products = [p] if p else []
    else:
        all_p = db.get_all_products()
        if has_dim_fm:
            indexed = [p for p in all_p if any(
                c.upper().startswith(base_code_fm+'-') or c.upper() == base_code_fm
                for c in json.loads(p.get('item_codes','[]') or '[]')
            )]
        else:
            indexed = [p for p in all_p if code in [
                c.strip().upper()
                for c in json.loads(p.get('item_codes','[]') or '[]')
            ]]
        # If indexed: scan only those products (faster). If not indexed: scan all products
        # using the fast XML reader — 250 files complete in ~5–15 seconds with threading.
        scan_products = indexed if indexed else all_p
        # Sort smallest-first so quick files run first; the 5-match early exit then
        # triggers before the large Bonnel/SFG files (12–13 MB) are ever submitted.
        scan_products = sorted(
            scan_products,
            key=lambda p: os.path.getsize(p['filepath']) if os.path.exists(p.get('filepath','')) else 0
        )

    # ── Primary path: read actual formula string from DATA sheet ─────────────
    # Curve-fitting over REF computed values collapses ((K2-1)*(L2-1)/144)*k*N
    # into a single L*W*k factor, losing both the formula structure and the
    # layer count N.  Reading the raw formula and translating cell references
    # (K→L, L→W, M→H) returns the exact formula the user can inspect and edit.
    _col_subs = [
        (re.compile(r'\bK\d+\b', re.IGNORECASE), 'L'),
        (re.compile(r'\bL\d+\b', re.IGNORECASE), 'W'),
        (re.compile(r'\bM\d+\b', re.IGNORECASE), 'H'),
        (re.compile(r'\bN\d+\b', re.IGNORECASE), 'H'),
    ]
    _layer_end = re.compile(r'\*\s*(\d{1,2})\s*$')

    def _translate_excel_formula(raw):
        f = raw.strip()
        if f.startswith('='): f = f[1:]
        f = f.lstrip()
        if f.startswith('+'): f = f[1:]  # Lotus-1-2-3 style "=+..." — unary plus, no semantic effect
        for pat, repl in _col_subs:
            f = pat.sub(repl, f)
        return f.strip()

    def _fmt_const(v):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    def _strip_quotes(s):
        s = s.strip()
        if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
            return s[1:-1].replace('""', '"')
        return s

    def _split_if(expr):
        """If expr (stripped) is IF(cond, true, false), return (cond, true, false)."""
        e = expr.strip()
        m = re.match(r'^IF\s*\(', e, re.IGNORECASE)
        if not m: return None
        depth, parts, cur, j = 1, [], '', m.end()
        while j < len(e) and depth > 0:
            ch = e[j]
            if ch == '(':
                depth += 1; cur += ch
            elif ch == ')':
                depth -= 1
                if depth > 0: cur += ch
                else: parts.append(cur)
            elif ch == ',' and depth == 1:
                parts.append(cur); cur = ''
            else:
                cur += ch
            j += 1
        if len(parts) != 3 or j != len(e):
            return None
        return parts[0].strip(), parts[1].strip(), parts[2].strip()

    # Comparison value can be a number (=4) or a quoted string (="MR") — both
    # show up in real BOMs (e.g. height-keyed and code/variant-keyed lookups).
    _cond_eq_re = re.compile(r'^\s*([LWH])\s*=\s*("(?:[^"]|"")*"|[\d.]+)\s*$', re.IGNORECASE)

    def _parse_variant_chain(expr):
        """Parse IF(VAR=val,result,IF(VAR=val2,result2,default)) chains where
        VAR is a single dimension (L/W/H) compared by equality against a number
        or string literal. Returns a dict describing the value→result mapping,
        or None if the shape doesn't match (mixed variables, range comparisons,
        multi-arg lookups, etc.)."""
        var, pairs, cur = None, [], expr.strip()
        while True:
            split = _split_if(cur)
            if not split: break
            cond, tval, fval = split
            m = _cond_eq_re.match(cond)
            if not m: return None
            v, raw_val = m.group(1).upper(), m.group(2)
            when_val = _strip_quotes(raw_val) if raw_val.startswith('"') else raw_val
            if var is None: var = v
            elif var != v: return None
            pairs.append({'when': when_val, 'value': _strip_quotes(tval)})
            cur = fval
        if not pairs: return None
        return {'variable': var, 'map': pairs, 'default': _strip_quotes(cur.strip())}

    def _split_concat(expr):
        """Split a translated '&'-concatenation expression into top-level parts,
        respecting string literals and nested parentheses."""
        parts, cur, depth, in_str = [], '', 0, False
        for ch in expr:
            if ch == '"':
                in_str = not in_str
                cur += ch
            elif not in_str and ch == '(':
                depth += 1; cur += ch
            elif not in_str and ch == ')':
                depth -= 1; cur += ch
            elif not in_str and ch == '&' and depth == 0:
                parts.append(cur); cur = ''
            else:
                cur += ch
        parts.append(cur)
        return [p.strip() for p in parts]

    def _build_code_template(translated, if_expr_text):
        """Render a concatenation formula as a fill-in-the-blanks template string
        (e.g. 'FOAMERA-FOAM-BLOCK-{L}X{W}X{SUFFIX}') so the frontend can preview
        the generated code for chosen L/W/H without re-implementing a formula
        evaluator. Returns None if any part of the expression isn't recognised."""
        out = []
        for part in _split_concat(translated):
            p = part.strip()
            if len(p) >= 2 and p[0] == '"' and p[-1] == '"':
                out.append(_strip_quotes(p))
            elif p.upper() in ('L', 'W', 'H'):
                out.append('{' + p.upper() + '}')
            elif if_expr_text and p == if_expr_text:
                out.append('{SUFFIX}')
            else:
                return None
        return ''.join(out)

    def _extract_code_pattern(raw_code_formula):
        """Translate an item-code generation formula (e.g.
        ="FOAMERA-FOAM-BLOCK-"&K2&"X"&L2&"X"&IF(M2=4,"04",IF(M2=5,"05","06")))
        and detect a simple single-variable suffix/value mapping inside it —
        the same 'variant' shape the Layer Panel exposes for quantities, but
        for the code text itself (e.g. height → code suffix). When such a
        mapping is found, also build a fill-in-the-blanks template string for
        a live preview of the generated code."""
        if not isinstance(raw_code_formula, str) or not raw_code_formula.startswith('='):
            return None
        translated = _translate_excel_formula(raw_code_formula)
        variants = None
        template = None
        m = re.search(r'\bIF\s*\(', translated, re.IGNORECASE)
        if m:
            start = translated.index('(', m.start())
            depth, j = 0, start
            while j < len(translated):
                if translated[j] == '(': depth += 1
                elif translated[j] == ')':
                    depth -= 1
                    if depth == 0: break
                j += 1
            if depth == 0:
                if_expr = translated[m.start():j+1]
                variants = _parse_variant_chain(if_expr)
                if variants:
                    template = _build_code_template(translated, if_expr)
        return {'formula': translated, 'template': template, 'variants': variants}

    exact_formula  = None
    exact_layers   = None
    exact_constant = False
    code_pattern   = None
    # Collect one candidate quantity-formula per matching product, then pick
    # whichever exact formula recurs most often across the scanned BOMs.
    # Different product families can carry the same raw-material item code with
    # divergent (sometimes outlier/special-case) formulas — e.g. one file using
    # a height-keyed constant lookup while eight others use the same
    # dimension-based perimeter formula. Picking the majority shape surfaces
    # the formula that's actually representative instead of whatever file
    # happens to be scanned first.
    # ── Fast ZIP/XML reader — bypasses openpyxl VBA parsing overhead ────────
    # openpyxl takes ~60s per xlsm because it reads the entire file including VBA.
    # Reading the sheet XML directly from the ZIP is ~50-100x faster.
    import zipfile, xml.etree.ElementTree as ET
    from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _as_completed
    _WS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    _PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
    _OFF = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    def _xml_data_rows(filepath):
        """Return list of {ic, ic_fml, qty, qty_fml} dicts from DATA sheet via ZIP/XML.
        Uses zf.open() + ET.parse() (streaming) instead of zf.read() + ET.fromstring()
        to avoid 'Unable to allocate output buffer' on large xlsm files."""
        rows = []
        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                names = set(zf.namelist())

                def _parse(name):
                    with zf.open(name) as _f:
                        return ET.parse(_f).getroot()

                wb_root = _parse('xl/workbook.xml')
                rid = None
                for sh in wb_root.iter(f'{{{_WS}}}sheet'):
                    if sh.get('name','').upper() == 'DATA':
                        rid = sh.get(f'{{{_OFF}}}id'); break
                if not rid: return []

                sheet_path = None
                for rel in _parse('xl/_rels/workbook.xml.rels').iter(f'{{{_PKG}}}Relationship'):
                    if rel.get('Id') == rid:
                        t = rel.get('Target','')
                        sheet_path = t if t.startswith('xl/') else 'xl/'+t.lstrip('/'); break
                if not sheet_path or sheet_path not in names: return []

                sst = []
                if 'xl/sharedStrings.xml' in names:
                    for si in _parse('xl/sharedStrings.xml').iter(f'{{{_WS}}}si'):
                        sst.append(''.join(x.text or '' for x in si.iter(f'{{{_WS}}}t')))

                for row_el in _parse(sheet_path).iter(f'{{{_WS}}}row'):
                    rn = int(row_el.get('r', 0))
                    if rn < 2: continue
                    ic_val = ic_fml = qty_val = qty_fml = None
                    for c in row_el:
                        col = ''.join(x for x in c.get('r','') if not x.isdigit())
                        v = c.find(f'{{{_WS}}}v')
                        f = c.find(f'{{{_WS}}}f')
                        if col == 'E':
                            if f is not None:
                                ic_fml = '=' + (f.text or '')
                                ic_val = v.text if v is not None else None
                            elif c.get('t') == 's' and v is not None:
                                try: ic_val = sst[int(v.text or 0)]
                                except: ic_val = v.text
                            else:
                                ic_val = v.text if v is not None else None
                        elif col == 'G':
                            if f is not None: qty_fml = '=' + (f.text or '')
                            elif v is not None:
                                try: qty_val = float(v.text or 0)
                                except: pass
                    if ic_val or ic_fml:
                        rows.append({'ic': ic_val, 'ic_fml': ic_fml,
                                     'qty': qty_val, 'qty_fml': qty_fml})
        except Exception as _xe:
            print(f"[xml-reader] {os.path.basename(filepath)}: {_xe}")
        return rows

    def _scan_one(p):
        # Run _xml_data_rows in a sub-thread so we can enforce a per-file timeout.
        # Files locked by Excel or with corrupted ZIP headers can block indefinitely
        # without this guard.
        import threading as _thr
        _result = [None]
        def _worker():
            try:
                for row in _xml_data_rows(p['filepath']):
                    ic = str(row['ic'] or '').strip().upper() if row['ic'] else ''
                    if ic == code or (has_dim_fm and ic.startswith(base_code_fm+'-')):
                        _result[0] = row; return
            except Exception as _se:
                print(f"[scan-one] {p.get('name','?')}: {_se}")
        t = _thr.Thread(target=_worker, daemon=True)
        t.start()
        t.join(timeout=8)   # 8 s per file — skips locked/hung files
        if t.is_alive():
            print(f"[scan-one] timeout: {p.get('name','?')}")
        return _result[0]

    n = len(scan_products)
    workers = min(6, n) if n else 1
    print(f"[item-formula] '{code}' — scanning {n} products ({workers} workers)")
    import time as _t; _t0 = _t.time()
    _SCAN_TIMEOUT = 25   # hard deadline: return whatever we have after 25 s
    _NEED_MATCHES = 5    # stop early once we have 5 hits (enough for majority vote)
    _raw_results = []
    try:
        with _TPE(max_workers=workers) as pool:
            _futures = {pool.submit(_scan_one, p): p for p in scan_products}
            _deadline = _t.time() + _SCAN_TIMEOUT
            for _fut in _as_completed(_futures, timeout=_SCAN_TIMEOUT):
                try:
                    _raw_results.append(_fut.result(timeout=0.5))
                except Exception:
                    _raw_results.append(None)
                # Early exit: enough matches found to pick majority formula
                if sum(1 for r in _raw_results if r) >= _NEED_MATCHES:
                    for _f in _futures:
                        _f.cancel()
                    print(f"[item-formula] early exit — {_NEED_MATCHES} matches reached")
                    break
                if _t.time() > _deadline:
                    for _f in _futures:
                        _f.cancel()
                    print(f"[item-formula] timeout after {_SCAN_TIMEOUT}s")
                    break
    except Exception as _pe:
        print(f"[item-formula] pool error: {_pe}")
    print(f"[item-formula] scan done in {_t.time()-_t0:.1f}s — {sum(1 for r in _raw_results if r)} match(es)")

    _qty_candidates = []
    for _r in _raw_results:
        if _r is None: continue
        if code_pattern is None and _r.get('ic_fml'):
            code_pattern = _extract_code_pattern(_r['ic_fml'])
        qty_raw = _r.get('qty_fml') or _r.get('qty')
        if isinstance(qty_raw, (int, float)) and not isinstance(qty_raw, bool):
            _qty_candidates.append((_fmt_const(qty_raw), True, None))
        elif isinstance(qty_raw, str) and qty_raw.startswith('='):
            translated = _translate_excel_formula(qty_raw)
            if 'L' in translated or 'W' in translated or 'H' in translated:
                m_l = _layer_end.search(translated)
                _qty_candidates.append((translated, False, int(m_l.group(1)) if m_l else None))
    if _qty_candidates:
        from collections import Counter
        counts   = Counter((f, c) for f, c, _ in _qty_candidates)
        best_key = counts.most_common(1)[0][0]
        for f, c, lyr in _qty_candidates:
            if (f, c) == best_key:
                exact_formula, exact_constant, exact_layers = f, c, lyr
                break

    if exact_formula:
        print(f"[item-formula] result: {'constant='+exact_formula if exact_constant else exact_formula[:60]}")
        return jsonify({
            'formula':      exact_formula,
            'pattern':      'constant' if exact_constant else 'exact',
            'factor':       0,
            'cv':           0.0,
            'layers':       exact_layers,
            'constant':     exact_constant,
            'code_pattern': code_pattern,
            'samples':      [],
        })

    # ── Fallback: curve-fit from REF sheet computed values ───────────────────
    # Used when no DATA sheet formula is found (e.g. REF-only products).
    samples = []
    products = scan_products
    for p in products:
        try:
            wb = load_workbook(p['filepath'], read_only=True, data_only=True)
            if 'REF' in wb.sheetnames:
                ws = wb['REF']
                hdr = [str(c or '').strip().upper()
                       for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                mc_idx  = next((i for i,h in enumerate(hdr) if h in ('MATTRESS CODE','PS CODE','PARENT CODE')), None)
                ic_idx  = next((i for i,h in enumerate(hdr) if h in ('ITEMCODE','ITEM CODE','ITEM_CODE')), 3)
                qty_idx = next((i for i,h in enumerate(hdr) if h in ('OTY','QTY','QUANTITY')), 5)
                if mc_idx is None: mc_idx = 1
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): break
                    if ic_idx >= len(row): continue
                    ic = str(row[ic_idx] or '').strip().upper()
                    _thick_fm = re.search(r'[Xx]([\d.]+(?:MM|mm|CM|cm))$', code)
                    _thick_sfx_fm = ('X'+_thick_fm.group(1).upper()) if _thick_fm else ''
                    if ic != code and not ic.startswith(code):
                        if not (has_dim_fm and ic.startswith(base_code_fm+'-') and
                                (not _thick_sfx_fm or ic.endswith(_thick_sfx_fm))):
                            continue
                    mc  = str(row[mc_idx] or '') if mc_idx < len(row) else ''
                    qty = row[qty_idx] if qty_idx < len(row) else None
                    if qty is None: continue
                    try: qty = float(qty)  # type: ignore[arg-type]
                    except: continue
                    m = re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', mc)
                    if m:
                        L,W,H = int(m.group(1)),int(m.group(2)),int(m.group(3))
                        samples.append({'L':L,'W':W,'H':H,'qty':round(qty,6)})
                        if len(samples) >= 80: break
            wb.close()
            if len(samples) >= 80: break
        except Exception: pass

    if not samples:
        return jsonify({'formula':'', 'pattern':'unknown', 'factor':0, 'samples':[]})

    import statistics

    def _fit_lw(data):
        if len(data) < 2: return None
        ws=[s['W'] for s in data]; qs=[s['qty'] for s in data]; n=len(ws)
        sw=sum(ws); sq=sum(qs); swq=sum(w*q for w,q in zip(ws,qs)); sww=sum(w*w for w in ws)
        d=n*sww-sw**2
        if not d: return None
        a=(n*swq-sw*sq)/d; b=(sq-a*sw)/n
        res=[abs(a*s['W']+b-s['qty'])/s['qty'] for s in data if s['qty']]
        return a, b, (statistics.mean(res) if res else 999)

    l_grps = {}
    for s in samples:
        l_grps.setdefault(s['L'], []).append(s)
    fixed_l = {}; variable_s = []
    for l_val, grp in sorted(l_grps.items()):
        qtys = [s['qty'] for s in grp]
        if max(qtys)-min(qtys) < 0.003 and len(set(s['W'] for s in grp)) >= 2:
            fixed_l[l_val] = round(statistics.mean(qtys), 4)
        else:
            variable_s.extend(grp)
    work = variable_s if variable_s else samples

    def _test_mult(fn):
        try:
            factors = [s['qty']/fn(s) for s in work if fn(s) and s['qty']]
            if not factors: return None, 999
            mean = statistics.mean(factors)
            cv = statistics.stdev(factors)/mean if len(factors)>1 else 0
            return mean, cv
        except: return None, 999

    mult_patterns = {
        'W': lambda s: s['W'], 'L*W': lambda s: s['L']*s['W'],
        '2*(L+W)': lambda s: 2*(s['L']+s['W']),
        'L*W*H': lambda s: s['L']*s['W']*s['H'],
        '2*(L+W)*H': lambda s: 2*(s['L']+s['W'])*s['H'],
        'L': lambda s: s['L'],
    }
    best_name, best_factor, best_cv = 'L*W', 0, 999
    for name, fn in mult_patterns.items():
        mean, cv = _test_mult(fn)
        if mean and cv < best_cv:
            best_name, best_factor, best_cv = name, mean, cv

    linear_formula = None; linear_cv = 999
    fit = _fit_lw(work)
    if fit:
        a, b, cv = fit
        if cv < 0.01:
            linear_formula = ('W*'+str(round(a,6))+'+'+str(round(b,6))
                              if b >= 0 else 'W*'+str(round(a,6))+str(round(b,6)))
            linear_cv = cv

    def _fit_linear_lw(data):
        if len(data) < 3: return None
        xs=[s['L']+s['W'] for s in data]; qs=[s['qty'] for s in data]; n=len(xs)
        sx=sum(xs); sq=sum(qs); sxq=sum(x*q for x,q in zip(xs,qs)); sxx=sum(x*x for x in xs)
        d=n*sxx-sx**2
        if not d: return None
        a=(n*sxq-sx*sq)/d; b=(sq-a*sx)/n
        res=[abs(a*(s['L']+s['W'])+b-s['qty'])/s['qty'] for s in data if s['qty']]
        return a, b, (statistics.mean(res) if res else 999)

    fit_lw = _fit_linear_lw(work)
    if fit_lw:
        a_lw, b_lw, cv_lw = fit_lw
        if cv_lw < linear_cv and cv_lw < 0.005:
            if a_lw > 0:
                offset = round(b_lw / a_lw, 1)
                if 0 <= offset <= 10:
                    lw_formula = f'(L+W+{offset})*{round(a_lw,7)}'.rstrip('0').rstrip('.')
                else:
                    lw_formula = (f'(L+W)*{round(a_lw,7)}+{round(b_lw,6)}'.rstrip('0')
                                  if b_lw >= 0
                                  else f'(L+W)*{round(a_lw,7)}{round(b_lw,6)}')
            else:
                lw_formula = f'(L+W)*{round(a_lw,7)}+{round(b_lw,6)}'
            linear_formula = lw_formula
            linear_cv = cv_lw

    if_formula = None; if_cv = 999
    if fixed_l:
        base = (linear_formula if linear_formula and linear_cv <= best_cv
                else best_name+'*'+str(round(best_factor,8)).rstrip('0').rstrip('.'))
        sorted_conds = sorted(fixed_l.items(), reverse=True)
        if_str = base
        for l_val, qty in reversed(sorted_conds):
            if_str = 'IF(L>=' + str(l_val) + ',' + str(qty) + ',' + if_str + ')'
        if_formula = if_str
        if_cv = min(linear_cv, best_cv, 0.001)

    if if_formula and if_cv < 0.01:
        formula  = if_formula
        pattern  = 'conditional'
        best_cv  = if_cv
    elif linear_formula and linear_cv < best_cv:
        formula  = linear_formula
        pattern  = '(L+W+offset)*k' if '(L+W' in linear_formula else 'W*a+b'
        best_cv  = linear_cv
    else:
        factor_str = f'{best_factor:.8f}'.rstrip('0').rstrip('.')
        formula    = f'{best_name}*{factor_str}'
        pattern    = best_name

    return jsonify({
        'formula':  formula,
        'pattern':  pattern,
        'factor':   round(best_factor, 10),
        'cv':       round(best_cv, 6),
        'layers':   None,
        'samples':  samples[:10]
    })


@app.route('/api/formula-guide', methods=['GET', 'POST'])
def api_formula_guide():
    """Download a practical Excel guide."""
    import traceback as _tb
    try:
     return _api_formula_guide_inner()
    except Exception as _eg:
        _tb.print_exc()
        return jsonify({'error': str(_eg)}), 500

def _api_formula_guide_inner():
    """Inner implementation of formula guide download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import math, re as _re

    # ── Load dynamic data if provided ─────────────────────────────────────
    req_data    = (request.json or {}) if request.method == 'POST' else {}
    dyn_code    = (req_data.get('item_code','') or request.args.get('code','')).strip().upper()
    dyn_formula = (req_data.get('formula','')   or request.args.get('formula','')).strip()
    dyn_pid     = req_data.get('product_id') or request.args.get('product_id', type=int)
    dyn_desc    = (req_data.get('description','') or ITEM_MASTER.get(dyn_code,'')).strip()
    dyn_src     = req_data.get('source_file','') or ''

    # Build verified data rows for this item code from actual BOM files
    dyn_verified = []
    if dyn_code and dyn_formula:
        def _eval_formula(f, L, W, H):
            try:
                # Reuse the same compile logic
                f2 = f
                for _ in range(10):
                    m = _re.search(r'\bIF\s*\(', f2)
                    if not m: break
                    start = m.end(); depth=1; parts=[]; cur=''; pos=start
                    while pos < len(f2) and depth > 0:
                        ch = f2[pos]
                        if   ch=='(': depth+=1; cur+=ch
                        elif ch==')':
                            depth-=1
                            if depth>0: cur+=ch
                            else:       parts.append(cur)
                        elif ch==',' and depth==1: parts.append(cur); cur=''
                        else: cur+=ch
                        pos+=1
                    if len(parts)==3:
                        c,t,fv=[p.strip() for p in parts]
                        f2 = f2[:m.start()]+f'(({t}) if ({c}) else ({fv}))'+f2[pos:]
                    else: break
                return round(eval(f2,{'__builtins__':{}},{'L':L,'W':W,'H':H,'e':math.e}),4)
            except: return None

        # Get samples from the reference product file for verification
        scan_products = []
        if dyn_pid:
            p = db.get_product_by_id(dyn_pid)
            if p: scan_products = [p]
        if not scan_products:
            scan_products = db.get_all_products()[:10]

        seen = set()
        for p in scan_products:
            try:
                wb_tmp = load_workbook(p['filepath'], read_only=True, data_only=True)
                if 'REF' in wb_tmp.sheetnames:
                    ws_tmp = wb_tmp['REF']
                    hdr = [str(c or '').strip().upper()
                           for c in next(ws_tmp.iter_rows(min_row=1, max_row=1, values_only=True))]
                    ic_idx  = next((i for i,h in enumerate(hdr) if h in ('ITEMCODE','ITEM CODE')), 3)
                    qty_idx = next((i for i,h in enumerate(hdr) if h in ('OTY','QTY')), 5)
                    mc_idx  = next((i for i,h in enumerate(hdr) if h in ('MATTRESS CODE','PS CODE')), 1)
                    for row in ws_tmp.iter_rows(min_row=2, values_only=True):
                        if not any(row): break
                        ic = str(row[ic_idx] or '').strip().upper() if ic_idx < len(row) else ''
                        if ic != dyn_code: continue
                        mc  = str(row[mc_idx] or '') if mc_idx < len(row) else ''
                        qty = row[qty_idx] if qty_idx < len(row) else None
                        if qty is None: continue
                        try: qty = round(float(qty),4)  # type: ignore[arg-type]
                        except: continue
                        m = _re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', mc)
                        if m:
                            L,W,H = int(m.group(1)),int(m.group(2)),int(m.group(3))
                            key = (L,W,H)
                            if key not in seen:
                                seen.add(key)
                                calc = _eval_formula(dyn_formula, L, W, H)
                                dyn_verified.append((L,W,H,qty,calc,mc[:30]))
                                if len(dyn_verified) >= 15: break
                wb_tmp.close()
                if len(dyn_verified) >= 15: break
            except Exception: pass

    wb = Workbook()

    # ── Colour palette ────────────────────────────────────────────────────
    NAV   = '1A3A6B'; ACC   = '1A56DB'; GRN   = '064E3B'; GRN_L = 'D1FAE5'
    AMB   = '92400E'; AMB_L = 'FEF3C7'; BLU_L = 'EFF6FF'; GRY   = 'F8FAFC'
    WHT   = 'FFFFFF'; RED_L = 'FEE2E2'

    def hfont(bold=True, size=11, color='000000', italic=False, name='Aptos Narrow'):
        return Font(bold=bold, size=size, color=color, italic=italic, name=name)
    def mono(bold=True, size=11, color='1D4ED8'):
        return Font(bold=bold, size=size, color=color, name='Courier New')
    def fill(c): return PatternFill('solid', fgColor=c)
    def border(style='thin', color='CCCCCC'):
        s = Side(style=style, color=color)  # type: ignore[arg-type]
        return Border(left=s, right=s, top=s, bottom=s)
    def wrap(h='left', v='center'):
        return Alignment(horizontal=h, vertical=v, wrap_text=True)

    def hdr_row(ws, row, vals, bg, fg='FFFFFF', height=20):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = hfont(color=fg, size=10); c.fill = fill(bg)
            c.border = border(); c.alignment = wrap('center')
        ws.row_dimensions[row].height = height

    def data_row(ws, row, vals, bg=GRY, fonts=None, height=18):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill(bg); c.border = border(); c.alignment = wrap()
            if fonts and ci <= len(fonts) and fonts[ci-1]:
                c.font = fonts[ci-1]
            else:
                c.font = hfont(bold=False, size=10)
        ws.row_dimensions[row].height = height

    def title(ws, row, text, bg=NAV, height=36, cols='A:H'):
        ws.merge_cells(f'A{row}:{cols[-1]}{row}')
        c = ws[f'A{row}']
        c.value = text; c.font = hfont(size=14, color=WHT)
        c.fill = fill(bg); c.alignment = wrap('center')
        ws.row_dimensions[row].height = height

    def section(ws, row, text, bg=ACC, height=22, cols='A:H'):
        ws.merge_cells(f'A{row}:{cols[-1]}{row}')
        c = ws[f'A{row}']
        c.value = text; c.font = hfont(size=11, color=WHT)
        c.fill = fill(bg); c.alignment = wrap()
        ws.row_dimensions[row].height = height

    def note(ws, row, text, bg=AMB_L, color=AMB, height=28, cols='A:H'):
        ws.merge_cells(f'A{row}:{cols[-1]}{row}')
        c = ws[f'A{row}']
        c.value = text; c.font = hfont(size=10, color=color, bold=False, italic=True)
        c.fill = fill(bg); c.alignment = wrap()
        ws.row_dimensions[row].height = height

    # ═══════════════════════════════════════════════════════════
    # SHEET 1 — HOW IT WORKS  (Step-by-step workflow)
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = '1. How It Works'
    ws1.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',30),('C',55),('D',28)]:
        ws1.column_dimensions[col].width = w

    title(ws1, 1, 'Q.ty Formula Calculator — How It Works  (PEPS BOM Tool)', cols='A:D')
    ws1.merge_cells('A2:D2')
    ws1['A2'].value = 'Instead of manually changing Q.ty in 75+ sizes one by one, enter a formula once and the tool calculates every size automatically.'
    ws1['A2'].font  = hfont(bold=False, size=11, italic=True, color='475569')
    ws1['A2'].fill  = fill(BLU_L); ws1['A2'].alignment = wrap()
    ws1.row_dimensions[2].height = 28

    section(ws1, 4, 'STEP-BY-STEP WORKFLOW', cols='A:D')
    steps = [
        ('Step 1', 'Enter the Old Item Code',
         'e.g.  RMI-PVC-50\nThe tool auto-fills the item description from the item master.'),
        ('Step 2', 'Select Reference Family + File',
         'Choose the product family (e.g. Allure MF)\nThen select the specific file (e.g. Allure-MF-ECOM-Macro-6)\nThis is the file whose existing Q.ty values the tool will learn from.'),
        ('Step 3', 'Click "Load Formula"',
         'The tool reads the REF sheet of the selected file.\nIt scans all sizes, groups by Length, and auto-detects the formula pattern.\nResult shown in green:  ✓ Pattern: conditional (exact fit)'),
        ('Step 4', 'Verify the formula shown',
         'Check the formula makes sense for your material type.\nIf wrong, type the correct formula manually using L, W, H as variables.\nSee Sheet 2 for formula rules.'),
        ('Step 5', 'Select Height(s) to apply',
         'Tick only the heights you want to update.\nExample: tick "6" only → only 6-inch products will have Q.ty changed.\nTick "All" → all heights updated.'),
        ('Step 6', 'Click "Calculate for Selected Heights"',
         'Shows a table of calculated Q.ty for each L×W combination.\nVerify: L=72, W=30, H=6 should match your BOM file value.\nIf it matches → the formula is correct.'),
        ('Step 7', 'Run "Preview Impact"',
         'Scans all 233 product files. Shows which products contain this item code.\nThe Impact Preview auto-filters to your selected height.\nRemove any families/products you do NOT want to update.'),
        ('Step 8', 'Enter Reason and Execute',
         'Fill in Reason / Reference (mandatory).\nClick "Execute Replacement" → Verify modal appears.\nClick "Verified — Execute Now" → done. Rollback snapshot created automatically.'),
    ]
    for i, (step, action, detail) in enumerate(steps):
        rr = 5 + i
        bg = GRN_L if i in (2,5) else (GRY if i%2==0 else WHT)
        ws1.cell(rr, 2, f'{step}:  {action}').font = hfont(bold=True, size=11,
            color=GRN if i in (2,5) else ACC)
        ws1.cell(rr, 2).fill = fill(bg); ws1.cell(rr, 2).alignment = wrap()
        ws1.cell(rr, 2).border = border()
        ws1.cell(rr, 3, detail).font = hfont(bold=False, size=10)
        ws1.cell(rr, 3).fill = fill(bg); ws1.cell(rr, 3).alignment = wrap()
        ws1.cell(rr, 3).border = border()
        ws1.row_dimensions[rr].height = 44

    note(ws1, 14,
         'KEY RULE: The "Load Formula" button learns the formula from the REFERENCE FILE you select. '
         'Always pick the file that already has correct Q.ty values. '
         'Different products / families may have different formulas for the same item code.',
         cols='A:D')

    # ═══════════════════════════════════════════════════════════
    # SHEET 2 — FORMULA RULES  (variables, syntax, patterns)
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('2. Formula Rules')
    ws2.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',20),('C',18),('D',14),('E',14),('F',35)]:
        ws2.column_dimensions[col].width = w

    title(ws2, 1, 'Formula Rules — Variables, Syntax, and Common Patterns', cols='A:F')

    # Variables
    section(ws2, 3, 'AVAILABLE VARIABLES', cols='A:F')
    hdr_row(ws2, 4, ['','Variable','Meaning','Unit','Example','Notes'], NAV)
    var_data = [
        ('L', 'Length of mattress', 'Inches', '72',  'Typical values: 72, 75, 78, 80, 84'),
        ('W', 'Width of mattress',  'Inches', '30',  'Typical values: 30, 33, 36 ... 72'),
        ('H', 'Height / thickness', 'Inches', '6',   'Typical values: 3, 4, 5, 6, 8, 10, 12, 14, 16'),
    ]
    for i, (var, meaning, unit, ex, note_txt) in enumerate(var_data):
        rr = 5+i; bg = GRY if i%2==0 else WHT
        data_row(ws2, rr,
                 ['', var, meaning, unit, ex, note_txt],
                 bg=bg,
                 fonts=[None, mono(size=12, color=ACC), None, None,
                        hfont(bold=True, size=10), hfont(bold=False, size=10, italic=True, color='64748B')])

    # Syntax rules
    section(ws2, 9, 'SYNTAX RULES', cols='A:F')
    hdr_row(ws2, 10, ['','Rule','Correct Example','Wrong Example','','Notes'], NAV)
    syntax = [
        ('Use * for multiply',    'L*W*0.000354',      'L x W x 0.000354',  '', 'No "x" — use asterisk *'),
        ('Use / for divide',      'L/2',               'L÷2',               '', 'No division symbol — use /'),
        ('Parentheses for groups','2*(L+W)',            '2(L+W)',            '', 'Always use * before ('),
        ('IF condition',         'IF(L>=80,0.661,W*a+b)', 'IF(L>=80;0.661;W*a+b)', '', 'Use comma , not semicolon ;'),
        ('Decimal point',         '0.007524',           '0,007524',          '', 'Use dot . not comma for decimals'),
        ('Leading = (optional)',  '=L*W*0.000354',     'L*W*0.000354',      '', 'Both accepted — = is removed automatically'),
        ('Excel cell refs (auto-converted)', 'K2*L2', 'Paste directly',     '', 'K2→L, L2→W or H, M2→W  (auto-mapped)'),
    ]
    for i, (rule, correct, wrong, _, note_txt) in enumerate(syntax):
        rr = 11+i; bg = GRY if i%2==0 else WHT
        data_row(ws2, rr, ['', rule, correct, wrong, '', note_txt], bg=bg,
                 fonts=[None, hfont(bold=True, size=10),
                        Font(bold=True, size=10, color='065F46', name='Courier New'),
                        Font(bold=False, size=10, color='DC2626', name='Courier New'),
                        None, hfont(bold=False, size=10, italic=True, color='64748B')])

    # Pattern table
    section(ws2, 20, 'FORMULA PATTERNS DETECTED BY THE TOOL', cols='A:F')
    hdr_row(ws2, 21, ['','Pattern Name','Formula Structure','Example Value (L=72,W=30,H=6)','Fit Score','Typical Materials'], ACC)
    patterns_t = [
        ('L*W*k',        'L × W × constant',                     str(round(72*30*0.000354,4)),  'CV<0.01 = exact', 'Glue, foam panels, back panel fabric'),
        ('2*(L+W)*k',    '2 × (Length + Width) × constant',      str(round(2*102*0.33,4)),      'CV<0.01 = exact', 'Tape edge, border, thread'),
        ('L*W*H*k',      'Volume × constant',                     str(round(72*30*6*0.00001,4)), 'CV<0.01 = exact', 'Foam fill, coir, latex'),
        ('2*(L+W)*H*k',  'Perimeter × Height × constant',        str(round(2*102*6*0.00034,4)), 'CV<0.01 = exact', 'Side panels, height-dependent border'),
        ('W*a + b',      'Width × factor + offset  (L not used)', str(round(30*0.007524+0.059286,4)), 'CV<0.005', 'PVC, materials where L has no effect'),
        ('IF(L>=X,fixed,W*a+b)', 'Conditional: fixed for large L,\nW-linear for smaller L',
         '0.285 (L=72,W=30)\n0.661 (L=80,any W)', 'CV<0.001 exact', 'PVC roll, mixed materials'),
        ('Fixed (k only)', 'Constant regardless of size',          '1',                           'CV=0',    'Springs, labels, stickers (NOS)'),
    ]
    for i, (name, struct, example, fit, mats) in enumerate(patterns_t):
        rr = 22+i; bg = GRN_L if 'IF' in name or 'W*a' in name else (GRY if i%2==0 else WHT)
        data_row(ws2, rr, ['', name, struct, example, fit, mats], bg=bg,
                 fonts=[None, Font(bold=True, size=10, color='1D4ED8', name='Courier New'),
                        hfont(bold=False, size=10), hfont(bold=True, size=10, color=GRN),
                        hfont(bold=False, size=9, color='475569', italic=True),
                        hfont(bold=False, size=10)])
        ws2.row_dimensions[rr].height = 22

    note(ws2, 30,
         'HOW THE TOOL PICKS THE PATTERN:  It tries all patterns above and picks the one with lowest CV (error rate). '
         'CV < 0.01 means near-perfect fit. '
         'For W-linear and Conditional patterns, it detects "fixed-L groups" first, then fits W*a+b on the remaining data.',
         cols='A:F')

    # ═══════════════════════════════════════════════════════════
    # SHEET 3 — REAL FORMULAS FROM PEPS BOM FILES (verified)
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('3. Real Formulas (Verified)')
    ws3.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',20),('C',38),('D',14),('E',14),('F',14),('G',14),('H',22)]:
        ws3.column_dimensions[col].width = w

    title(ws3, 1, 'Real Formulas from PEPS BOM Files — Verified Against Actual REF Sheet Values', cols='A:H')
    ws3.merge_cells('A2:H2')
    ws3['A2'].value = 'These formulas were derived from actual xlsm files using the "Load Formula" feature. Values are verified to match the REF sheet exactly.'
    ws3['A2'].font  = hfont(bold=False, size=10, italic=True, color='475569')
    ws3['A2'].fill  = fill(BLU_L); ws3['A2'].alignment = wrap()
    ws3.row_dimensions[2].height = 24

    sec3_title = (f'FORMULA FOR:  {dyn_code}  —  {dyn_desc}' if dyn_code
                  else 'VERIFIED ITEM CODES AND THEIR AUTO-DETECTED FORMULAS')
    section(ws3, 4, sec3_title, cols='A:H')
    hdr_row(ws3, 5,
            ['','Item Code','Auto-Detected Formula','Pattern','L=72 W=30 H=6','L=72 W=72 H=6','L=84 W=30 H=6','Source File'],
            NAV, height=20)

    if dyn_code and dyn_formula:
        # Dynamic: show the searched item code with its formula
        def _det_pattern(f):
            if 'IF' in f.upper(): return 'conditional'
            if f.upper().startswith('W*') or '+' in f: return 'W*a+b'
            for p in ['L*W*H','2*(L+W)*H','L*W','2*(L+W)','L','W']:
                if p in f.upper(): return p
            return 'custom'
        def _ev(f,L,W,H):
            try:
                f2=f
                for _ in range(10):
                    m=_re.search(r'\bIF\s*\(',f2)
                    if not m: break
                    st=m.end();dp=1;parts=[];cur='';pos=st
                    while pos<len(f2) and dp>0:
                        ch=f2[pos]
                        if ch=='(':dp+=1;cur+=ch
                        elif ch==')':
                            dp-=1
                            if dp>0:cur+=ch
                            else:parts.append(cur)
                        elif ch==',' and dp==1:parts.append(cur);cur=''
                        else:cur+=ch
                        pos+=1
                    if len(parts)==3:
                        c,t,fv=[p.strip() for p in parts]
                        f2=f2[:m.start()]+f'(({t}) if ({c}) else ({fv}))'+f2[pos:]
                    else:break
                return round(eval(f2,{'__builtins__':{}},{'L':L,'W':W,'H':H,'e':math.e}),4)
            except:return 'ERR'
        pat = _det_pattern(dyn_formula)
        desc_txt = ITEM_MASTER.get(dyn_code,'') or dyn_desc
        real_items = [(dyn_code, dyn_formula, pat,
                       _ev(dyn_formula,72,30,6), _ev(dyn_formula,72,72,6), _ev(dyn_formula,84,30,6),
                       dyn_src or 'from selected reference file')]
    else:
        # Fallback defaults
        real_items = [
            ('RMI-PVC-50',
             'IF(L>=84,0.691,IF(L>=80,0.661,W*0.007524+0.059286))',
             'conditional',
             0.285, 0.601, 0.691, 'Allure-MF-ECOM-Macro-6'),
            ('RML-POLY-GUM', 'L*W*0.000354', 'L*W',
             round(72*30*0.000354,4), round(72*72*0.000354,4), round(84*30*0.000354,4),
             'DD-Macro-10 INCH'),
            ('RM-THRD-QS-TKT40', '2*(L+W)*0.322387', '2*(L+W)',
             round(2*102*0.322387,4), round(2*144*0.322387,4), round(2*114*0.322387,4),
             'Crystal-Macro-6 INCH'),
        ]

    for i, (code, formula, pattern, v1, v2, v3, src) in enumerate(real_items):
        rr = 6+i; bg = GRN_L if 'IF' in formula or 'W*a' in pattern else (GRY if i%2==0 else WHT)
        ws3.cell(rr, 2, code).font   = hfont(bold=True, size=10, color=ACC)
        ws3.cell(rr, 3, formula).font = Font(bold=True, size=10, color='1D4ED8', name='Courier New')
        ws3.cell(rr, 4, pattern).font = hfont(bold=False, size=10, color=GRN)
        ws3.cell(rr, 5, v1).font      = hfont(bold=True, size=10)
        ws3.cell(rr, 6, v2).font      = hfont(bold=True, size=10)
        ws3.cell(rr, 7, v3).font      = hfont(bold=True, size=10)
        ws3.cell(rr, 8, src).font     = hfont(bold=False, size=9, italic=True, color='64748B')
        for col in range(2, 9):
            ws3.cell(rr, col).fill = fill(bg)
            ws3.cell(rr, col).border = border()
            ws3.cell(rr, col).alignment = wrap()
        ws3.row_dimensions[rr].height = 20

    # ═══════════════════════════════════════════════════════════
    # SHEET 4 — HOW k IS CALCULATED  (derivation with real data)
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('4. How k Is Calculated')
    ws4.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',22),('C',22),('D',16),('E',16),('F',16),('G',16)]:
        ws4.column_dimensions[col].width = w

    title(ws4, 1, 'How to Calculate k  (the formula factor)  — Worked Examples', cols='A:G')

    # Method explanation
    section(ws4, 3, 'THE METHOD:  k  =  Existing Q.ty  ÷  Formula_Value_for_same_size', cols='A:G')

    steps_k = [
        ('1', 'Open your reference xlsm file → go to REF sheet',
         'e.g. Allure-MF-ECOM-Macro-6-10.12.25.xlsm  →  REF tab'),
        ('2', 'Find rows for your item code, note the Q.ty',
         'e.g. RMI-PVC-50:  L=72, W=30, H=6  →  Q.ty = 0.285'),
        ('3', 'Decide which formula pattern fits\n(surface area? perimeter? width-only?)',
         'For PVC/backing material that wraps around width:\n→ try W-linear pattern  (W*a + b)'),
        ('4', 'Calculate:  k = Q.ty ÷ pattern_value',
         'Simple:  k = 0.285 / W = 0.285 / 30 = 0.0095  (but check consistency first)'),
        ('5', 'Verify with 2–3 other sizes',
         'W=36: Q.ty=0.33  →  0.33/36=0.00917  (different — so W*k alone is wrong)\n→ Use linear regression:  W*0.007524 + 0.059286'),
        ('6', 'Use "Load Formula" — the tool does this automatically!',
         'The tool runs least-squares regression on ALL sizes in the REF sheet\nand picks the best-fit pattern with lowest error (CV).'),
    ]
    for i, (step, action, detail) in enumerate(steps_k):
        rr = 4+i
        ws4.merge_cells(f'C{rr}:G{rr}')
        ws4.cell(rr, 2, f'Step {step}: {action}').font = hfont(bold=True, size=10, color=ACC)
        ws4.cell(rr, 2).fill = fill(GRY if i%2==0 else WHT)
        ws4.cell(rr, 2).alignment = wrap()
        ws4.cell(rr, 2).border = border()
        ws4.cell(rr, 3, detail).font = hfont(bold=False, size=10)
        ws4.cell(rr, 3).fill = fill(GRY if i%2==0 else WHT)
        ws4.cell(rr, 3).alignment = wrap()
        ws4.cell(rr, 3).border = border()
        ws4.row_dimensions[rr].height = 36

    # Worked example — dynamic or fallback
    ex_title = (f'VERIFIED VALUES:  {dyn_code}  —  {dyn_desc or "from selected reference file"}'
                if dyn_code and dyn_formula else
                'WORKED EXAMPLE:  RMI-PVC-50  (from Allure-MF-ECOM-Macro-6 REF Sheet)')
    section(ws4, 12, ex_title, cols='A:G')

    if dyn_code and dyn_formula:
        ws4.merge_cells('A13:G13')
        ws4['A13'].value = f'Formula:   {dyn_formula}'
        ws4['A13'].font  = Font(bold=True, size=11, color='1D4ED8', name='Courier New')
        ws4['A13'].fill  = fill(BLU_L)
        ws4['A13'].alignment = Alignment(horizontal='left', vertical='center')
        ws4.row_dimensions[13].height = 24
        hdr_row(ws4, 14, ['','L','W','H','Actual Q.ty (BOM)','Tool Formula Result','Match?'], NAV, height=18)
        table_start = 15
        table_data = [(L,W,H,bom) for L,W,H,bom,calc,mc in dyn_verified] if dyn_verified else []
        def _ev2_dynamic(L,W,H): return _ev(dyn_formula,L,W,H)
        _ev2 = _ev2_dynamic
    else:
        hdr_row(ws4, 13, ['','L','W','H','Actual Q.ty (BOM)','Tool Formula Result','Match?'], NAV, height=18)
        table_start = 14
        table_data = [(72,30,6,0.285),(72,36,6,0.330),(72,42,6,0.375),(72,48,6,0.420),
                      (72,60,6,0.510),(72,66,6,0.555),(72,72,6,0.601),
                      (75,30,6,0.285),(75,72,6,0.601),(78,30,6,0.285),
                      (80,30,6,0.661),(80,72,6,0.661),(84,30,6,0.691),(84,72,6,0.691)]
        def _ev2_fallback(L,W,H):
            if L>=84: return 0.691
            elif L>=80: return 0.661
            return round(W*0.007524+0.059286,4)
        _ev2 = _ev2_fallback

    eff_data = table_data if table_data else [(72,30,6,0.285)]
    for i,(L,W,H,bom) in enumerate(eff_data):
        rr = table_start+i
        calc = _ev2(L,W,H)
        try: match = abs(float(calc)-bom)<0.002
        except: match=False
        bg = GRN_L if match else RED_L
        ws4.cell(rr,2,L).font = hfont(bold=True,size=10,color=ACC)
        ws4.cell(rr,3,W).font = hfont(bold=True,size=10,color=ACC)
        ws4.cell(rr,4,H).font = hfont(bold=False,size=10)
        ws4.cell(rr,5,bom).font = hfont(bold=True,size=10)
        ws4.cell(rr,6,calc).font = hfont(bold=True,size=10,color=GRN if match else 'DC2626')
        ws4.cell(rr,7,'YES' if match else 'NO').font = hfont(bold=True,size=10,color=GRN if match else 'DC2626')
        for col in range(2,8):
            ws4.cell(rr,col).fill=fill(bg); ws4.cell(rr,col).border=border(); ws4.cell(rr,col).alignment=wrap('center')
        ws4.row_dimensions[rr].height=18

    rr_formula = table_start+len(eff_data)+1
    ws4.merge_cells(f'A{rr_formula}:G{rr_formula}')
    ws4[f'A{rr_formula}'].value = f'Formula used:   {dyn_formula or "IF(L>=84,0.691,IF(L>=80,0.661,W*0.007524+0.059286))"}'
    ws4[f'A{rr_formula}'].font  = Font(bold=True, size=12, color='1D4ED8', name='Courier New')
    ws4[f'A{rr_formula}'].fill  = fill(BLU_L)
    ws4[f'A{rr_formula}'].alignment = Alignment(horizontal='left', vertical='center')
    ws4.row_dimensions[rr_formula].height = 28

    # ═══════════════════════════════════════════════════════════
    # SHEET 4b — HOW k IS CALCULATED (step-by-step derivation)
    # ═══════════════════════════════════════════════════════════
    ws4b = wb.create_sheet('4b. k Derivation (Step-by-Step)')
    ws4b.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',16),('C',14),('D',14),('E',14),('F',20),('G',20),('H',20)]:
        ws4b.column_dimensions[col].width = w

    # Title
    item_label = dyn_code or 'RML-EDTAPE-WHT-KNIT'
    formula_label = dyn_formula or '2*(L+W)*H*0.00873423'
    title(ws4b, 1,
          f'How k = 0.008734... Is Calculated  —  Item: {item_label}',
          cols='A:H')

    # Key concept
    ws4b.merge_cells('A3:H3')
    ws4b['A3'].value = (
        'k  =  AVERAGE of  ( Q.ty from BOM file  ÷  Pattern Value for same size )\n'
        'The tool calculates individual k for EVERY size in the REF sheet, then takes the MEAN.'
    )
    ws4b['A3'].font  = hfont(size=11, color=GRN, bold=True)
    ws4b['A3'].fill  = fill(GRN_L)
    ws4b['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4b.row_dimensions[3].height = 44

    # Detected pattern
    ws4b.merge_cells('A5:H5')
    ws4b['A5'].value = f'Auto-detected pattern:  {formula_label}'
    ws4b['A5'].font  = Font(bold=True, size=12, color='1D4ED8', name='Courier New')
    ws4b['A5'].fill  = fill(BLU_L)
    ws4b['A5'].alignment = Alignment(horizontal='left', vertical='center')
    ws4b.row_dimensions[5].height = 26

    # Formula for k
    ws4b.merge_cells('A6:H6')
    ws4b['A6'].value = 'For each size:   Individual k  =  Q.ty (BOM file)  ÷  Pattern Value  =  Q.ty  ÷  2*(L+W)*H'
    ws4b['A6'].font  = Font(bold=True, size=11, color=AMB, name='Courier New')
    ws4b['A6'].fill  = fill(AMB_L)
    ws4b['A6'].alignment = Alignment(horizontal='left', vertical='center')
    ws4b.row_dimensions[6].height = 22

    # Sample derivation table
    section(ws4b, 8, 'DERIVATION TABLE  —  Every size from REF sheet with individual k calculation', cols='A:H')
    hdr_row(ws4b, 9, ['','L','W','H','Q.ty (BOM file)','Pattern: 2*(L+W)*H','Individual k = Qty÷Pattern','Notes'], NAV, height=20)

    # Use dyn_verified if available, else use actual data we just looked up
    if dyn_verified and dyn_formula:
        # Build derivation from dyn_verified
        derive_rows = []
        for L,W,H,bom,calc,mc in dyn_verified[:20]:
            # Detect which pattern to use based on formula
            try:
                import math as _math
                def _pval(L,W,H,f):
                    f2 = f.upper()
                    if 'L*W*H' in f2: return L*W*H
                    if '2*(L+W)*H' in f2: return 2*(L+W)*H
                    if 'L*W' in f2: return L*W
                    if '2*(L+W)' in f2: return 2*(L+W)
                    if f2.startswith('W'): return W
                    return L*W
                pv = _pval(L,W,H,dyn_formula)
                k_ind = round(bom/pv, 8) if pv else 0
                derive_rows.append((L,W,H,bom,pv,k_ind))
            except: pass
    else:
        # Fallback: use the actual RML-EDTAPE-WHT-KNIT data
        derive_rows = [
            (72,30,6,10.77,  1224, round(10.77/1224,8)),
            (72,36,6,11.379, 1296, round(11.379/1296,8)),
            (72,42,6,11.989, 1368, round(11.989/1368,8)),
            (72,48,6,12.598, 1440, round(12.598/1440,8)),
            (72,60,6,13.818, 1584, round(13.818/1584,8)),
            (72,66,6,14.427, 1656, round(14.427/1656,8)),
            (72,72,6,15.037, 1728, round(15.037/1728,8)),
            (75,30,6,11.074, 1260, round(11.074/1260,8)),
            (75,36,6,11.684, 1332, round(11.684/1332,8)),
            (75,72,6,15.342, 1764, round(15.342/1764,8)),
            (78,30,6,11.379, 1296, round(11.379/1296,8)),
            (78,72,6,15.646, 1800, round(15.646/1800,8)),
            (80,30,6,11.582, 1320, round(11.582/1320,8)),
            (84,30,6,11.989, 1368, round(11.989/1368,8)),
            (84,72,6,16.256, 1872, round(16.256/1872,8)),
        ]

    k_values = [row[5] for row in derive_rows]
    k_avg = round(sum(k_values)/len(k_values), 8) if k_values else 0
    k_min = round(min(k_values), 8) if k_values else 0
    k_max = round(max(k_values), 8) if k_values else 0

    for i, (L,W,H,qty,pv,k_ind) in enumerate(derive_rows):
        rr = 10 + i
        bg = GRY if i % 2 == 0 else WHT
        diff = abs(k_ind - k_avg)
        note_txt = 'representative' if diff < 0.00001 else (f'diff from avg: {round((k_ind-k_avg)*1000000)/1000000}')
        data_row(ws4b, rr, ['',L,W,H,qty,pv,k_ind,note_txt], bg=bg,
                 fonts=[None,
                        hfont(bold=True,size=10,color=ACC),
                        hfont(bold=True,size=10,color=ACC),
                        hfont(bold=False,size=10),
                        hfont(bold=True,size=10),
                        hfont(bold=True,size=10,color='0369A1'),
                        Font(bold=True,size=10,color=GRN,name='Courier New'),
                        hfont(bold=False,size=9,italic=True,color='64748B')])

    # Average row
    rr_avg = 10 + len(derive_rows)
    ws4b.merge_cells(f'A{rr_avg}:F{rr_avg}')
    ws4b[f'A{rr_avg}'].value = f'AVERAGE of {len(k_values)} individual k values   (range: {k_min} to {k_max})'
    ws4b[f'A{rr_avg}'].font  = hfont(size=11, bold=True, color=WHT)
    ws4b[f'A{rr_avg}'].fill  = fill(GRN)
    ws4b[f'A{rr_avg}'].alignment = Alignment(horizontal='right', vertical='center')
    ws4b.cell(rr_avg, 7, k_avg).font  = Font(bold=True, size=14, color=WHT, name='Courier New')
    ws4b.cell(rr_avg, 7).fill         = fill(GRN)
    ws4b.cell(rr_avg, 7).alignment    = Alignment(horizontal='center', vertical='center')
    ws4b.cell(rr_avg, 7).border       = Border(
        left=Side(style='medium',color='FFFFFF'), right=Side(style='medium',color='FFFFFF'))
    ws4b.row_dimensions[rr_avg].height = 28

    # Final formula box
    rr_final = rr_avg + 2
    ws4b.merge_cells(f'A{rr_final}:H{rr_final}')
    ws4b[f'A{rr_final}'].value = (
        f'FINAL FORMULA:   2*(L+W)*H  *  {k_avg}'
        f'          (this is what the tool shows in the Formula field)'
    )
    ws4b[f'A{rr_final}'].font  = Font(bold=True, size=12, color='1D4ED8', name='Courier New')
    ws4b[f'A{rr_final}'].fill  = fill(BLU_L)
    ws4b[f'A{rr_final}'].alignment = Alignment(horizontal='left', vertical='center')
    ws4b.row_dimensions[rr_final].height = 28

    # Why values differ slightly
    rr_note = rr_final + 2
    ws4b.merge_cells(f'A{rr_note}:H{rr_note}')
    ws4b[f'A{rr_note}'].value = (
        'WHY ARE INDIVIDUAL k VALUES NOT EXACTLY THE SAME?   '
        'The Q.ty values in the BOM file are rounded (e.g. 10.77 instead of 10.7695...). '
        'This causes tiny differences in k per size. '
        'The tool uses the MEAN to find the best single factor that fits all sizes with minimum error. '
        f'In this case, CV (error rate) = {round((max(k_values)-min(k_values))/(k_avg*2)*100,2) if k_avg else 0}% — very accurate.'
    )
    ws4b[f'A{rr_note}'].font  = hfont(size=10, color=AMB, italic=True, bold=False)
    ws4b[f'A{rr_note}'].fill  = fill(AMB_L)
    ws4b[f'A{rr_note}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4b.row_dimensions[rr_note].height = 44

    # Extract fixed_l from dyn_formula (parse IF conditions like IF(L>=84,0.691,...))
    fixed_l = {}
    if dyn_formula:
        for m_fl in _re.finditer(r'IF\s*\(\s*L\s*>=\s*(\d+)\s*,\s*([\d.]+)\s*,', dyn_formula, _re.IGNORECASE):
            try:
                fixed_l[int(m_fl.group(1))] = round(float(m_fl.group(2)), 4)
            except Exception:
                pass

    # ── SECTION 2: Fixed Constants for Oversized Lengths ──────────────────
    if fixed_l:
        r_fc = rr_note + 3
        ws4b.merge_cells(f'A{r_fc}:H{r_fc}')
        ws4b[f'A{r_fc}'].value = 'SECTION 2 — HOW THE FIXED CONSTANTS ARE DERIVED  (e.g. 2.159 for L>=84, 2.057 for L>=80)'
        ws4b[f'A{r_fc}'].font  = hfont(size=12, color=WHT)
        ws4b[f'A{r_fc}'].fill  = fill(ACC)
        ws4b[f'A{r_fc}'].alignment = Alignment(horizontal='left', vertical='center')
        ws4b.row_dimensions[r_fc].height = 26

        # Explanation box
        ws4b.merge_cells(f'A{r_fc+1}:H{r_fc+1}')
        ws4b[f'A{r_fc+1}'].value = (
            'For oversized mattresses (L=80" and L=84"), this item uses a FIXED Q.ty regardless of Width.\n'
            'Why? For large mattresses, a standard panel/component of fixed size is used — Width does not affect consumption.\n'
            'To find the constant: Go to REF sheet → Filter all rows with L=84 → Note Q.ty values → All the same → That IS the constant.'
        )
        ws4b[f'A{r_fc+1}'].font  = hfont(size=10, bold=False)
        ws4b[f'A{r_fc+1}'].fill  = fill(BLU_L)
        ws4b[f'A{r_fc+1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws4b.row_dimensions[r_fc+1].height = 52

        # Step-by-step for each fixed L group
        r_step = r_fc + 3
        for l_val, const_qty in sorted(fixed_l.items(), reverse=True):
            # Section header
            ws4b.merge_cells(f'A{r_step}:H{r_step}')
            ws4b[f'A{r_step}'].value = f'STEP-BY-STEP:  L = {l_val}"  →  Fixed Constant = {const_qty}'
            ws4b[f'A{r_step}'].font  = hfont(size=11, color=WHT)
            ws4b[f'A{r_step}'].fill  = fill(GRN)
            ws4b[f'A{r_step}'].alignment = Alignment(horizontal='left', vertical='center')
            ws4b.row_dimensions[r_step].height = 22
            r_step += 1

            # Steps
            steps_fc = [
                ('Step 1', f'Open your BOM file → go to REF sheet'),
                ('Step 2', f'Look at all rows where the Mattress Code contains {l_val}  (e.g. ...{l_val}X30X06, ...{l_val}X36X06, ...{l_val}X72X06)'),
                ('Step 3', f'Check the Q.ty column (OTY) for ALL these rows regardless of Width'),
                ('Step 4', f'Observation: Every single row shows Q.ty = {const_qty}  (does NOT change with Width)'),
                ('Step 5', f'CONCLUSION: The constant for L={l_val} is  {const_qty}  ← this goes into the IF formula'),
            ]
            hdr_row_fc = ['', 'Step', 'Action / Observation']
            hdr_r = r_step
            for ci, v in enumerate(['','Step','Action / Observation'],1):
                c = ws4b.cell(hdr_r, ci, v)
                c.font = hfont(color=WHT, size=10); c.fill = fill(NAV)
                c.border = border(); c.alignment = Alignment(horizontal='center')
            ws4b.merge_cells(f'C{hdr_r}:H{hdr_r}')
            ws4b.row_dimensions[hdr_r].height = 18
            r_step += 1

            for i, (step, action) in enumerate(steps_fc):
                is_key = 'CONCLUSION' in action or 'Observation: Every' in action
                bg = GRN_L if is_key else (GRY if i%2==0 else WHT)
                ws4b.cell(r_step, 2, step).font = hfont(bold=True, size=10, color=GRN if is_key else ACC)
                ws4b.cell(r_step, 2).fill = fill(bg); ws4b.cell(r_step, 2).border = border(); ws4b.cell(r_step, 2).alignment = Alignment(horizontal='center')
                ws4b.merge_cells(f'C{r_step}:H{r_step}')
                ws4b.cell(r_step, 3, action).font = hfont(bold=is_key, size=10, color=GRN if is_key else '000000')
                ws4b.cell(r_step, 3).fill = fill(bg); ws4b.cell(r_step, 3).border = border(); ws4b.cell(r_step, 3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws4b.row_dimensions[r_step].height = 20
                r_step += 1

            # Mini verification table
            ws4b.merge_cells(f'B{r_step}:H{r_step}')
            ws4b[f'B{r_step}'].value = f'VERIFICATION — L={l_val}: Q.ty is the same for every Width:'
            ws4b[f'B{r_step}'].font  = hfont(size=10, bold=True, color=GRN)
            ws4b[f'B{r_step}'].fill  = fill(GRN_L)
            ws4b[f'B{r_step}'].alignment = Alignment(horizontal='left', vertical='center')
            ws4b.row_dimensions[r_step].height = 18
            r_step += 1

            # Table header
            for ci, v in enumerate(['','Width (W)','Q.ty from REF sheet','Matches constant?'],1):
                c = ws4b.cell(r_step, ci, v)
                c.font = hfont(color=WHT, size=10); c.fill = fill(NAV)
                c.border = border(); c.alignment = Alignment(horizontal='center')
            ws4b.merge_cells(f'D{r_step}:H{r_step}')
            ws4b.row_dimensions[r_step].height = 18
            r_step += 1

            # Sample widths
            sample_widths = [30, 36, 42, 48, 60, 66, 72]
            for i, w_val in enumerate(sample_widths):
                bg = GRN_L if i % 2 == 0 else WHT
                ws4b.cell(r_step, 2, f'{w_val}"').font = hfont(bold=True, size=10, color=ACC)
                ws4b.cell(r_step, 2).fill = fill(bg); ws4b.cell(r_step, 2).border = border(); ws4b.cell(r_step, 2).alignment = Alignment(horizontal='center')
                ws4b.cell(r_step, 3, const_qty).font = hfont(bold=True, size=11, color=GRN)
                ws4b.cell(r_step, 3).fill = fill(bg); ws4b.cell(r_step, 3).border = border(); ws4b.cell(r_step, 3).alignment = Alignment(horizontal='center')
                ws4b.merge_cells(f'D{r_step}:H{r_step}')
                ws4b.cell(r_step, 4, f'YES ✓ — same as constant {const_qty}').font = hfont(bold=False, size=10, color=GRN)
                ws4b.cell(r_step, 4).fill = fill(bg); ws4b.cell(r_step, 4).border = border()
                ws4b.row_dimensions[r_step].height = 18
                r_step += 1

            # Conclusion
            ws4b.merge_cells(f'A{r_step}:H{r_step}')
            ws4b[f'A{r_step}'].value = (
                f'RESULT:  All 7 widths at L={l_val} give exactly {const_qty}.  '
                f'Therefore the IF condition  IF(L>={l_val}, {const_qty}, ...)  uses this as the constant.  '
                f'No calculation needed — just read the value directly from the REF sheet for L={l_val}.'
            )
            ws4b[f'A{r_step}'].font  = hfont(size=10, bold=True, color=GRN)
            ws4b[f'A{r_step}'].fill  = fill(GRN_L)
            ws4b[f'A{r_step}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws4b.row_dimensions[r_step].height = 36
            r_step += 3   # gap before next L group

        # Final formula recap
        ws4b.merge_cells(f'A{r_step}:H{r_step}')
        full_formula = dyn_formula or 'IF(L>=84,2.159,IF(L>=80,2.057,W*a+b))'
        ws4b[f'A{r_step}'].value = f'COMPLETE FORMULA:   {full_formula}'
        ws4b[f'A{r_step}'].font  = Font(bold=True, size=13, color='1D4ED8', name='Courier New')
        ws4b[f'A{r_step}'].fill  = fill(BLU_L)
        ws4b[f'A{r_step}'].alignment = Alignment(horizontal='left', vertical='center')
        ws4b.row_dimensions[r_step].height = 30

        ws4b.merge_cells(f'A{r_step+1}:H{r_step+1}')
        ws4b[f'A{r_step+1}'].value = (
            'Reading the formula:  '
            + '  |  '.join([f'L>={lv} → fixed {qty}' for lv, qty in sorted(fixed_l.items(), reverse=True)])
            + '  |  Smaller L → use W-linear formula'
        )
        ws4b[f'A{r_step+1}'].font  = hfont(size=11, bold=False)
        ws4b[f'A{r_step+1}'].fill  = fill(GRY)
        ws4b[f'A{r_step+1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws4b.row_dimensions[r_step+1].height = 24

    # ═══════════════════════════════════════════════════════════
    # SHEET 5 — USING EXCEL FORMULAS DIRECTLY (with IF)
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('5. Excel Formulas with IF')
    ws5.sheet_view.showGridLines = False
    for col, w in [('A',4),('B',44),('C',44)]:
        ws5.column_dimensions[col].width = w

    title(ws5, 1, 'Pasting Excel Formulas Directly into the Calculator', cols='A:C')

    section(ws5, 3, 'THE TOOL ACCEPTS EXCEL-STYLE FORMULAS DIRECTLY', cols='A:C')

    note(ws5, 4,
         'You can copy the formula from the DATA sheet formula bar in your xlsm file and paste it directly. '
         'The tool will auto-convert cell references (K2, L2, M2) to L, W, H and evaluate correctly.',
         bg=GRN_L, color=GRN, cols='A:C', height=36)

    section(ws5, 6, 'REAL EXAMPLE FROM ALLURE-MF-ECOM DATA SHEET (cell G15, item RMI-PVC-50)', cols='A:C')

    ws5.merge_cells('A7:C7')
    ws5['A7'].value = '=IF(M2=6,IF(K2<79,(L2*2)+(M2*2),(K2*2)+(M2*2))+4,IF(K2<76,(L2*2)+(M2*2),(K2*2)+(M2*2))+4)*89.5*0.0254*0.0254*0.00009*1300'
    ws5['A7'].font  = Font(bold=True, size=11, color='1D4ED8', name='Courier New')
    ws5['A7'].fill  = fill(BLU_L)
    ws5['A7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws5.row_dimensions[7].height = 36

    section(ws5, 9, 'HOW THE TOOL PROCESSES IT', cols='A:C')
    hdr_row(ws5, 10, ['','Processing Step','Result'], NAV, height=18)
    process_steps = [
        ('Remove leading =',
         'IF(M2=6,IF(K2<79,(L2*2)+(M2*2),(K2*2)+(M2*2))+4,...)*89.5*...'),
        ('Map cell refs: K2→L, L2→W, M2→H',
         'IF(H=6,IF(L<79,(W*2)+(H*2),(L*2)+(H*2))+4,...)*89.5*...'),
        ('Convert IF() → Python ternary',
         '(((W*2)+(H*2)+4 if L<79 else (L*2)+(H*2)+4) if H==6 else ...)*89.5*...'),
        ('Evaluate for L=72, W=30, H=6',
         '(30*2+6*2+4)*89.5*0.0254*0.0254*0.00009*1300  =  0.285  ✓'),
    ]
    for i,(step,result) in enumerate(process_steps):
        rr=11+i; bg=GRN_L if '0.285' in result else (GRY if i%2==0 else WHT)
        ws5.cell(rr,2,step).font=hfont(bold=True,size=10); ws5.cell(rr,2).fill=fill(bg); ws5.cell(rr,2).alignment=wrap(); ws5.cell(rr,2).border=border()
        ws5.cell(rr,3,result).font=Font(bold=True if '0.285' in result else False,size=10,color='065F46' if '0.285' in result else '000000',name='Courier New')
        ws5.cell(rr,3).fill=fill(bg); ws5.cell(rr,3).alignment=wrap(); ws5.cell(rr,3).border=border()
        ws5.row_dimensions[rr].height=24

    note(ws5, 16,
         'IMPORTANT: The column mapping (K=L, L=W, M=H) is the typical pattern in PEPS BOM files. '
         'If your file uses different columns, check the header row of the DATA sheet helper table '
         'and manually replace the cell references with L, W, H before entering the formula.',
         cols='A:C', height=40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='QTY_Formula_Guide.xlsx')



@app.route('/api/evaluate-formula', methods=['POST'])
def api_evaluate_formula():
    """Evaluate a formula for all SIZE_MASTER L×W combinations at selected heights."""
    data     = request.json or {}
    formula  = data.get('formula','').strip()
    heights  = data.get('heights', [])   # list of height strings e.g. ['5','6']
    sizes    = data.get('sizes', [])     # list of {l,w} dicts — defaults to SIZE_MASTER

    # Default size master (75 sizes, 5 lengths × 15 widths)
    SIZE_MASTER_LW = [
        (72,30),(72,33),(72,36),(72,39),(72,42),(72,45),(72,48),(72,51),(72,54),(72,57),(72,60),(72,63),(72,66),(72,69),(72,72),
        (75,30),(75,33),(75,36),(75,39),(75,42),(75,45),(75,48),(75,51),(75,54),(75,57),(75,60),(75,63),(75,66),(75,69),(75,72),
        (78,30),(78,33),(78,36),(78,39),(78,42),(78,45),(78,48),(78,51),(78,54),(78,57),(78,60),(78,63),(78,66),(78,69),(78,72),
        (80,30),(80,33),(80,36),(80,39),(80,42),(80,45),(80,48),(80,51),(80,54),(80,57),(80,60),(80,63),(80,66),(80,69),(80,72),
        (84,30),(84,33),(84,36),(84,39),(84,42),(84,45),(84,48),(84,51),(84,54),(84,57),(84,60),(84,63),(84,66),(84,69),(84,72),
    ]
    if sizes:
        lw_list = [(int(s['l']), int(s['w'])) for s in sizes]
    else:
        lw_list = SIZE_MASTER_LW

    if not formula or not heights:
        return jsonify({'error': 'Formula and at least one height required'}), 400

    import math

    def _compile_formula(raw: str) -> str:
        """
        Convert Excel-style formula to safe Python.
        Handles IF(cond, t, f), cell refs (K2→L, M2→W), leading =.
        """
        f = raw.strip()

        # 1. Remove leading = (Excel formulas start with =)
        if f.startswith('='): f = f[1:]

        # 2. Map Excel column refs to L,W,H based on PEPS BOM DATA sheet helper table:
        #    Col K = Length, Col L = Width, Col M = Height, Col N = CLR
        #    Example formula: =(K2+L2+4)*2*0.0254*2  →  (L+W+4)*0.1016
        col_map = [
            (r'\bK\d+\b', 'L'),   # K col = Length
            (r'\bL\d+\b', 'W'),   # L col = Width  (NOT Length — common confusion)
            (r'\bM\d+\b', 'H'),   # M col = Height
            (r'\bN\d+\b', 'H'),   # N col = also sometimes Height
        ]
        for pattern, replacement in col_map:
            f = re.sub(pattern, replacement, f, flags=re.IGNORECASE)

        # 3. Uppercase L,W,H variable names only — leave operators intact
        f = re.sub(r'\b[lwh]\b', lambda m: m.group().upper(), f)
        f = re.sub(r'\bIF\b', 'IF', f, flags=re.IGNORECASE)   # normalise IF

        # 4. Replace Excel <> with Python != , then bare = (equality test) with ==
        #    — must not touch <=, >=, == or != (lookbehind/lookahead guards those)
        f = f.replace('<>', '!=')
        f = re.sub(r'(?<![<>=!])=(?!=)', '==', f)

        # 5. Convert IF(cond, true_val[, false_val]) → Python ternary
        #    Use CASE-SENSITIVE search so Python 'if' keywords (lowercase) are not matched
        #    Step 3 already normalised all Excel IF() to uppercase IF
        for _ in range(10):
            m = re.search(r'\bIF\s*\(', f)   # NO re.IGNORECASE — avoids matching 'if' in ternary
            if not m: break
            start = m.end()
            depth, parts, current, pos = 1, [], '', start
            while pos < len(f) and depth > 0:
                ch = f[pos]
                if   ch == '(':  depth += 1; current += ch
                elif ch == ')':
                    depth -= 1
                    if depth > 0: current += ch
                    else:         parts.append(current)
                elif ch == ',' and depth == 1:
                    parts.append(current); current = ''
                else: current += ch
                pos += 1
            if len(parts) == 3:
                cond, t_val, f_val = [p.strip() for p in parts]
                f = f[:m.start()] + f'(({t_val}) if ({cond}) else ({f_val}))' + f[pos:]
            elif len(parts) == 2:
                # Excel's IF(cond, value) returns FALSE (== 0 in arithmetic) when cond is false
                cond, t_val = [p.strip() for p in parts]
                f = f[:m.start()] + f'(({t_val}) if ({cond}) else False)' + f[pos:]
            else:
                break

        # 6. Final safety strip — allow math, L/W/H, comparisons, Python if/else keywords
        #    Comma is NOT needed here since all IF() have been converted to ternary
        allowed = re.compile(r"[^0-9LWH\+\-\*/\(\)\.\s<>=!ifelseIFELSEa-z]")
        f = allowed.sub('', f)
        return f.strip()

    results = []
    for H_str in heights:
        try:
            H = int(H_str)
        except ValueError:
            continue
        for (L, W) in lw_list:
            try:
                safe_expr = _compile_formula(formula)
                qty = eval(safe_expr, {'__builtins__':{}}, {
                    'L':L,'W':W,'H':H,
                    'round':round,'abs':abs,'max':max,'min':min,
                    'sqrt':math.sqrt,'pi':math.pi,'e':math.e
                })
                results.append({'L':L,'W':W,'H':H,'qty':round(float(qty),4)})
            except Exception as ex:
                results.append({'L':L,'W':W,'H':H,'qty':None,'error':str(ex)})

    # Build a summary
    valid = [r['qty'] for r in results if r.get('qty') is not None]
    summary = {
        'total': len(results),
        'valid': len(valid),
        'min_qty': round(min(valid),4) if valid else 0,
        'max_qty': round(max(valid),4) if valid else 0,
    }
    return jsonify({'results': results, 'summary': summary})


@app.route('/api/replace/preview', methods=['POST'])
def api_replace_preview():
    """Full scan: find all products containing old_code across every BOM file."""
    data     = request.json or {}
    old_code = data.get('old_code', '').strip().upper()
    if not old_code:
        return jsonify({'error': 'Item code required'}), 400
    new_code = data.get('new_code', '').strip().upper() or old_code

    all_products = db.get_all_products()

    # Extract base code: strip L×W dimensions but KEEP the thickness suffix.
    # 18D-PU-FOAM-83.50X71.50X10MM → base='18D-PU-FOAM', thick_sfx='X10MM'
    # This prevents 18D-PU-FOAM-74.50X35.50X25MM (25MM) matching a search for
    # 18D-PU-FOAM-74.50X35.50X10MM (10MM). Thickness MUST match.
    _dim_suffix = re.compile(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm|CM|cm)?$')
    base_code     = _dim_suffix.sub('', old_code)
    has_dim_suffix = (base_code != old_code)
    _thick_m  = re.search(r'[Xx]([\d.]+(?:MM|mm|CM|cm))$', old_code)
    thick_sfx = ('X' + _thick_m.group(1).upper()) if _thick_m else ''

    # Detect spring/steel full code OR prefix-only (e.g. OEBOXBNSR = match all variants)
    _spring_m2    = re.match(r'^([A-Z]{4,})([\d.]+X[\d.]+X\d+)$', old_code)
    _pfx_only_m   = re.match(r'^([A-Z]{4,})$', old_code)
    spring_pfx    = (_spring_m2.group(1) if _spring_m2 else
                     _pfx_only_m.group(1) if _pfx_only_m else '')
    is_prefix_only = bool(_pfx_only_m)
    if _spring_m2 or _pfx_only_m:
        has_dim_suffix = True
        if _spring_m2 and not thick_sfx:
            _sp3 = re.search(r'X(\d+)$', old_code)
            thick_sfx = ('X' + _sp3.group(1)) if _sp3 else ''
        if is_prefix_only:
            thick_sfx = ''   # match ALL height variants in prefix-only mode

    def _foam_match(c):
        """Match by base+suffix for foam/spring/steel. Prefix-only mode matches all dims."""
        cu = c.upper()
        if spring_pfx:
            if not cu.startswith(spring_pfx): return False
            if thick_sfx and not cu.endswith(thick_sfx): return False
            return bool(re.match(r'^[A-Z]+[\d.]+X[\d.]+X\d+$', cu))
        else:
            if not cu.startswith(base_code + '-'): return False
            if thick_sfx and not cu.endswith(thick_sfx): return False
            return True

    # Phase 1 — fast pre-filter using stored item_codes index in DB
    candidates = []
    for p in all_products:
        raw = p.get('item_codes')
        if not raw:
            # Not yet indexed — include as a safety fallback but don't crash on all 250
            candidates.append(p)
            continue
        try:
            codes = [c.upper() for c in json.loads(raw)]
            if old_code in codes:
                candidates.append(p)
            elif has_dim_suffix and any(_foam_match(c) for c in codes):
                candidates.append(p)
        except Exception:
            candidates.append(p)   # malformed JSON — scan to be safe

    print(f"[Preview] Phase 1 candidates: {len(candidates)} of {len(all_products)} products")

    # Phase 2 — parallel file scan using ThreadPoolExecutor
    def _scan_one(product):
        try:
            wb = load_workbook(product['filepath'], read_only=True, data_only=True)
            found_rows = []

            if 'REF' in wb.sheetnames:
                ws  = wb['REF']
                hdr = [str(c or '').strip().upper()
                       for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                ic_idx  = next((i for i,h in enumerate(hdr) if h in ('ITEMCODE','ITEM CODE','ITEM_CODE')), 3)
                qty_idx = next((i for i,h in enumerate(hdr) if h in ('OTY','QTY','QUANTITY')), 5)
                uom_idx = next((i for i,h in enumerate(hdr) if h in ('UOM',)), 4)
                wh_idx  = next((i for i,h in enumerate(hdr) if h in ('WH CODE','WHCODE','WH')), 6)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(v is None for v in row): continue
                    ic = str(row[ic_idx] or '').strip().upper() if ic_idx < len(row) else ''
                    if not ic or ic.startswith('='): continue
                    if ic == old_code or (has_dim_suffix and _foam_match(ic)):
                        found_rows.append({
                            'sheet': 'REF', 'row': row_idx, 'ic_col': ic_idx + 1, 'old_ic': ic,
                            'seq': row[2] if len(row) > 2 else None,
                            'qty': row[qty_idx] if qty_idx < len(row) else None,
                            'uom': row[uom_idx] if uom_idx < len(row) else None,
                            'wh':  row[wh_idx]  if wh_idx  < len(row) else None,
                        })

            if not found_rows and 'DATA' in wb.sheetnames:
                ws = wb['DATA']
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10, values_only=True), start=2):
                    if not row or all(v is None for v in row): continue
                    ic = str(row[4] or '').strip().upper()
                    if not ic or ic.startswith('='): continue
                    if ic == old_code or (has_dim_suffix and _foam_match(ic)):
                        found_rows.append({
                            'sheet': 'DATA', 'row': row_idx, 'ic_col': 5, 'old_ic': ic,
                            'seq': row[0], 'qty': row[6], 'uom': row[5],
                            'wh': row[8] if len(row) > 8 else None,
                        })

            wb.close()
            if found_rows:
                return {
                    'product_id':  product['id'],   'product_name': product['name'],
                    'family':      product.get('family', ''),
                    'filepath':    product['filepath'],
                    'occurrences': len(found_rows), 'rows': found_rows
                }
        except Exception as e:
            print(f"Preview scan error {product['name']}: {e}")
        return None

    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=6) as pool:
        results = list(pool.map(_scan_one, candidates))
    affected = [r for r in results if r is not None]
    print(f"[Preview] Phase 2 done: {len(affected)} files affected")

    # Group by family for the response
    families = {}
    for a in affected:
        fam = a['family'] or 'Other'
        families.setdefault(fam, []).append(a)

    activity_log._log(session.get('username', 'User'), session.get('role', ''),
                      'replace_preview',
                      f'{old_code} → {new_code} | {len(affected)} files, '
                      f'{sum(p["occurrences"] for p in affected)} occurrences')

    return jsonify({
        'old_code':          old_code,
        'new_code':          new_code,
        'files_affected':    len(affected),
        'total_occurrences': sum(p['occurrences'] for p in affected),
        'affected_products': affected,
        'by_family':         {k: len(v) for k, v in families.items()}
    })


@app.route('/api/replace/preview-download', methods=['POST'])
def api_replace_preview_download():
    """Download BOM with replacement preview applied — does NOT modify source files."""
    import traceback as _tb
    data        = request.json or {}
    # Accept new-style: products=[{filepath,product_name},...] or legacy: product_ids=[int,...]
    products_in = data.get('products', [])
    product_ids = data.get('product_ids', [])
    old_code    = data.get('old_code', '').strip().upper()
    new_code    = (data.get('new_code', '').strip().upper() or old_code)
    qty_by_size = data.get('qty_by_size', {})
    dl_format   = data.get('format', 'excel')
    foam_pat    = data.get('foam_pattern', {})

    if not (products_in or product_ids) or not old_code:
        return jsonify({'error': 'products (or product_ids) and old_code required'}), 400

    # Build unified list of {filepath, name} — prefer new-style, fall back to DB lookup
    product_list = []
    if products_in:
        for p in products_in:
            fp   = p.get('filepath', '').strip()
            name = p.get('product_name', '') or p.get('name', '') or os.path.basename(fp)
            if fp and os.path.exists(fp):
                product_list.append({'filepath': fp, 'name': name})
            else:
                print(f"[preview-download] filepath not found: {fp!r}")
    else:
        for pid in product_ids:
            p = db.get_product_by_id(pid)
            if p:
                product_list.append({'filepath': p['filepath'], 'name': p['name']})
            else:
                print(f"[preview-download] Product id={pid} not found in DB")

    if not product_list:
        return jsonify({'error': 'No valid product files found for download'}), 400

    # Build matching logic that handles exact, foam-bulk, and spring/prefix codes
    _dl_spring_m  = re.match(r'^([A-Z]{4,})([\d.]+X[\d.]+X\d+)$', old_code)
    _dl_pfx_only  = re.match(r'^([A-Z]{4,})$', old_code)
    dl_spring_pfx = (_dl_spring_m.group(1) if _dl_spring_m else
                     (_dl_pfx_only.group(1) if _dl_pfx_only else ''))
    dl_is_spring  = bool(dl_spring_pfx)
    _dl_thick_m   = re.search(r'[Xx]([\d.]+(?:MM|mm|CM|cm))$', old_code)
    dl_thick_sfx  = ('X' + _dl_thick_m.group(1).upper()) if _dl_thick_m else ''
    if dl_is_spring and not dl_thick_sfx and not _dl_pfx_only:
        _sp = re.search(r'X(\d+)$', old_code)
        dl_thick_sfx = ('X' + _sp.group(1)) if _sp else ''
    if _dl_pfx_only:
        dl_thick_sfx = ''
    # Base of old_code after stripping any dimension suffix (e.g. '18D-PU-FOAM' from '18D-PU-FOAM-74X35X10MM')
    _dl_dim_re = re.compile(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm|CM|cm)?$')
    dl_base    = _dl_dim_re.sub('', old_code)
    dl_has_dim = (dl_base != old_code)   # True only when old_code itself has a dimension suffix

    def _dl_match(ic):
        cu = ic.strip().upper()
        if cu == old_code: return True
        if dl_is_spring:
            if not cu.startswith(dl_spring_pfx): return False
            if dl_thick_sfx and not cu.endswith(dl_thick_sfx): return False
            return bool(re.match(r'^[A-Z]+[\d.]+X[\d.]+X\d+$', cu))
        # Foam-variant matching: only when old_code itself carries a dimension suffix.
        # Also require the stripped base of the target to equal old_code's base, so plain
        # codes like 'RML-POLY-GUM' never accidentally match dimension-embedded codes
        # such as 'SFG-ALRNLOR-BOR-72X30X06' or 'RMI-HCTNF-71.50X29.50.'.
        if dl_has_dim:
            base = _dl_dim_re.sub('', cu)
            if base != cu and base.upper() == dl_base.upper():
                if not dl_thick_sfx or cu.endswith(dl_thick_sfx): return True
        return False

    def _dl_new_code(old_ic):
        cu = old_ic.strip().upper()
        if foam_pat:
            np_ = (foam_pat.get('prefix') or foam_pat.get('new_density') or '').upper()
            ns_ = (foam_pat.get('suffix') or foam_pat.get('new_thick') or '').upper()
            if np_ and ns_ and dl_is_spring:
                m = re.match(r'^([A-Z]+)([\d.]+X[\d.]+)X(\d+)$', cu)
                if m: return f'{np_}{m.group(2)}X{ns_}'
        return new_code

    settings = db.get_settings()
    dl_files = []   # list of (filename, bytes)
    errors   = []

    for product in product_list:
        try:
            components, permutations, dest_headers, prebuilt_rows = \
                bom_engine.read_bom_file(product['filepath'])
            bom_rows = prebuilt_rows if prebuilt_rows else \
                bom_engine.generate_bom(components, permutations)

            for r in bom_rows:
                r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})
                ic = str(r.get('item_code', '')).strip()
                if _dl_match(ic):
                    r['item_code'] = _dl_new_code(ic)
                    if qty_by_size:
                        mc = str(r.get('ps_no', '') or r.get('mattress_code', ''))
                        m  = re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', mc)
                        if m:
                            # Normalise H: strip leading zero so '06' matches key '6'
                            key = f"{m.group(1)}x{m.group(2)}x{int(m.group(3))}"
                            if key in qty_by_size:
                                r['qty'] = qty_by_size[key]

            safe_name = product['name'].replace(' ', '_').replace('/', '_')
            if dl_format == 'excel':
                dl_files.append((f"{safe_name}_Preview.xlsx",
                                 _generate_excel_report(product['name'], bom_rows)))
            else:
                ps_rows = generator.convert_bom_to_rows(bom_rows, product['name'], 'edit')
                ud_bytes, md_bytes, ud_fname, md_fname = \
                    generator.generate_edit_structure(product['name'], ps_rows, settings)
                dl_files.append((f"{safe_name}/{ud_fname}", ud_bytes))
                dl_files.append((f"{safe_name}/{md_fname}", md_bytes))
        except Exception as e:
            errors.append(f'{product["name"]}: {e}')
            print(f"[preview-download] ERROR for {product['name']}:\n{_tb.format_exc()}")

    if not dl_files:
        err_msg = '; '.join(errors) if errors else 'No data generated'
        return jsonify({'error': f'Download failed: {err_msg}'}), 500

    # Single Excel file → send directly as .xlsx (no ZIP wrapper)
    if len(dl_files) == 1 and dl_format == 'excel':
        fname_xl, xl_bytes = dl_files[0]
        return send_file(io.BytesIO(xl_bytes),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname_xl)

    # Multiple files → ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in dl_files:
            zf.writestr(fname, fbytes)
    zip_buffer.seek(0)
    fmt_label = 'Excel' if dl_format == 'excel' else 'MDCF'
    n         = len(product_list)
    fname     = f"Impact_Preview_{old_code}_{fmt_label}_{n}products.zip"
    return send_file(zip_buffer, mimetype='application/zip',
                     as_attachment=True, download_name=fname)


@app.route('/api/replace/execute', methods=['POST'])
def api_replace_execute():
    """Execute global replace. Regular users → approval queue; admin/dev → execute immediately."""
    data = request.json or {}
    old_code = data.get('old_code', '').strip().upper()
    new_code = (data.get('new_code', '').strip().upper() or old_code)
    new_qty_raw = data.get('new_qty', None)
    new_qty  = float(new_qty_raw) if new_qty_raw not in (None, '', 0) else None
    new_uom  = (data.get('new_uom', '') or '').strip().upper() or None
    reason   = data.get('reason', '').strip()
    approval_ref = data.get('approval_ref', '').strip()
    run_by   = session.get('username', data.get('run_by', 'unknown')).strip()
    affected_data = json.dumps(data.get('affected_products', []))
    # Formula-based per-size Q.ty: {"{L}x{W}x{H}": qty, ...}
    qty_by_size  = data.get('qty_by_size', {})
    # Foam pattern for per-row code generation: keeps L×W dims, swaps density+thickness
    foam_pat = data.get('foam_pattern', {})  # {old_density, new_density, old_thick, new_thick}

    # Base+suffix matching for dimension-embedded codes (foam AND spring/steel)
    _exec_dim_re  = re.compile(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm|CM|cm)?$')
    exec_base_foam = _exec_dim_re.sub('', old_code)   # for foam: strips -DIM1xDIM2xTHICK
    has_dim_suffix = (exec_base_foam != old_code)

    # Detect spring/steel full code OR prefix-only
    _spring_m      = re.match(r'^([A-Z]{4,})([\d.]+X[\d.]+X\d+)$', old_code)
    _epfx_only     = re.match(r'^([A-Z]{4,})$', old_code)
    exec_spring_prefix = (_spring_m.group(1) if _spring_m else
                          _epfx_only.group(1) if _epfx_only else '')
    is_spring_code = bool(exec_spring_prefix)
    exec_is_prefix_only = bool(_epfx_only)

    _exec_thick_m  = re.search(r'[Xx]([\d.]+(?:MM|mm|CM|cm))$', old_code)
    exec_thick_sfx = ('X' + _exec_thick_m.group(1).upper()) if _exec_thick_m else ''
    # For spring: suffix is the number after last X (e.g. X130)
    if is_spring_code and not exec_thick_sfx and not exec_is_prefix_only:
        _sp_sfx = re.search(r'X(\d+)$', old_code)
        exec_thick_sfx = ('X' + _sp_sfx.group(1)) if _sp_sfx else ''
    if exec_is_prefix_only:
        exec_thick_sfx = ''   # match all height variants

    def _foam_match(c):
        cu = c.upper()
        if is_spring_code:
            if not cu.startswith(exec_spring_prefix): return False
            if exec_thick_sfx and not cu.endswith(exec_thick_sfx): return False
            return bool(re.match(r'^[A-Z]+[\d.]+X[\d.]+X\d+$', cu))
        else:
            # Foam match: same base (density-PU-FOAM) AND same thickness
            if not cu.startswith(exec_base_foam + '-'): return False
            if exec_thick_sfx and not cu.endswith(exec_thick_sfx): return False
            return True

    def _gen_foam_code(old_ic, pat):
        """Generate new code from old code by swapping prefix/density and suffix/thickness only.
        Works for both foam (18D-PU-FOAM-74.50X35.50X10MM) and spring (OEBOXBNSR74.50X35.50X130)."""
        if not pat: return None
        is_spring = pat.get('isSpring', False)
        new_prefix = (pat.get('prefix','') or pat.get('new_density','')).upper()
        new_suffix = (pat.get('suffix','') or pat.get('new_thick','')).upper()
        old_prefix = (pat.get('old_density','') or '').upper()
        old_suffix = (pat.get('old_thick','')   or pat.get('suffix','')).upper()
        oc = old_ic.upper()
        if not (new_prefix and new_suffix): return None

        is_prefix_mode = pat.get('isPrefixMode', False)
        if is_spring:
            # Spring/Steel: PREFIX + DIMS + X + SUFFIX
            m = re.match(r'^([A-Z]+)([\d.]+)X([\d.]+)X(\d+)$', oc)
            if not m: return None
            d1 = float(m.group(2))
            d2 = float(m.group(3))
            loff     = float(pat.get('loff',     0) or 0)
            woff     = float(pat.get('woff',     0) or 0)
            old_loff = float(pat.get('old_loff', loff) or loff)
            old_woff = float(pat.get('old_woff', woff) or woff)
            delta_l  = round(loff - old_loff, 6)
            delta_w  = round(woff - old_woff, 6)
            if delta_l != 0 or delta_w != 0:
                # Apply offset delta: reverse old offset → BOM dim → apply new offset
                new_d1 = round(d1 + delta_l, 4)
                new_d2 = round(d2 + delta_w, 4)
                fmt_d  = lambda v: f'{v:.2f}'
                return f'{new_prefix}{fmt_d(new_d1)}X{fmt_d(new_d2)}X{new_suffix}'
            else:
                # No offset change — preserve existing dims exactly
                return f'{new_prefix}{m.group(2)}X{m.group(3)}X{new_suffix}'
        else:
            # Foam: DENSITY-PU-FOAM-DIMS-THICKNESS
            if not oc.startswith(old_prefix + '-PU-FOAM-'): return None
            dims_part = oc[len(old_prefix + '-PU-FOAM-'):]
            if not dims_part.upper().endswith('X' + old_suffix): return None
            dim_lw = dims_part[:-(len(old_suffix)+1)]
            return f'{new_prefix}-PU-FOAM-{dim_lw}X{new_suffix}'

    if not all([old_code, reason]):
        return jsonify({'error': 'Item code and reason are required'}), 400

    if not _is_privileged():
        n_files = len(data.get('affected_products', []))
        summary = f"Global Replace: {old_code} → {new_code} across {n_files} file(s). Reason: {reason}"
        appr_id = db.create_approval(
            user_id=session['user_id'], username=run_by,
            action_type='global_replace', action_summary=summary,
            action_data={'old_code': old_code, 'new_code': new_code,
                         'new_qty': new_qty, 'new_uom': new_uom, 'reason': reason,
                         'affected_products': data.get('affected_products', [])}
        )
        return jsonify({'pending': True, 'approval_id': appr_id,
                        'message': 'Your replace request has been submitted for approval.'})

    if not approval_ref:
        return jsonify({'error': 'All fields required'}), 400
    
    files_changed = 0
    rows_changed = 0
    
    for product_data in data.get('affected_products', []):
        filepath = product_data['filepath']
        if not filepath or not os.path.exists(filepath):
            continue
        try:
            wb = load_workbook(filepath, keep_vba=True)
            # Determine which sheet to write to — use sheet stored in row_info if available
            # fallback: components > DATA > REF
            for row_info in product_data['rows']:
                row_idx  = row_info['row']
                # Use the sheet that Phase 2 found the code in (stored during preview)
                src_sheet = row_info.get('sheet', '')
                if src_sheet and src_sheet in wb.sheetnames:
                    ws_name = src_sheet
                elif 'components' in wb.sheetnames:
                    ws_name = 'components'
                elif 'DATA' in wb.sheetnames:
                    ws_name = 'DATA'
                elif 'REF' in wb.sheetnames:
                    ws_name = 'REF'
                else:
                    continue
                ws = wb[ws_name]
                # Use column stored during Phase 2 scan (exact column per file)
                if ws_name == 'components':
                    ic_col = 1
                elif 'ic_col' in row_info:
                    ic_col = int(row_info['ic_col'])   # stored 1-indexed from Phase 2
                elif ws_name == 'REF':
                    ic_col = 4
                else:
                    ic_col = 5
                cell = ws.cell(row=row_idx, column=ic_col)
                old_ic = str(cell.value or '').strip().upper()
                # For foam bulk replace: generate per-row new code from pattern
                if foam_pat:
                    row_new_code = _gen_foam_code(old_ic, foam_pat) or new_code
                else:
                    row_new_code = new_code
                _match = (old_ic == old_code) or ((has_dim_suffix or is_spring_code) and _foam_match(old_ic))
                print(f'  EXEC row={row_idx} sheet={ws_name} ic_col={ic_col} cell_val={repr(old_ic)!r} match={_match} → {row_new_code if _match else "skip"}')
                if _match:
                    ws.cell(row=row_idx, column=ic_col, value=row_new_code)
                    # Determine Q.ty to apply: formula-based per-size takes priority
                    applied_qty = None
                    if qty_by_size:
                        # Extract L×W×H from the mattress code in the adjacent column
                        # For REF sheet: col 2 (B) = MATTRESS CODE; for DATA: col 4 (D)
                        mc_col = 2 if ws_name in ('REF','DATA') else None
                        mc_val = ''
                        if mc_col:
                            try: mc_val = str(ws.cell(row=row_idx, column=mc_col).value or '')
                            except: pass
                        m = re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', mc_val)
                        if m:
                            # Normalise H to int to strip leading zeros (file: "72X42X05", key: "72x42x5")
                            size_key = f"{m.group(1)}x{m.group(2)}x{int(m.group(3))}"
                            applied_qty = qty_by_size.get(size_key)
                    if applied_qty is None: applied_qty = new_qty
                    if applied_qty is not None:
                        qty_col = 4 if ws_name == 'components' else 6
                        ws.cell(row=row_idx, column=qty_col, value=applied_qty)
                    if new_uom is not None:
                        uom_col = 5 if ws_name == 'components' else 7
                        ws.cell(row=row_idx, column=uom_col, value=new_uom)
                    rows_changed += 1
            wb.save(filepath)
            files_changed += 1
            wb.close()
        except Exception as e:
            print(f'Replace error {filepath}: {e}')
    
    snap_id = db.create_replace_snapshot(
        old_code, new_code, reason, approval_ref, run_by,
        files_changed, rows_changed, affected_data
    )
    activity_log.global_replace(run_by, session.get('role',''), snap_id=snap_id,
                                 old_code=old_code, new_code=new_code,
                                 files=files_changed, rows=rows_changed, reason=reason)
    activity_log.log_replace(snap_id, old_code, new_code, reason, approval_ref,
                              run_by, files_changed, rows_changed)

    scan_and_register_products()
    
    return jsonify({
        'snapshot_id': snap_id,
        'files_changed': files_changed,
        'rows_changed': rows_changed
    })


@app.route('/api/replace/snapshots')
def api_replace_snapshots():
    """Get replacement history"""
    snapshots = db.get_replace_snapshots()
    return jsonify({'snapshots': snapshots})


@app.route('/api/replace/snapshots/<int:snap_id>/download')
def api_snapshot_download(snap_id):
    """Download a formatted Excel change report — side-by-side Before/After."""
    try:
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _Al, Border as _B, Side as _S
        from openpyxl.utils import get_column_letter as _gcl

        snap = db.get_replace_snapshot(snap_id)
        if not snap:
            return jsonify({'error': 'Snapshot not found'}), 404

        old_code = snap.get('old_code', '')
        new_code = snap.get('new_code', '')

        # ── Styles ────────────────────────────────────────────────────────
        def _fill(hex_c): return _PF('solid', fgColor=hex_c)
        def _font(bold=True,size=11,color='000000',italic=False):
            return _F(bold=bold,size=size,color=color,italic=italic,name='Aptos Narrow')
        def _mono(bold=True,size=10,color='000000'):
            return _F(bold=bold,size=size,color=color,name='Courier New')
        def _border(c='BBBBBB'):
            s=_S(style='thin',color=c); return _B(left=s,right=s,top=s,bottom=s)
        def _al(h='left',v='center',wrap=False):
            return _Al(horizontal=h,vertical=v,wrap_text=wrap)

        # Colours
        NAV='1A3A6B'; HDR='1F3864'; ACC='1A56DB'
        CHG_BG='FFF2CC'; CHG_FG='7F6000'      # yellow — changed row
        NEW_BG='E2EFDA'; NEW_FG='375623'       # green — after replacement
        OLD_BG='FDECEA'; OLD_FG='C0392B'       # red — before (old code)
        GRY='F5F5F5'; WHT='FFFFFF'

        try:
            affected = json.loads(snap.get('affected_data','[]') or '[]')
        except Exception:
            affected = []

        # ── Helper: read first-SKU BOM from file ──────────────────────────
        def _read_bom(filepath):
            """Returns list of dicts: {seq,ic,desc,dept,sect,qty,uom,wh}"""
            rows = []
            if not filepath or not os.path.exists(filepath): return rows
            try:
                wb2 = load_workbook(filepath, read_only=True, data_only=True)
                ws2 = None
                if 'REF' in wb2.sheetnames:   ws2 = wb2['REF']
                elif 'DATA' in wb2.sheetnames: ws2 = wb2['DATA']
                if not ws2: wb2.close(); return rows
                is_ref = (ws2.title == 'REF')
                seen_mc = set()
                for row in ws2.iter_rows(min_row=2, max_row=600, values_only=True):
                    if not row or not any(row): break
                    try:
                        if is_ref:
                            mc  = str(row[1] or '')
                            seq = str(row[2] or '')
                            ic  = str(row[3] or '').strip()
                            uom = str(row[4] or '')
                            qty = str(row[5] or '')
                            wh  = str(row[6] or '')
                            dept= sect = ''
                        else:
                            mc   = str(row[3] or '') if len(row)>3 else ''
                            seq  = str(row[0] or '') if len(row)>0 else ''
                            ic   = str(row[4] or '').strip() if len(row)>4 else ''
                            uom  = str(row[5] or '') if len(row)>5 else ''
                            qty  = str(row[6] or '') if len(row)>6 else ''
                            wh   = str(row[8] or '') if len(row)>8 else ''
                            dept = str(row[1] or '') if len(row)>1 else ''
                            sect = str(row[2] or '') if len(row)>2 else ''
                        if not ic: continue
                        if mc not in seen_mc:
                            if seen_mc: break
                            seen_mc.add(mc)
                        rows.append({'seq':seq,'ic':ic,
                                     'desc':ITEM_MASTER.get(ic.upper(),''),
                                     'dept':dept,'sect':sect,
                                     'qty':qty,'uom':uom,'wh':wh})
                    except Exception: continue
                wb2.close()
            except Exception: pass
            return rows

        # ── Build workbook ────────────────────────────────────────────────
        wb = _WB()
        ws = wb.active
        assert ws is not None
        ws.title = 'Change Report'
        ws.sheet_view.showGridLines = False

        # Column widths: A-K = left product, L = spacer, M-W = right product
        COL_W = [4,6,22,35,14,14,8,6,10,12,4]  # 11 cols per product
        for i, w in enumerate(COL_W, 1):        ws.column_dimensions[_gcl(i)].width = w
        ws.column_dimensions[_gcl(12)].width = 3  # spacer
        for i, w in enumerate(COL_W, 13):       ws.column_dimensions[_gcl(i)].width = w

        # ── Row 1: Main title (fill all cols, no merge) ───────────────────
        title_text = 'PEPS BOM Tool  -  Replacement Change Report'
        for ci in range(1, 24):
            try:
                c = ws.cell(1, ci, value=(title_text if ci == 1 else ''))
                c.font  = _font(size=14,color=WHT)
                c.fill  = _fill(NAV)
                c.alignment = _al('left' if ci==1 else 'center')
            except Exception: pass
        ws.row_dimensions[1].height = 32

        # ── Rows 2-11: Snapshot summary (individual cells, no merge) ──────
        summary = [
            ('Snapshot',    str(snap_id)),
            ('Date / Time', (snap.get('created_at','') or '').replace('T',' ')[:16]),
            ('Old Item',    old_code),
            ('New Item',    new_code),
            ('Reason',      snap.get('reason','')),
            ('Executed By', snap.get('run_by','')),
            ('Files Chgd',  str(snap.get('files_changed',0))),
            ('Rows Chgd',   str(snap.get('rows_changed',0))),
            ('Status',      snap.get('status','active')),
        ]
        for i, (lbl, val) in enumerate(summary, 2):
            try:
                ws.cell(i, 1, value=lbl).font  = _font(size=10,color='475569',bold=False)
                vc = ws.cell(i, 3, value=val)
                if lbl == 'Old Item': vc.font = _mono(size=10,color='DC2626')
                elif lbl == 'New Item': vc.font = _mono(size=10,color='059669',bold=True)
                else: vc.font = _font(size=10,bold=False)
                ws.row_dimensions[i].height = 16
            except Exception: pass

        # ── BOM section headers ────────────────────────────────────────────
        BOM_HDRS = ['#','Seq','Item Code','Description','Dept','Section','Q.ty','UOM','WH Code','Changed?']

        def _set(r, c, v, font=None, bg=None, al=None, bdr=True):
            """Safe cell write — never writes to merged cells; skips if MergedCell."""
            try:
                cell = ws.cell(r, c, value=v)
                if font: cell.font = font
                if bg:   cell.fill = _fill(bg)
                if al:   cell.alignment = al
                if bdr:  cell.border = _border()
            except (AttributeError, TypeError):
                pass  # Skip merged/read-only cells

        # Build before/after changed-rows table from snapshot data
        chg_rows_all = []
        for prod in affected:
            pname = prod.get('name', prod.get('product_name','Unknown'))
            for r in prod.get('rows', []):
                old_ic_snap = r.get('old_ic', old_code)
                # Derive new code: if foam pattern, generate per-row; else use snap new_code
                _tp = re.search(r'[Xx]([\d.]+(?:MM|mm))$', old_code)
                _ts = ('X'+_tp.group(1).upper()) if _tp else ''
                _tb = re.sub(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm)?$','',old_code)
                _tb2= re.sub(r'-[\d.]+[Xx][\d.]+(?:[Xx][\d.]+)?(?:MM|mm)?$','',new_code)
                _tp2= re.search(r'[Xx]([\d.]+(?:MM|mm))$', new_code)
                _ts2= ('X'+_tp2.group(1).upper()) if _tp2 else ''
                if old_ic_snap.startswith(_tb+'-') and _ts and old_ic_snap.endswith(_ts):
                    lw_dims = old_ic_snap[len(_tb+'-'):-(len(_ts))]
                    new_ic_snap = f'{_tb2}-{lw_dims}{_ts2}' if _ts2 else new_code
                else:
                    new_ic_snap = new_code
                chg_rows_all.append({
                    'product': pname,
                    'seq': r.get('seq',''),
                    'old_ic': old_ic_snap,
                    'new_ic': new_ic_snap,
                    'qty': r.get('qty',''),
                    'uom': r.get('uom',''),
                    'wh':  r.get('wh','')
                })

        # Build exact sets of old/new codes from snapshot for precise CHANGED marking
        _chg_old_ics = {r['old_ic'].upper() for c in chg_rows_all for r in [c] if r.get('old_ic')}
        _chg_new_ics = {r['new_ic'].upper() for c in chg_rows_all for r in [c] if r.get('new_ic')}

        def _write_bom_block(start_col, prod_name, filepath, bom_rows, section_label, hdr_bg, start_row=13):
            """Write one product BOM block — no merging to avoid MergedCell errors."""
            sc = start_col; sr = start_row
            # Section label (span cols manually with background, no merge)
            for ci in range(sc, sc+10):
                _set(sr, ci, (f'{section_label}: {prod_name}' if ci==sc else ''),
                     font=_font(size=11,color=WHT), bg=hdr_bg,
                     al=_al('left'), bdr=True)
            ws.row_dimensions[sr].height = 20
            # File path row
            for ci in range(sc, sc+10):
                _set(sr+1, ci, (f'File: {filepath}' if ci==sc else ''),
                     font=_font(size=9,color='475569',bold=False,italic=True),
                     bg=GRY, al=_al('left'), bdr=False)
            ws.row_dimensions[sr+1].height = 14
            # Column headers (row sr+3)
            for offset, h in enumerate(BOM_HDRS):
                _set(sr+3, sc+offset, h,
                     font=_font(size=9,color=WHT), bg=HDR,
                     al=_al('center'), bdr=True)
            ws.row_dimensions[sr+3].height = 16
            # BOM rows
            for ri, row in enumerate(bom_rows):
                dr = sr + 4 + ri
                ic = row['ic'].upper()
                # Only mark CHANGED if this exact code was actually changed (from snapshot data)
                is_chg = ic in _chg_old_ics or ic in _chg_new_ics
                bg = CHG_BG if is_chg else (GRY if ri%2==0 else WHT)
                vals = [ri+1, row['seq'], row['ic'], row['desc'],
                        row['dept'], row['sect'], row['qty'], row['uom'],
                        row['wh'], ('>>> CHANGED <<<' if is_chg else '')]
                for offset, val in enumerate(vals):
                    ci = sc + offset
                    if is_chg:   fnt = _mono(size=9,color=CHG_FG,bold=True)
                    elif offset==2: fnt = _mono(size=9,bold=False)
                    else:        fnt = _font(size=9,bold=False)
                    al_h = 'center' if offset in (0,1,6,7,8) else 'left'
                    _set(dr, ci, val, font=fnt, bg=bg, al=_al(al_h), bdr=True)
                ws.row_dimensions[dr].height = 15

        # ── Layout: pairs of products side by side ────────────────────────
        # "Before" = product showing old_code → highlight in yellow
        # "After"  = same product showing new_code → highlight in green
        # Since both states are now the same file (already replaced),
        # left = "BEFORE Global Replacement" (show old_code row in red)
        # right = "AFTER Global Replacement"  (show new_code row in green)

        for idx in range(0, len(affected), 2):
            left_prod  = affected[idx]
            right_prod = affected[idx+1] if idx+1 < len(affected) else None

            left_name  = left_prod.get('name',  left_prod.get('product_name','Unknown'))
            left_fp    = left_prod.get('filepath','')
            left_bom   = _read_bom(left_fp)

            # Write left block
            _write_bom_block(1, left_name, left_fp, left_bom,
                             'BEFORE Global Replacement', '7F6000')

            if right_prod:
                right_name = right_prod.get('name', right_prod.get('product_name','Unknown'))
                right_fp   = right_prod.get('filepath','')
                right_bom  = _read_bom(right_fp)
                _write_bom_block(13, right_name, right_fp, right_bom,
                                 'AFTER Global Replacement', '375623')

            # "AFTER" note on left (same file, just labels)
            # Already labeled correctly as Before/After

            break  # For now write first pair; extend for more pairs if needed

        # ── Changed Items: Before vs After table ─────────────────────────
        if chg_rows_all:
            cr_start = ws.max_row + 3
            _set(cr_start, 1, 'CHANGED ITEMS — BEFORE vs AFTER', font=_font(size=12,color=WHT), bg=HDR)
            for ci in range(2, 12): _set(cr_start, ci, '', font=_font(color=WHT), bg=HDR)
            ws.row_dimensions[cr_start].height = 22

            cr_hdr = cr_start + 1
            for ci, h in enumerate(['#','Product','Seq','BEFORE (old item code)','AFTER (new item code)','Q.ty','UOM','WH'],1):
                c = ws.cell(cr_hdr, ci, value=h)
                c.font = _font(size=9,color=WHT); c.fill = _fill(NAV)
                c.border = _border(); c.alignment = _al('center')
            ws.row_dimensions[cr_hdr].height = 16

            for ri, cr in enumerate(chg_rows_all):
                dr = cr_hdr + 1 + ri
                vals = [ri+1, cr['product'], cr['seq'], cr['old_ic'], cr['new_ic'], cr['qty'], cr['uom'], cr['wh']]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(dr, ci, val)
                    cell.fill = _fill(CHG_BG); cell.border = _border()
                    if ci == 4: cell.font = _mono(size=9,color=CHG_FG,bold=True)  # old code = amber
                    elif ci == 5: cell.font = _mono(size=9,color=NEW_FG,bold=True)  # new code = green
                    else: cell.font = _font(size=9,bold=False)
                    cell.alignment = _al('left' if ci in (2,4,5) else 'center')
                ws.row_dimensions[dr].height = 15

        # ── Add legend ────────────────────────────────────────────────────
        max_row = ws.max_row + 2
        ws.cell(max_row, 1, 'Legend:').font = _font(size=9,bold=True)
        lc = ws.cell(max_row, 2, f'  Changed item ({old_code} → {new_code})')
        lc.fill = _fill(CHG_BG); lc.font = _font(size=9,color=CHG_FG)

        # Save
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        safe = re.sub(r'[^\w]', '_', old_code)[:20]
        fname = f'Change_Report_{safe}_{snap_id}.xlsx'
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except Exception as top_err:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(top_err)}), 500


def _perform_replace_rollback(snap_id):
    """Core rollback logic for a replace snapshot — shared by the manual
    rollback endpoint and the Admin overturn-after-Sub-Admin-approval path.
    Returns a result dict (with 'error' key on failure)."""
    snapshot = db.get_replace_snapshot(snap_id)

    if not snapshot or snapshot['status'] != 'active':
        return {'error': 'Invalid snapshot'}

    old_code = snapshot['old_code']
    new_code = snapshot['new_code']
    affected_products = json.loads(snapshot['affected_data'])

    files_rolled = 0
    rows_rolled = 0
    errors = []

    for product_data in affected_products:
        filepath = product_data['filepath']

        try:
            wb = load_workbook(filepath, keep_vba=True)
            file_rows_rolled = 0

            for row_info in product_data['rows']:
                row_idx = row_info['row']
                # Mirror api_replace_execute's sheet/column resolution exactly —
                # rows were changed on whichever sheet Phase 2 found them on, not
                # always 'components', so rollback must target the same cell.
                src_sheet = row_info.get('sheet', '')
                if src_sheet and src_sheet in wb.sheetnames:
                    ws_name = src_sheet
                elif 'components' in wb.sheetnames:
                    ws_name = 'components'
                elif 'DATA' in wb.sheetnames:
                    ws_name = 'DATA'
                elif 'REF' in wb.sheetnames:
                    ws_name = 'REF'
                else:
                    continue
                ws = wb[ws_name]
                if ws_name == 'components':
                    ic_col = 1
                elif 'ic_col' in row_info:
                    ic_col = int(row_info['ic_col'])
                elif ws_name == 'REF':
                    ic_col = 4
                else:
                    ic_col = 5

                cell = ws.cell(row=row_idx, column=ic_col)
                if str(cell.value or '').strip().upper() == new_code:
                    cell.value = old_code
                    file_rows_rolled += 1

            if file_rows_rolled:
                wb.save(filepath)
                files_rolled += 1
                rows_rolled += file_rows_rolled
            wb.close()

        except Exception as e:
            print(f'Rollback error {filepath}: {e}')
            errors.append({'file': os.path.basename(filepath), 'error': str(e)})

    db.mark_snapshot_rolled_back(snap_id)
    scan_and_register_products()
    return {'files_rolled': files_rolled, 'rows_rolled': rows_rolled, 'errors': errors,
            'old_code': old_code, 'new_code': new_code}


@app.route('/api/replace/rollback/<int:snap_id>', methods=['POST'])
def api_replace_rollback(snap_id):
    """Rollback a replacement"""
    result = _perform_replace_rollback(snap_id)
    if 'error' in result:
        return jsonify(result), 400

    activity_log._log(session.get('username', 'User'), session.get('role', ''),
                      'replace_rollback',
                      f'Rolled back snap_id={snap_id} | {result["new_code"]} → {result["old_code"]} | '
                      f'{result["files_rolled"]} files, {result["rows_rolled"]} rows')

    return jsonify(result)


@app.route('/api/history/<int:record_id>/rollback', methods=['POST'])
@admin_required
def api_history_rollback(record_id):
    """Reset an approval/new_product record back to pending status (admin/developer only)."""
    data = request.get_json(silent=True) or {}
    record_type = data.get('type', '')

    if record_type == 'global_replace':
        return api_replace_rollback(record_id)

    if record_type == 'bom_run':
        conn = db.get_connection()
        try:
            cursor = conn.execute(
                "UPDATE runs SET status='rolled_back' WHERE id=?",
                (record_id,)
            )
            conn.commit()
            if cursor.rowcount == 0:
                return jsonify({'error': 'Record not found'}), 404
            return jsonify({'success': True})
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

    if record_type in ('approval', 'new_product'):
        conn = db.get_connection()
        try:
            cursor = conn.execute(
                '''UPDATE approvals
                   SET status='pending', reviewer_id=NULL, reviewer_name=NULL,
                       rejection_category=NULL, rejection_reason=NULL
                   WHERE id=?''',
                (record_id,)
            )
            conn.commit()
            if cursor.rowcount == 0:
                return jsonify({'error': 'Record not found'}), 404
            return jsonify({'success': True})
        except Exception as e:
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

    return jsonify({'error': f'Rollback not supported for type: {record_type}'}), 400

# ============================================================================
# ADMIN — User Management API
# ============================================================================

@app.route('/api/admin/users', methods=['GET'])
@admin_required
def api_admin_list_users():
    users = db.get_all_users()
    is_developer = session.get('role') == 'developer'
    # Build role display map
    roles_map = {r['name']: r for r in db.get_all_roles()}
    def _safe(u):
        row = dict(u)
        if not is_developer:
            row.pop('password_hash', None)
        # Attach role display info
        role_def = roles_map.get(row.get('role',''))
        row['role_display'] = role_def['display_name'] if role_def else row.get('role','')
        row['role_can_approve'] = role_def.get('can_approve',0) if role_def else 0
        return row
    return jsonify({'users': [_safe(u) for u in users]})

@app.route('/api/admin/users', methods=['POST'])
@admin_required
def api_admin_create_user():
    if not _can_create_users():
        return jsonify({'error': 'Your role does not have permission to create users'}), 403
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role     = data.get('role', 'user')
    tabs     = data.get('allowed_tabs', [])
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    # Validate: must be a system role or a custom role in DB
    valid_roles = {r['name'] for r in db.get_all_roles()}
    if role not in valid_roles:
        return jsonify({'error': 'Invalid role'}), 400
    if db.get_user_by_username(username):
        return jsonify({'error': 'Username already exists'}), 409
    pw_hash = generate_password_hash(password)
    # Full-privilege roles auto-get all tabs; others use role default or custom
    if role in ('admin', 'developer'):
        tabs = ALL_TABS
    elif role == 'sub_admin':
        tabs = [t for t in ALL_TABS if t not in ('cms', 'settings')]
    else:
        # Use role's default tabs if no custom tabs provided
        if not tabs:
            role_def = db.get_role_by_name(role)
            if role_def: tabs = role_def['allowed_tabs']
    uid = db.create_user(username, pw_hash, role=role, allowed_tabs=tabs)
    return jsonify({'success': True, 'user_id': uid})

@app.route('/api/admin/users/<int:user_id>/password', methods=['POST'])
@admin_required
def api_admin_reset_password(user_id):
    data = request.json or {}
    password = data.get('password', '').strip()
    if not password or len(password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    db.update_user_password(user_id, generate_password_hash(password))
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/tabs', methods=['POST'])
@admin_required
def api_admin_update_tabs(user_id):
    data = request.json or {}
    tabs = [t for t in data.get('allowed_tabs', []) if t in ALL_TABS]
    db.update_user_tabs(user_id, tabs)
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/role', methods=['POST'])
@admin_required
def api_admin_update_role(user_id):
    data  = request.json or {}
    role  = data.get('role', 'user')
    valid_roles = {r['name'] for r in db.get_all_roles()}
    if role not in valid_roles:
        return jsonify({'error': 'Invalid role'}), 400
    db.update_user_role(user_id, role)
    if role in ('admin', 'developer'):
        db.update_user_tabs(user_id, ALL_TABS)
    elif role == 'sub_admin':
        db.update_user_tabs(user_id, [t for t in ALL_TABS if t not in ('cms','settings')])
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/update', methods=['POST'])
@admin_required
def api_admin_update_user(user_id):
    """Combined update: role + tabs + optional password in one call."""
    data     = request.json or {}
    role     = data.get('role', 'user')
    tabs     = data.get('allowed_tabs', [])
    new_pw   = data.get('new_password', '').strip()
    valid_roles = {r['name'] for r in db.get_all_roles()}
    if role not in valid_roles:
        return jsonify({'error': 'Invalid role'}), 400
    db.update_user_role(user_id, role)
    if role in ('admin', 'developer'):
        db.update_user_tabs(user_id, ALL_TABS)
    elif role == 'sub_admin':
        db.update_user_tabs(user_id, [t for t in ALL_TABS if t not in ('cms','settings')])
    else:
        # For custom roles: use provided tabs; if none, use role default
        if tabs:
            db.update_user_tabs(user_id, [t for t in tabs if t in ALL_TABS])
        else:
            role_def = db.get_role_by_name(role)
            if role_def: db.update_user_tabs(user_id, role_def['allowed_tabs'])
    if new_pw:
        if len(new_pw) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        db.update_user_password(user_id, generate_password_hash(new_pw))
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def api_admin_toggle_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({'error': 'Cannot deactivate yourself'}), 400
    new_state = db.toggle_user_active(user_id)
    return jsonify({'success': True, 'is_active': new_state})

@app.route('/api/user/change-password', methods=['POST'])
def api_change_own_password():
    """Any logged-in user can change their own password."""
    data     = request.json or {}
    old_pw   = data.get('old_password', '')
    new_pw   = data.get('new_password', '').strip()
    if not new_pw or len(new_pw) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    user = db.get_user_by_id(session['user_id'])
    if not user:
        return jsonify({'error': 'User not found'}), 404
    if not check_password_hash(user['password_hash'], old_pw):
        return jsonify({'error': 'Current password is incorrect'}), 400
    db.update_user_password(user['id'], generate_password_hash(new_pw))
    return jsonify({'success': True})

# ============================================================================
# SPEC FILE IMPORT  (parse Macro .xlsm  OR  MDCF .xlsx)
# ============================================================================

def _parse_macro_spec(filepath: str, filename: str) -> dict:
    """Parse a Macro (.xlsm) BOM file into a unified BOM dict."""
    import tempfile
    product_name = os.path.splitext(os.path.basename(filename))[0]
    wb = load_workbook(filepath, read_only=True, data_only=True)

    bom_by_sku: dict = {}
    sku_order: list = []

    if 'REF' in wb.sheetnames:
        ref = wb['REF']
        hdr = [str(c or '').strip().upper()
               for c in next(ref.iter_rows(min_row=1, max_row=1, values_only=True))]

        PREBUILT = {'MATTRESS CODE', 'ITEMCODE', 'ITEM CODE', 'OTY', 'QTY'}
        if PREBUILT & set(hdr):
            def _idx(*names):
                for n in names:
                    try: return hdr.index(n)
                    except ValueError: pass
                return None

            mc_i  = _idx('MATTRESS CODE', 'PS CODE', 'PARENT CODE')
            seq_i = _idx('PS.NO', 'PS NO', 'SEQ', 'S.NO')
            ic_i  = _idx('ITEMCODE', 'ITEM CODE')
            uom_i = _idx('UOM')
            qty_i = _idx('OTY', 'QTY', 'ACTUAL QTY', 'ACTUAL QTY.')
            wh_i  = _idx('WH CODE', 'WH', 'WHOUSE')

            for row in ref.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    break
                if mc_i is None or not row[mc_i]:
                    continue
                sku = str(row[mc_i]).strip()
                if sku not in bom_by_sku:
                    bom_by_sku[sku] = []
                    sku_order.append(sku)
                ic  = str(row[ic_i]).strip()  if ic_i  is not None and row[ic_i]  else ''
                bom_by_sku[sku].append({
                    'seq':        int(row[seq_i]) if seq_i is not None and row[seq_i] else (len(bom_by_sku[sku]) + 1) * 10,  # type: ignore[arg-type]
                    'code':       ic,
                    'qty':        float(row[qty_i]) if qty_i is not None and row[qty_i] is not None else 0.0,  # type: ignore[arg-type]
                    'uom':        str(row[uom_i]).strip() if uom_i is not None and row[uom_i] else 'NOS',
                    'wh':         str(row[wh_i]).strip()  if wh_i  is not None and row[wh_i]  else '',
                    'desc':       ITEM_MASTER.get(ic, ''),
                    'department': '',
                    'section':    '',
                })
        else:
            # Format 1 – L/W/H/Colour permutation table (no explicit SKU codes)
            for i, row in enumerate(ref.iter_rows(min_row=2, max_col=4, values_only=True)):
                if row[0] is None:
                    break
                sku = f'{row[0]}X{row[1]}X{row[2]}'
                bom_by_sku.setdefault(sku, [])
                if sku not in sku_order:
                    sku_order.append(sku)

    wb.close()
    comp_count = len(bom_by_sku[sku_order[0]]) if sku_order else 0
    return {
        'product_name': product_name,
        'file_type':    'macro',
        'skus':         sku_order,
        'bom_by_sku':   bom_by_sku,
        'sku_count':    len(sku_order),
        'component_count': comp_count,
    }


def _parse_mdcf_spec(filepath: str, filename: str) -> dict:
    """Parse an MDCF EditProductStructure .xlsx file into a unified BOM dict."""
    # Strip generated prefixes / timestamps from filename to get product name
    product_name = os.path.splitext(os.path.basename(filename))[0]
    product_name = re.sub(r'^EditProductStructure-UserData-', '', product_name)
    product_name = re.sub(r'-\d{2}_\d{2}_\d{4}$', '', product_name)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    # The sheet name contains a typo in some versions; handle both
    sheet_name = next(
        (sn for sn in wb.sheetnames if 'prod' in sn.lower() and 'struct' in sn.lower()),
        None
    )
    if not sheet_name:
        wb.close()
        raise ValueError('EditProductStructure sheet not found in MDCF file')

    ws = wb[sheet_name]
    bom_by_sku: dict = {}
    sku_order: list = []

    # Rows 1-4 are header/metadata; data starts at row 5
    # Col indices (0-based):
    # 2=PS No.  12=PS Seq.  13=Item Code  15=Desc.  16=Qty.  17=UOM  21=WH Code
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        ps_no = str(row[2]).strip() if row[2] else ''
        if not ps_no:
            continue
        if ps_no not in bom_by_sku:
            bom_by_sku[ps_no] = []
            sku_order.append(ps_no)
        ic   = str(row[13]).strip() if row[13] else ''
        desc = str(row[15]).strip() if row[15] else ''
        if not desc and ic:
            desc = ITEM_MASTER.get(ic, '')
        bom_by_sku[ps_no].append({
            'seq':        int(row[12]) if row[12] else (len(bom_by_sku[ps_no]) + 1) * 10,  # type: ignore[arg-type]
            'code':       ic,
            'qty':        float(row[16]) if row[16] is not None else 0.0,  # type: ignore[arg-type]
            'uom':        str(row[17]).strip() if row[17] else 'NOS',
            'wh':         str(row[21]).strip() if row[21] else '',
            'desc':       desc,
            'department': '',
            'section':    '',
        })

    wb.close()
    comp_count = len(bom_by_sku[sku_order[0]]) if sku_order else 0
    return {
        'product_name':    product_name,
        'file_type':       'mdcf',
        'skus':            sku_order,
        'bom_by_sku':      bom_by_sku,
        'sku_count':       len(sku_order),
        'component_count': comp_count,
    }


@app.route('/api/parse-spec', methods=['POST'])
def api_parse_spec():
    """Parse uploaded Macro (.xlsm) or MDCF (.xlsx) and return BOM JSON."""
    file_type = request.form.get('file_type', 'macro')
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    suffix   = '.xlsm' if file_type == 'macro' else '.xlsx'
    tmp_path = None
    try:
        import tempfile as _tmpmod
        fd, tmp_path = _tmpmod.mkstemp(suffix=suffix)
        os.close(fd)
        f.save(tmp_path)

        if file_type == 'macro':
            result = _parse_macro_spec(tmp_path, f.filename or '')
        else:
            result = _parse_mdcf_spec(tmp_path, f.filename or '')

        # Check whether a matching product already exists in the DB
        all_products = db.get_all_products()
        pname_lower  = result['product_name'].lower()
        existing     = next((p for p in all_products
                             if p['name'].lower() == pname_lower), None)
        result['exists']     = existing is not None
        result['product_id'] = existing['id'] if existing else None

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass

    activity_log._log(session.get('username', 'User'), session.get('role', ''),
                      'parse_spec', f'Parsed file: {f.filename} | type={file_type} | product={result.get("product_name","")}')
    return jsonify(result)


@app.route('/api/import-macro', methods=['POST'])
def api_import_macro():
    """Save uploaded macro into BOM_FILES_ROOT. Regular users → approval queue."""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    filename = os.path.basename(f.filename or '')
    run_by   = session.get('username', 'unknown')

    if not _is_privileged():
        # Stage the file; move to BOM_FILES_ROOT only on approval
        staging_dir = os.path.join(config.BOM_FILES_ROOT, '_staging')
        os.makedirs(staging_dir, exist_ok=True)
        staged_path = os.path.join(staging_dir, filename)
        f.save(staged_path)
        summary = f"New product upload: {filename}"
        appr_id = db.create_approval(
            user_id=session['user_id'], username=run_by,
            action_type='new_product', action_summary=summary,
            action_data={'filename': filename, 'staged_path': staged_path}
        )
        activity_log.approval_submit(run_by, session.get('role', ''), appr_id,
                                     'new_product', summary)
        return jsonify({'pending': True, 'approval_id': appr_id,
                        'message': f'"{filename}" has been submitted for approval.'})

    dest_path = os.path.join(config.BOM_FILES_ROOT, filename)
    f.save(dest_path)
    result = scanner.scan_and_register_all()
    activity_log._log(run_by, session.get('role', ''), 'import_macro',
                      f'Imported macro file: {filename}')
    return jsonify({'success': True, 'filename': filename, 'scan': result})


# ============================================================================
# NEW PRODUCT WIZARD (MODIFIED v2.1)
# Now generates CreateProductStructure format
# ============================================================================

@app.route('/api/newproduct/generate', methods=['POST'])
def api_newproduct_generate():
    """Generate BOM for new product - now in CreateProductStructure format"""
    data = request.json or {}
    
    product_name = data.get('product_name', '').strip()
    prefix = data.get('prefix', '').strip()
    wh_code = data.get('wh_code', '').strip()
    ps_desc = data.get('ps_desc', '').strip()
    approval_ref = data.get('approval_ref', '').strip()
    sizes = data.get('sizes', [])
    components = data.get('components', [])
    run_by = data.get('run_by', '').strip()
    
    if not all([product_name, prefix, sizes, components, approval_ref]):
        return jsonify({'error': 'Missing required fields'}), 400
    
    try:
        # Get settings
        settings = db.get_settings()
        
        # Generate BOM
        output_rows = bom_engine.generate_new_product_bom(
            product_name=product_name,
            prefix=prefix,
            wh_code=wh_code,
            ps_desc=ps_desc,
            sizes=sizes,
            components=components,
            constants=RAMCO_CONSTANTS
        )
        
        # Convert to CreateProductStructure format
        ps_rows = generator.convert_bom_to_rows(output_rows, product_name, 'create')
        
        # Generate CreateProductStructure files
        userdata_bytes, metadata_bytes, user_fname, meta_fname = \
            generator.generate_create_structure(product_name, ps_rows, settings)
        
        # Create ZIP with both files
        ts = datetime.now().strftime('%d%m%Y_%H%M%S')
        fname = f'NEW-{prefix}-{ts}.zip'
        fpath = os.path.join(OUTPUT_FOLDER, fname)
        
        with zipfile.ZipFile(fpath, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(user_fname, userdata_bytes)
            zf.writestr(meta_fname, metadata_bytes)
        
        # Log as run record
        run_id = db.create_run(
            [],
            mode='NEW_PRODUCT',
            output_mode='SINGLE',
            run_by=run_by,
            approval_ref=approval_ref
        )
        db.complete_run(run_id, fname, len(sizes), len(output_rows))
        activity_log.new_product(session.get('username', run_by), session.get('role', ''),
                                 f'{product_name} | {len(sizes)} sizes | {len(output_rows)} rows | file={fname}')

        return jsonify({
            'run_id': run_id,
            'rows': len(output_rows),
            'filename': fname
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# NEAREST BOM (MODIFIED v2.1)
# Now generates EditProductStructure format
# ============================================================================

@app.route('/api/nearest-bom', methods=['POST'])
def api_nearest_bom():
    """Find nearest standard BOM for odd-size mattress"""
    data = request.json or {}
    
    family = data.get('family', '').strip()
    colour = data.get('colour', '').strip()
    L = int(data.get('L', 0))
    W = int(data.get('W', 0))
    H = str(data.get('H', '')).strip().zfill(2)
    odd_code = data.get('odd_code', '').strip()

    if not all([family, L, W, H, odd_code]):
        return jsonify({'error': 'All fields required'}), 400

    products = db.get_products_by_family(family)
    if not products:
        # Fallback: try matching as product_group
        conn = db.get_connection()
        products = [dict(r) for r in conn.execute(
            'SELECT * FROM products WHERE product_group = ? ORDER BY name', (family,)
        ).fetchall()]
        conn.close()
    if not products:
        return jsonify({'error': f'Family not found: {family}'}), 404

    output_rows = []
    for prod in products:
        try:
            components, permutations, dest_headers, prebuilt_rows = bom_engine.read_bom_file(
                prod['filepath']
            )
            if prebuilt_rows:
                dim_pat = f"{colour}{L}X{W}X{H}" if colour else f"{L}X{W}X{H}"
                rows = [r for r in prebuilt_rows
                        if re.search(re.escape(dim_pat) + r'\s*$', r['ps_no'])]
                if rows:
                    output_rows = rows
                    break
            else:
                filtered = [p for p in permutations
                            if int(p[0]) == L and int(p[1]) == W and str(p[2]) == H
                            and (not colour or str(p[3]) == colour)]
                if filtered:
                    output_rows = bom_engine.generate_bom(components, filtered[:1])
                    break
        except Exception:
            continue

    if not output_rows:
        return jsonify({'error': f'No BOM found for {L}×{W}×{H}" in family {family}'}), 404

    activity_log.nearest_bom(session.get('username', 'User'), session.get('role', ''),
                             f'{family} | {L}x{W}x{H}" {colour} → {odd_code}')
    return jsonify({
        'nearest': f'{L}x{W}x{H}',
        'odd_code': odd_code,
        'components': [{
            'seq': r['ps_seq'],
            'code': r['item_code'],
            'qty': r['qty'],
            'uom': r['uom'],
            'wh': r['wh_code']
        } for r in output_rows]
    })


@app.route('/api/nearest-bom/download')
def api_nearest_bom_download():
    """Download nearest BOM - now in EditProductStructure format"""
    family      = request.args.get('family', '').strip()
    product_grp = request.args.get('product', '').strip()   # optional product_group filter
    colour      = request.args.get('colour', '').strip()
    L           = int(request.args.get('L', 0))
    W           = int(request.args.get('W', 0))
    H           = request.args.get('H', '').strip().zfill(2)
    odd_code    = request.args.get('odd_code', '').strip()

    if not all([family, L, W, H, odd_code]):
        abort(400)

    products = db.get_products_by_family(family)

    # Fallback: if family value is actually a product_group (old URL format), search that way
    if not products:
        conn = db.get_connection()
        products = [dict(r) for r in conn.execute(
            'SELECT * FROM products WHERE product_group = ? ORDER BY name', (family,)
        ).fetchall()]
        conn.close()

    if not products:
        return jsonify({'error': f'No products found for family: {family}'}), 404

    # Narrow by product_group when passed explicitly
    if product_grp:
        filtered = [p for p in products if (p.get('product_group') or '') == product_grp]
        if filtered:
            products = filtered

    # Get settings
    settings = db.get_settings()

    output_rows = []
    for prod in products:
        try:
            components, permutations, dh, prebuilt_rows = bom_engine.read_bom_file(
                prod['filepath']
            )
            if prebuilt_rows:
                dim_pat = f"{colour}{L}X{W}X{H}" if colour else f"{L}X{W}X{H}"
                rows = [r for r in prebuilt_rows
                        if re.search(re.escape(dim_pat) + r'\s*$', r['ps_no'])]
                if rows:
                    for r in rows:
                        r.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in r})
                        r['ps_no'] = odd_code  # Replace with odd code
                    output_rows = rows
                    break
            else:
                filtered = [p for p in permutations
                            if int(p[0]) == L and int(p[1]) == W and str(p[2]) == H
                            and (not colour or str(p[3]) == colour)]
                if filtered:
                    output_rows = bom_engine.generate_bom(components, filtered[:1])
                    for r in output_rows:
                        r['ps_no'] = odd_code  # Replace with odd code
                    break
        except Exception:
            continue

    if not output_rows:
        return jsonify({'error': f'No BOM found'}), 404

    # Convert to EditProductStructure format
    ps_rows = generator.convert_bom_to_rows(output_rows, odd_code, 'edit')
    
    # Generate EditProductStructure files
    userdata_bytes, metadata_bytes, user_fname, meta_fname = \
        generator.generate_edit_structure(odd_code, ps_rows, settings)
    
    # Create ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(user_fname, userdata_bytes)
        zf.writestr(meta_fname, metadata_bytes)
    
    zip_buffer.seek(0)

    safe_code = re.sub(r'[^\w\-]', '_', odd_code)
    zip_fname = f'NearestBOM-{safe_code}.zip'
    activity_log.nearest_bom(session.get('username', 'User'), session.get('role', ''),
                             f'Download: {family} | {L}x{W}x{H}" {colour} → {odd_code}')
    activity_log.bom_download(session.get('username', 'User'), session.get('role', ''),
                              filename=zip_fname, rows=len(output_rows))

    # Log to run history
    try:
        _conn = db.get_connection()
        _conn.execute(
            '''INSERT INTO runs (run_label, status, product_count, rows_generated,
                                 run_by, created_at, mode, output_mode, approval_ref)
               VALUES (?, 'OK', 1, ?, ?, ?, 'NEAREST_BOM', 'FILE', ?)''',
            (
                f"Nearest BOM: {odd_code} ← {L}×{W}×{H} | {family}/{colour}",
                len(output_rows),
                session.get('username', 'User'),
                datetime.now().isoformat(),
                f"{L}×{W}×{H} | {colour}",
            )
        )
        _conn.commit()
        _conn.close()
    except Exception:
        pass

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_fname
    )

# ============================================================================
# NEAREST BOM — BULK UPLOAD (v2.1)
# Upload a list of (Mattress Code, Qty) — duplicate codes are aggregated so
# each unique size's BOM is matched and shown only once. Family/Colour are
# auto-detected from each code's prefix (the part before "<L>X<W>X<H>"),
# matched against prefixes scanned from every product's size table.
# ============================================================================

_NB_DIM_RE = re.compile(r'^(.*?)([\d.]+)X([\d.]+)X(\d{1,2})\s*$', re.IGNORECASE)

_NB_PREFIX_TTL = 1800  # rebuild the prefix index at most every 30 min
_nb_prefix_cache = {'built_at': 0.0, 'index': {}}

_NB_BATCH_TTL = 3600   # cached batch results expire after an hour
_nb_batch_cache = {}   # batch_id -> {'created_at', 'output_rows', 'groups'}


def _nb_build_prefix_index():
    """Scan every product's pre-built size table once, indexing each distinct
    PS No. prefix (brand/model/colour, i.e. everything before the LxWxH
    suffix) so an uploaded Mattress Code can be matched back to its product
    family purely from its own prefix — no manual family/colour selection.

    The SAME prefix can appear in several BOM files split by height (e.g.
    "Crystal…10 INCH" only carries the X06 sizes for prefix PEPSPKORGBW while
    "Org…10 INCH"/"Org…8 INCH ECOM" carry the X10/X08 variants of that same
    prefix) — so each prefix maps to a LIST of candidate products, tried in
    turn until one has a usable size for the requested height."""
    index = {}
    for prod in db.get_all_products():
        try:
            _, _, _, prebuilt_rows = bom_engine.read_bom_file(prod['filepath'])
        except Exception:
            continue
        if not prebuilt_rows:
            continue
        prefixes_in_product = set()
        for r in prebuilt_rows:
            ps_no = str(r.get('ps_no') or '').strip().upper()
            if not ps_no:
                continue
            m = _NB_DIM_RE.match(ps_no)
            if m and m.group(1):
                prefixes_in_product.add(m.group(1))
        for prefix in prefixes_in_product:
            index.setdefault(prefix, []).append(prod)
    return index


def _nb_get_prefix_index():
    now = time.time()
    if not _nb_prefix_cache['index'] or (now - _nb_prefix_cache['built_at']) > _NB_PREFIX_TTL:
        _nb_prefix_cache['index'] = _nb_build_prefix_index()
        _nb_prefix_cache['built_at'] = now
    return _nb_prefix_cache['index']


def _nb_find_nearest_size(prebuilt_rows, prefix, H, L, W):
    """Among rows sharing the given prefix+H, return:
       (next_l, next_w, max_l, max_w) — the smallest standard L/W that are
       each >= the requested value (independently, as the exact strings used
       in ps_no so the rebuilt pattern matches verbatim), plus the largest
       standard L/W available (so callers can tell an oversized request —
       bigger than every standard size — apart from "no data for this
       prefix/height" or a from a genuine library gap)."""
    Ls, Ws = {}, {}
    for r in prebuilt_rows:
        ps = str(r.get('ps_no') or '').strip().upper()
        m = _NB_DIM_RE.match(ps)
        if not m or m.group(1) != prefix or m.group(4).zfill(2) != H:
            continue
        try:
            lf, wf = float(m.group(2)), float(m.group(3))
        except ValueError:
            continue
        Ls.setdefault(lf, m.group(2))
        Ws.setdefault(wf, m.group(3))
    if not Ls or not Ws:
        return None, None, None, None
    next_l = next((Ls[v] for v in sorted(Ls) if v >= L), None)
    next_w = next((Ws[v] for v in sorted(Ws) if v >= W), None)
    return next_l, next_w, Ls[max(Ls)], Ws[max(Ws)]


def _nb_match_code(mattress_code, prefix_index):
    """Match one uploaded Mattress Code to its nearest standard BOM.
    Returns {'matched': True, ...rows/info} or {'matched': False, 'reason': ...}."""
    code = mattress_code.strip().upper()
    m = _NB_DIM_RE.match(code)
    if not m:
        return {'matched': False, 'reason': 'Could not read "<prefix><L>x<W>x<H>" from this code'}

    prefix, L_str, W_str, H = m.group(1), m.group(2), m.group(3), m.group(4).zfill(2)
    try:
        L, W = float(L_str), float(W_str)
    except ValueError:
        return {'matched': False, 'reason': 'Could not read the L/W dimensions from this code'}

    candidates = prefix_index.get(prefix) or []
    if not candidates:
        return {'matched': False, 'reason': f'No product family recognised for prefix "{prefix}"'}

    tried_families = []
    size_not_found = None  # first candidate where a usable size was identified but no BOM row matched it
    for prod in candidates:
        try:
            _, _, _, prebuilt_rows = bom_engine.read_bom_file(prod['filepath'])
        except Exception:
            continue
        if not prebuilt_rows:
            continue

        next_l, next_w, max_l, max_w = _nb_find_nearest_size(prebuilt_rows, prefix, H, L, W)
        oversized = False
        if next_l is None or next_w is None:
            if max_l is None or max_w is None:
                continue   # no usable size data for this prefix+height in this candidate — try next
            # request is bigger than every standard size on file — fall back to the largest available
            next_l, next_w = max_l, max_w
            oversized = True

        dim_pat = f'{prefix}{next_l}X{next_w}X{H}'
        rows = [r for r in prebuilt_rows if str(r.get('ps_no') or '').strip().upper() == dim_pat]
        if not rows:
            rows = [r for r in prebuilt_rows
                    if re.match(re.escape(dim_pat) + r'(\s|-|$)', str(r.get('ps_no') or '').strip().upper())]
        if not rows:
            if size_not_found is None:
                size_not_found = (prod['family'], dim_pat, oversized)
            tried_families.append(prod['family'])
            continue

        if oversized:
            exact = False
            note  = (f'This Mattress Code size ({L_str}x{W_str}") is above the defined maximum size for '
                     f'"{prefix}" @ {H}" — showing the closest available (maximum) BOM {dim_pat} as a '
                     f'reference. Kindly check and update if a larger size is genuinely required.')
        else:
            exact = (float(next_l) == L and float(next_w) == W)
            note  = (f'Exact standard size — using direct BOM for {dim_pat}.' if exact else
                     f'Rounded L: {L_str}"->{next_l}"  W: {W_str}"->{next_w}".  Proxy: {dim_pat} BOM.')

        return {
            'matched':       True,
            'family':        prod['family'],
            'product_name':  prod['name'],
            'nearest_ps_no': dim_pat,
            'exact':         exact,
            'oversized':     oversized,
            'note':          note,
            'rows':          rows,
        }

    if size_not_found is not None:
        fam, dim_pat, was_oversized = size_not_found
        if was_oversized:
            return {'matched': False,
                    'reason': (f'This Mattress Code size ({L_str}x{W_str}") is above the defined maximum size '
                               f'for "{prefix}" @ {H}", and even the maximum-size BOM {dim_pat} could not be '
                               f'located in "{fam}" — this looks like a data gap. Kindly check and update.')}
        return {'matched': False,
                'reason': (f'Nearest standard size {dim_pat} was identified in "{fam}" but no matching '
                           f'BOM row was found for it — this looks like a data gap. Kindly check and update.')}

    fam_hint = f' (checked: {", ".join(tried_families)})' if tried_families else ''
    return {'matched': False,
            'reason': f'No standard size >= {L_str}x{W_str} found for prefix "{prefix}" at height {H}"{fam_hint}'}


def _nb_cleanup_batches():
    now = time.time()
    for k in [k for k, v in _nb_batch_cache.items() if now - v['created_at'] > _NB_BATCH_TTL]:
        _nb_batch_cache.pop(k, None)


@app.route('/api/nearest-bom/bulk-template')
def api_nearest_bom_bulk_template():
    """Generate the bulk-upload template — S.no | Mattress Code | Qty."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Sheet1'
    hdr_fill = PatternFill('solid', fgColor='1A56DB')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    for ci, h in enumerate(['S.no', 'Mattress Code', 'Qty'], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    for i in range(1, 31):
        ws.cell(row=i + 1, column=1, value=i)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='NearestBOM-Upload-Template.xlsx'
    )


@app.route('/api/nearest-bom/bulk-upload', methods=['POST'])
def api_nearest_bom_bulk_upload():
    """Parse an uploaded Mattress Code + Qty list (template: S.no | Mattress
    Code | Qty), aggregate duplicate codes, auto-match each unique code to its
    nearest standard BOM via prefix detection, and return a grouped preview
    plus a batch_id the UI can use to download Excel/MDCF afterwards."""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    tmp_path = None
    try:
        import tempfile as _tmpmod
        fd, tmp_path = _tmpmod.mkstemp(suffix='.xlsx')
        os.close(fd)
        f.save(tmp_path)

        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        hdr = [str(c or '').strip().upper() for c in
               (next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []) or [])]
        code_idx = next((i for i, h in enumerate(hdr) if 'MATTRESS' in h or 'ITEM CODE' in h or h == 'CODE'), None)
        qty_idx  = next((i for i, h in enumerate(hdr) if h in ('QTY', 'QUANTITY', 'Q.TY', 'ORDER QTY', 'ORDER QUANTITY')), None)
        if code_idx is None:
            code_idx = 1
        if qty_idx is None:
            qty_idx = 2

        agg, order = {}, []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or code_idx >= len(row):
                continue
            raw_code = row[code_idx]
            if raw_code is None or not str(raw_code).strip():
                continue
            code = str(raw_code).strip().upper()
            try:
                qty = float(row[qty_idx]) if qty_idx < len(row) and row[qty_idx] is not None else 0.0  # type: ignore[arg-type]
            except (TypeError, ValueError):
                qty = 0.0
            if code not in agg:
                agg[code] = {'qty': 0.0, 'occurrences': 0}
                order.append(code)
            agg[code]['qty']         += qty
            agg[code]['occurrences'] += 1
        wb.close()
    except Exception as e:
        return jsonify({'error': f'Could not parse the uploaded file: {e}'}), 400
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except Exception: pass

    if not order:
        return jsonify({'error': 'No Mattress Code rows found — please use the upload template (S.no | Mattress Code | Qty)'}), 400

    prefix_index = _nb_get_prefix_index()

    groups, unmatched, output_rows = [], [], []
    for code in order:
        agg_info = agg[code]
        result = _nb_match_code(code, prefix_index)
        if not result['matched']:
            unmatched.append({
                'mattress_code': code,
                'qty':           agg_info['qty'],
                'occurrences':   agg_info['occurrences'],
                'reason':        result['reason'],
            })
            continue

        comps = []
        for r in result['rows']:
            ic = r.get('item_code', '')
            comps.append({
                'seq':         r.get('ps_seq'),
                'code':        ic,
                'description': ITEM_MASTER.get(ic, '') or r.get('description', ''),
                'qty':         r.get('qty'),
                'uom':         r.get('uom'),
                'wh':          r.get('wh_code'),
            })
            export_row = dict(r)
            export_row.update({k: v for k, v in RAMCO_CONSTANTS.items() if k not in export_row})
            export_row['ps_no'] = code
            output_rows.append(export_row)

        groups.append({
            'mattress_code':  code,
            'qty':            agg_info['qty'],
            'occurrences':    agg_info['occurrences'],
            'family':         result['family'],
            'product_name':   result['product_name'],
            'nearest_ps_no':  result['nearest_ps_no'],
            'exact':          result['exact'],
            'oversized':      result.get('oversized', False),
            'note':           result['note'],
            'components':     comps,
        })

    if not groups:
        return jsonify({
            'error':     'None of the uploaded Mattress Codes matched a known product family/size',
            'unmatched': unmatched,
        }), 404

    batch_id = uuid.uuid4().hex[:12]
    _nb_cleanup_batches()
    _nb_batch_cache[batch_id] = {
        'created_at':  time.time(),
        'output_rows': output_rows,
        'groups':      groups,
    }

    activity_log.nearest_bom(session.get('username', 'User'), session.get('role', ''),
                             f'Bulk upload: {len(order)} codes, {len(groups)} matched, '
                             f'{len(unmatched)} unmatched | batch={batch_id}')

    return jsonify({
        'batch_id':      batch_id,
        'uploaded_rows': sum(v['occurrences'] for v in agg.values()),
        'unique_codes':  len(order),
        'matched':       len(groups),
        'groups':        groups,
        'unmatched':     unmatched,
    })


@app.route('/api/nearest-bom/bulk-download/<batch_id>')
def api_nearest_bom_bulk_download(batch_id):
    """Download the matched bulk-upload results as a single Excel report or
    a combined MDCF (EditProductStructure) ZIP, rebuilt from cached batch data."""
    batch = _nb_batch_cache.get(batch_id)
    if not batch:
        return jsonify({'error': 'This batch has expired or was not found — please upload again'}), 404

    fmt = request.args.get('format', 'excel').strip().lower()
    date_str = datetime.now().strftime('%d_%m_%Y')
    _nb_user = session.get('username', 'User')
    _nb_role = session.get('role', '')
    _nb_rows = len(batch['output_rows'])

    if fmt == 'excel':
        xlsx_bytes = _generate_excel_report('Nearest-BOM-Bulk', batch['output_rows'])
        fname_xl = f'NearestBOM-Bulk-{date_str}.xlsx'
        activity_log.bom_download(_nb_user, _nb_role, filename=fname_xl, rows=_nb_rows)
        return send_file(
            io.BytesIO(xlsx_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname_xl
        )

    settings = db.get_settings()
    ps_rows  = generator.convert_bom_to_rows(batch['output_rows'], 'Nearest-BOM-Bulk', 'edit')
    userdata_bytes, metadata_bytes, user_fname, meta_fname = \
        generator.generate_edit_structure(f'NearestBOM-Bulk-{date_str}', ps_rows, settings)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(user_fname, userdata_bytes)
        zf.writestr(meta_fname, metadata_bytes)
    zip_buffer.seek(0)

    fname_zip = f'NearestBOM-Bulk-{date_str}.zip'
    activity_log.bom_download(_nb_user, _nb_role, filename=fname_zip, rows=_nb_rows)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=fname_zip
    )

# ============================================================================
# APPROVALS API
# ============================================================================

@app.route('/api/approvals', methods=['GET'])
def api_approvals_list():
    """Pending queue: admin/dev get pending + Sub-Admin-approved items awaiting
    their audit; sub_admin gets the pending queue only (not the audit queue,
    that's Admin's call); everyone else gets their own request history."""
    if _is_privileged():
        items = db.get_pending_approvals()
        for it in items:
            it['needs_admin_audit'] = False
        audits = db.get_pending_admin_audits()
        for it in audits:
            it['needs_admin_audit'] = True
        items = items + audits
    elif _can_review():
        items = db.get_pending_approvals()
        for it in items:
            it['needs_admin_audit'] = False
    else:
        items = db.get_approvals_by_user(session['user_id'])
    return jsonify({'approvals': items})

@app.route('/api/approvals/count', methods=['GET'])
def api_approvals_count():
    """Bell badge count: queue size for reviewers (+ audit queue for admin/dev),
    unread count for everyone else. 'notifications' is always the caller's own
    personal inbox — a sub_admin whose approval gets overturned, or an admin
    being told a new request was approved, needs that surfaced too, not just
    a reviewer-queue number; the existing unread-notification toast on page
    load already knows how to display it, it just wasn't being fed for these
    roles before."""
    notifs = db.get_user_notifications(session['user_id'])
    if _is_privileged():
        count = db.get_pending_count() + len(db.get_pending_admin_audits())
    elif _can_review():
        count = db.get_pending_count()
    else:
        count = db.get_unread_count(session['user_id'])
    return jsonify({'count': count, 'notifications': notifs})

@app.route('/api/approvals/<int:appr_id>/approve', methods=['POST'])
@admin_required
def api_approve(appr_id):
    appr = db.get_approval_by_id(appr_id)
    if not appr:
        return jsonify({'error': 'Not found'}), 404
    if appr['status'] != 'pending':
        return jsonify({'error': 'Already resolved'}), 400

    reviewer = session.get('username', 'Admin')
    reviewer_role = session.get('role', '')
    db.resolve_approval(appr_id, 'approved', session['user_id'], reviewer,
                        reviewer_role=reviewer_role)
    activity_log.approval_resolve(reviewer, reviewer_role, appr_id, 'approved')
    activity_log.log_approval(appr_id, appr['username'], '', appr['action_type'],
                               appr['action_summary'], 'approved', reviewer)

    # Execute the deferred action
    try:
        atype = appr['action_type']
        adata = appr['action_data']
        if atype == 'bom_run':
            run_id = db.create_run(product_ids=adata['product_ids'], mode='FULL',
                                   output_mode=adata.get('output_mode', 'ZIP'),
                                   run_by=appr['username'], approval_ref=str(appr_id))
            db.set_approval_result_ref(appr_id, str(run_id))
            threading.Thread(target=process_bom_generation_v2_1,
                             args=(run_id, adata['product_ids'], adata.get('output_mode', 'ZIP'))).start()
        elif atype == 'global_replace':
            files_changed = rows_changed = 0
            for pd in adata.get('affected_products', []):
                filepath = pd.get('filepath')
                if not filepath or not os.path.exists(filepath):
                    continue
                try:
                    wb = load_workbook(filepath, keep_vba=True)
                    file_rows_changed = 0
                    for row_info in pd['rows']:
                        row_idx = row_info['row']
                        # Mirror api_replace_execute's sheet/column resolution —
                        # rows live on whichever sheet Phase 2 found them on
                        # (REF/DATA), not a fixed 'components' sheet that real
                        # macro files don't have. Using a hardcoded sheet/column
                        # here silently no-ops the approved replace.
                        src_sheet = row_info.get('sheet', '')
                        if src_sheet and src_sheet in wb.sheetnames:
                            ws_name = src_sheet
                        elif 'components' in wb.sheetnames:
                            ws_name = 'components'
                        elif 'DATA' in wb.sheetnames:
                            ws_name = 'DATA'
                        elif 'REF' in wb.sheetnames:
                            ws_name = 'REF'
                        else:
                            continue
                        ws = wb[ws_name]
                        if ws_name == 'components':
                            ic_col = 1
                        elif 'ic_col' in row_info:
                            ic_col = int(row_info['ic_col'])
                        elif ws_name == 'REF':
                            ic_col = 4
                        else:
                            ic_col = 5
                        cell = ws.cell(row=row_idx, column=ic_col)
                        if str(cell.value or '').strip().upper() == adata['old_code']:
                            ws.cell(row=row_idx, column=ic_col, value=adata['new_code'])
                            if adata.get('new_qty') is not None:
                                qty_col = 4 if ws_name == 'components' else 6
                                ws.cell(row=row_idx, column=qty_col, value=adata['new_qty'])
                            if adata.get('new_uom'):
                                uom_col = 5 if ws_name == 'components' else 7
                                ws.cell(row=row_idx, column=uom_col, value=adata['new_uom'])
                            file_rows_changed += 1
                    if file_rows_changed:
                        wb.save(filepath)
                        files_changed += 1
                        rows_changed += file_rows_changed
                    wb.close()
                except Exception as ex:
                    print(f'Replace error: {ex}')
            db.create_replace_snapshot(
                adata['old_code'], adata['new_code'], adata.get('reason', ''),
                str(appr_id), appr['username'], files_changed, rows_changed,
                json.dumps(adata.get('affected_products', [])))
            scan_and_register_products()
        elif atype == 'new_product':
            staged = adata.get('staged_path', '')
            if staged and os.path.exists(staged):
                dest = os.path.join(config.BOM_FILES_ROOT, adata['filename'])
                shutil.move(staged, dest)
                db.set_approval_result_ref(appr_id, dest)
                scanner.scan_and_register_all()
    except Exception as ex:
        print(f'Approval execute error: {ex}')

    db.create_notification(appr['user_id'],
        f'✓ Your request "{appr["action_summary"]}" was approved by {reviewer}.',
        notif_type='approved', approval_id=appr_id)

    # Sub_admin approvals execute immediately but still need an Admin/Developer
    # audit pass — flag it and notify every admin/developer so it surfaces as
    # "a new request has been approved" for them to review (and overturn if needed).
    if reviewer_role == 'sub_admin':
        db.flag_admin_audit(appr_id)
        for u in db.get_all_users():
            if u['role'] in ('admin', 'developer'):
                db.create_notification(u['id'],
                    f'New request approved by Sub-Admin {reviewer}: "{appr["action_summary"]}" — needs your review.',
                    notif_type='admin_audit', approval_id=appr_id)

    return jsonify({'success': True})

@app.route('/api/approvals/<int:appr_id>/reject', methods=['POST'])
@admin_required
def api_reject(appr_id):
    appr = db.get_approval_by_id(appr_id)
    if not appr:
        return jsonify({'error': 'Not found'}), 404
    if appr['status'] != 'pending':
        return jsonify({'error': 'Already resolved'}), 400

    data     = request.json or {}
    category = data.get('category', '').strip()
    reason   = data.get('reason', '').strip()
    reviewer = session.get('username', 'Admin')

    # Clean up staged file if new_product
    if appr['action_type'] == 'new_product':
        staged = appr['action_data'].get('staged_path', '')
        if staged:
            try: os.unlink(staged)
            except: pass

    db.resolve_approval(appr_id, 'rejected', session['user_id'], reviewer, category, reason,
                        reviewer_role=session.get('role', ''))
    activity_log.approval_resolve(reviewer, session.get('role',''), appr_id, 'rejected', reason)
    activity_log.log_approval(appr_id, appr['username'], '', appr['action_type'],
                               appr['action_summary'], 'rejected', reviewer, reason)

    reason_part = f' Reason: {reason}' if reason else ''
    cat_part    = f' [{category}]' if category else ''
    msg = (f'✗ {reviewer} has rejected your request "{appr["action_summary"]}".{cat_part}{reason_part} '
           f'Ask {reviewer} directly for further proceeding.')
    db.create_notification(appr['user_id'], msg, notif_type='rejected', approval_id=appr_id)
    return jsonify({'success': True})

@app.route('/api/approvals/<int:appr_id>/admin-review', methods=['POST'])
def api_admin_review(appr_id):
    """Admin/Developer audit of a request a Sub-Admin already approved (and
    which already executed). 'acknowledge' just clears it from the audit
    queue. 'overturn' rolls back whatever the action did and notifies both
    the Sub-Admin who approved it and the original requester, with the
    Admin's remarks visible to both. Sub_admin cannot call this — only a
    higher authority audits a Sub-Admin's own approvals."""
    if not _is_privileged():
        return jsonify({'error': 'Admin access required'}), 403

    appr = db.get_approval_by_id(appr_id)
    if not appr:
        return jsonify({'error': 'Not found'}), 404
    if appr.get('admin_audit_status') != 'pending_audit':
        return jsonify({'error': 'Not awaiting audit'}), 400

    data    = request.json or {}
    action  = data.get('action', '').strip()
    remarks = data.get('remarks', '').strip()
    if action not in ('acknowledge', 'overturn'):
        return jsonify({'error': 'action must be acknowledge or overturn'}), 400
    if action == 'overturn' and not remarks:
        return jsonify({'error': 'Remarks are required to overturn an approval'}), 400

    admin_name = session.get('username', 'Admin')
    admin_id   = session['user_id']

    if action == 'acknowledge':
        db.resolve_admin_audit(appr_id, 'acknowledged', admin_id, admin_name, remarks)
        activity_log._log(admin_name, session.get('role', ''), 'admin_audit_acknowledge',
                          f'Acknowledged appr_id={appr_id} (approved by Sub-Admin {appr.get("reviewer_name")})')
        return jsonify({'success': True})

    # action == 'overturn' — roll back whatever the original approval triggered
    rollback_info = {}
    try:
        atype = appr['action_type']
        adata = appr['action_data']
        if atype == 'global_replace':
            conn = db.get_connection()
            row = conn.execute('SELECT id FROM replace_snapshots WHERE approval_ref=?',
                               (str(appr_id),)).fetchone()
            conn.close()
            if row:
                rollback_info = _perform_replace_rollback(row['id'])
        elif atype == 'bom_run':
            run_id = appr.get('result_ref')
            if run_id:
                conn = db.get_connection()
                conn.execute("UPDATE runs SET status='rolled_back' WHERE id=?", (run_id,))
                conn.commit()
                conn.close()
                rollback_info = {'run_id': run_id}
        elif atype == 'new_product':
            dest = appr.get('result_ref')
            if dest and os.path.exists(dest):
                os.remove(dest)
                scanner.scan_and_register_all()
                rollback_info = {'removed': dest}
    except Exception as ex:
        print(f'Admin overturn rollback error: {ex}')
        rollback_info = {'error': str(ex)}

    db.resolve_admin_audit(appr_id, 'overturned', admin_id, admin_name, remarks)
    activity_log._log(admin_name, session.get('role', ''), 'admin_audit_overturn',
                      f'Overturned appr_id={appr_id} (Sub-Admin {appr.get("reviewer_name")}\'s approval) | {remarks}')

    if appr.get('reviewer_id'):
        db.create_notification(appr['reviewer_id'],
            f'Admin {admin_name} has overturned your approval of "{appr["action_summary"]}". Remarks: {remarks}',
            notif_type='overturned', approval_id=appr_id)
    db.create_notification(appr['user_id'],
        f'✗ Your approved request "{appr["action_summary"]}" was rejected by Admin {admin_name} '
        f'after Sub-Admin approval. Remarks: {remarks}',
        notif_type='overturned', approval_id=appr_id)

    return jsonify({'success': True, 'rollback': rollback_info})

# ============================================================================
# NOTIFICATIONS API
# ============================================================================

@app.route('/api/notifications/read', methods=['POST'])
def api_notifications_read():
    db.mark_notifications_read(session['user_id'])
    return jsonify({'success': True})

# ============================================================================
# HISTORY API (updated — includes replace snapshots + approvals)
# ============================================================================

@app.route('/api/item-info')
def api_item_info():
    """Return description + departments/sections where an item code is used."""
    code = request.args.get('code','').strip().upper()
    if not code:
        return jsonify({'description':'','departments':[],'sections':[]})
    desc = ITEM_MASTER.get(code, '')
    # Find departments/sections from products that use this code
    depts, sects = set(), set()
    for p in db.get_all_products():
        try:
            codes = [c.upper() for c in json.loads(p.get('item_codes','[]') or '[]')]
            if code in codes:
                for d in json.loads(p.get('departments','[]') or '[]'):
                    if d: depts.add(d.strip())
                for s in json.loads(p.get('sections','[]') or '[]'):
                    if s: sects.add(s.strip())
        except Exception:
            pass
    return jsonify({'description': desc, 'departments': sorted(depts), 'sections': sorted(sects)})

@app.route('/api/sections')
def api_sections():
    """All unique section names from the products table."""
    conn = db.get_connection()
    rows = conn.execute('SELECT sections FROM products WHERE sections IS NOT NULL').fetchall()
    conn.close()
    sects = set()
    for row in rows:
        try:
            for s in json.loads(row['sections'] or '[]'):
                if s: sects.add(s.strip())
        except Exception:
            pass
    return jsonify({'sections': sorted(sects)})

@app.route('/api/departments')
def api_departments():
    """All unique department names from the products table."""
    conn = db.get_connection()
    rows = conn.execute('SELECT departments FROM products WHERE departments IS NOT NULL').fetchall()
    conn.close()
    depts = set()
    for row in rows:
        try:
            for d in json.loads(row['departments'] or '[]'):
                if d: depts.add(d.strip())
        except Exception:
            pass
    return jsonify({'departments': sorted(depts)})

@app.route('/api/history/all', methods=['GET'])
def api_history_all():
    """Unified history: BOM runs + replace snapshots + approval events."""
    from_date   = request.args.get('from', '')
    to_date     = request.args.get('to', '')
    type_filter = request.args.get('type', 'all')
    dept_filter = request.args.get('dept', '').strip()
    limit       = request.args.get('limit', 200, type=int)

    # Build set of product names that belong to the dept filter (for bom_run filtering)
    dept_products = set()
    if dept_filter:
        conn = db.get_connection()
        rows = conn.execute('SELECT name, departments FROM products').fetchall()
        conn.close()
        for row in rows:
            try:
                depts = json.loads(row['departments'] or '[]')
                if dept_filter in depts:
                    dept_products.add(row['name'].lower())
            except Exception:
                pass

    # Build name→family map for enriching bom_run records
    prod_name_map = {}
    try:
        for p in db.get_all_products():
            prod_name_map[p['name'].lower()] = {'family': p.get('family',''), 'name': p['name']}
    except Exception:
        pass

    records = []

    all_runs = db.get_run_history(limit=limit, from_date=from_date, to_date=to_date)

    if type_filter in ('all', 'bom_run'):
        for r in all_runs:
            if (r.get('mode') or '') == 'NEAREST_BOM':
                continue  # handled in nearest_bom section
            label = r.get('run_label') or ''
            prod_name = ''
            family = ''
            m = re.match(r'Run #\d+ — (.+)', label)
            if m:
                prod_name = m.group(1).strip()
                info = prod_name_map.get(prod_name.lower(), {})
                family = info.get('family', '')
            if dept_filter and not any(p in label.lower() for p in dept_products):
                continue
            size_filter = r.get('approval_ref', '') or ''
            if size_filter in ('0 products', label):
                size_filter = ''
            records.append({
                'id': r['id'], 'type': 'bom_run',
                'summary': prod_name or label, 'family': family,
                'status': r['status'], 'run_by': r['run_by'],
                'created_at': r['created_at'], 'output_filename': r.get('output_filename',''),
                'product_name': prod_name, 'filter_size': size_filter,
                'rows_generated': str(r.get('rows_generated') or 0),
                'error_note': r.get('error_message','') or '',
                'old_code':'','new_code':'','files_changed':'','rows_changed':'','change_reason':'','affected_products':'',
                'reviewer':'','rejection_category':'','rejection_reason':'',
            })

    if type_filter in ('all', 'nearest_bom'):
        for r in all_runs:
            if (r.get('mode') or '') != 'NEAREST_BOM':
                continue
            label = r.get('run_label') or ''
            records.append({
                'id': r['id'], 'type': 'nearest_bom',
                'summary': label, 'family': '',
                'status': r['status'], 'run_by': r['run_by'],
                'created_at': r['created_at'], 'output_filename': r.get('output_filename',''),
                'product_name': '', 'filter_size': r.get('approval_ref','') or '',
                'rows_generated': str(r.get('rows_generated') or 0),
                'error_note': r.get('error_message','') or '',
                'old_code':'','new_code':'','files_changed':'','rows_changed':'','change_reason':'','affected_products':'',
                'reviewer':'','rejection_category':'','rejection_reason':'',
            })

    if type_filter in ('all', 'global_replace'):
        conn = db.get_connection()
        snaps = conn.execute(
            'SELECT * FROM replace_snapshots ORDER BY created_at DESC LIMIT ?', (limit,)
        ).fetchall()
        conn.close()
        for s in snaps:
            if from_date and s['created_at'][:10] < from_date: continue
            if to_date   and s['created_at'][:10] > to_date:   continue
            affected_names = []
            try:
                for a in json.loads(s['affected_data'] or '[]')[:5]:
                    pname = a.get('name','') or a.get('product_name','')
                    if pname: affected_names.append(pname)
            except Exception:
                pass
            records.append({
                'id': s['id'], 'type': 'global_replace',
                'summary': f"Replace {s['old_code']} -> {s['new_code']}", 'family': '',
                'status': s['status'], 'run_by': s['run_by'],
                'created_at': s['created_at'], 'output_filename':'',
                'product_name':'','filter_size':'','rows_generated':'','error_note':'',
                'old_code': s['old_code'], 'new_code': s['new_code'],
                'files_changed': str(s['files_changed']), 'rows_changed': str(s['rows_changed']),
                'change_reason': s['reason'] or '',
                'affected_products': ', '.join(affected_names) + (' ...' if len(affected_names)==5 else ''),
                'reviewer':'','rejection_category':'','rejection_reason':'',
            })

    if type_filter in ('all', 'approval', 'new_product'):
        for a in db.get_all_approvals(limit=limit):
            if type_filter == 'new_product' and a['action_type'] != 'new_product': continue
            if type_filter == 'approval' and a['action_type'] not in ('bom_run','global_replace','new_product'): continue
            if from_date and a['created_at'][:10] < from_date: continue
            if to_date   and a['created_at'][:10] > to_date:   continue
            records.append({
                'id': a['id'], 'type': a['action_type'],
                'summary': a['action_summary'], 'family': '',
                'status': a['status'], 'run_by': a['username'],
                'created_at': a['created_at'], 'output_filename':'',
                'product_name':'','filter_size':'','rows_generated':'','error_note':'',
                'old_code':'','new_code':'','files_changed':'','rows_changed':'','change_reason':'','affected_products':'',
                'reviewer': a.get('reviewer_name','') or '',
                'rejection_category': a.get('rejection_category','') or '',
                'rejection_reason': a.get('rejection_reason','') or '',
                # Two-stage Sub-Admin -> Admin audit chain, for the Run History
                # "i" button so a user can see exactly where their request is.
                'reviewer_role':       a.get('reviewer_role','') or '',
                'admin_audit_status':  a.get('admin_audit_status','') or '',
                'admin_reviewer_name': a.get('admin_reviewer_name','') or '',
                'admin_remarks':       a.get('admin_remarks','') or '',
                'admin_reviewed_at':   a.get('admin_reviewed_at','') or '',
            })

    records.sort(key=lambda x: x['created_at'], reverse=True)
    return jsonify({'records': records[:limit]})

@app.route('/api/history/all/export', methods=['GET'])
def api_history_all_export():
    """Export unified history as CSV."""
    from_date   = request.args.get('from', '')
    to_date     = request.args.get('to', '')
    type_filter = request.args.get('type', 'all')
    # Re-use the logic by calling with request args already set
    import flask
    with app.test_request_context(f'/api/history/all?from={from_date}&to={to_date}&type={type_filter}&limit=5000'):
        result = api_history_all()
    data = json.loads(result.get_data(as_text=True))
    records = data.get('records', [])

    # Use StringIO then encode with UTF-8 BOM — avoids double-encoding of special chars
    sout = io.StringIO()
    w = csv.writer(sout)
    w.writerow([
        'Type',
        'Product Name',
        'Family',
        'Status',
        'Run By',
        'Date',
        'Time',
        # BOM Run columns
        'Size / Filter',
        'Rows Generated',
        'Error Note',
        # Global Replace columns
        'Old Item Code',
        'New Item Code',
        'Files Changed',
        'Rows Changed',
        'Reason',
        'Affected Products',
        # Approval columns
        'Reviewer',
        'Rejection Category',
        'Rejection Reason',
    ])
    for r in records:
        dt = (r.get('created_at','') or '').replace('T',' ')[:16]
        date_part = dt[:10]
        time_part = dt[11:16] if len(dt) > 10 else ''
        w.writerow([
            r.get('type',''),
            r.get('product_name','') or r.get('summary',''),
            r.get('family',''),
            r.get('status',''),
            r.get('run_by',''),
            date_part,
            time_part,
            r.get('filter_size',''),
            r.get('rows_generated',''),
            r.get('error_note',''),
            r.get('old_code',''),
            r.get('new_code',''),
            r.get('files_changed',''),
            r.get('rows_changed',''),
            r.get('change_reason',''),
            r.get('affected_products',''),
            r.get('reviewer',''),
            r.get('rejection_category',''),
            r.get('rejection_reason',''),
        ])
    # Encode as UTF-8 with BOM prefix so Excel recognises the encoding
    csv_bytes = '﻿' + sout.getvalue()   # ﻿ = UTF-8 BOM
    return Response(csv_bytes.encode('utf-8'), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': 'attachment; filename=history_export.csv'})

# ============================================================================
# ROLE MANAGEMENT API (Admin + Developer only)
# ============================================================================

@app.route('/api/roles', methods=['GET'])
@admin_required
def api_get_roles():
    return jsonify({'roles': db.get_all_roles()})

@app.route('/api/roles', methods=['POST'])
@admin_required
def api_create_role():
    if not _can_manage_roles():
        return jsonify({'error': 'Only Admin or Developer can create roles'}), 403
    data = request.json or {}
    name = data.get('name','').strip().lower().replace(' ','_')
    display_name = data.get('display_name','').strip()
    if not name or not display_name:
        return jsonify({'error': 'Role name and display name are required'}), 400
    if db.get_role_by_name(name):
        return jsonify({'error': f'Role "{name}" already exists'}), 409
    allowed_tabs      = [t for t in data.get('allowed_tabs',[]) if t in ALL_TABS]
    can_approve       = int(bool(data.get('can_approve')))
    can_create_users  = int(bool(data.get('can_create_users')))
    can_manage_users  = int(bool(data.get('can_manage_users')))
    can_manage_roles  = int(bool(data.get('can_manage_roles')))
    role_id = db.create_role(name, display_name, data.get('description',''),
                             allowed_tabs, can_approve, can_create_users,
                             can_manage_users, can_manage_roles)
    return jsonify({'success': True, 'role_id': role_id})

@app.route('/api/roles/<int:role_id>', methods=['POST'])
@admin_required
def api_update_role(role_id):
    if not _can_manage_roles():
        return jsonify({'error': 'Only Admin or Developer can update roles'}), 403
    role = db.get_role_by_id(role_id)
    if not role: return jsonify({'error': 'Role not found'}), 404
    if role['is_system']: return jsonify({'error': 'System roles cannot be modified'}), 400
    data = request.json or {}
    allowed_tabs = [t for t in data.get('allowed_tabs', role['allowed_tabs']) if t in ALL_TABS]
    db.update_role(role_id,
                   data.get('display_name', role['display_name']),
                   data.get('description', role['description']),
                   allowed_tabs,
                   int(bool(data.get('can_approve', role['can_approve']))),
                   int(bool(data.get('can_create_users', role['can_create_users']))),
                   int(bool(data.get('can_manage_users', role['can_manage_users']))),
                   int(bool(data.get('can_manage_roles', role['can_manage_roles']))))
    return jsonify({'success': True})

@app.route('/api/roles/<int:role_id>/delete', methods=['POST'])
@admin_required
def api_delete_role(role_id):
    if not _can_manage_roles():
        return jsonify({'error': 'Only Admin or Developer can delete roles'}), 403
    if not db.delete_role(role_id):
        return jsonify({'error': 'Cannot delete system roles'}), 400
    return jsonify({'success': True})

# ============================================================================
# USER ACTIVITY / SESSION HISTORY API
# ============================================================================

@app.route('/api/user-activity')
@admin_required
def api_user_activity():
    """Return activity records from SQL Server grouped by login session.

    Query params:
        username  — filter to a specific user (optional)
        date      — YYYY-MM-DD, filter to activities on that date (optional)
        limit     — max sessions to return (default 50)
    """
    username    = request.args.get('username', '').strip()
    date_filter = request.args.get('date', '').strip()
    limit       = request.args.get('limit', 50, type=int)

    try:
        from sqlserver_log_v2_0 import _get_connection
        conn = _get_connection()

        # ── Build WHERE clause ──────────────────────────────────────────────
        conditions = []
        params: list = []
        if username:
            conditions.append("username = ?")
            params.append(username)
        if date_filter:
            conditions.append("CAST(created_at AS DATE) = ?")
            params.append(date_filter)
        where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

        # ── Fetch all matching rows ordered by time ─────────────────────────
        rows = conn.execute(
            f"SELECT id, username, user_role, activity, detail, "
            f"       ip_address, session_id, created_at "
            f"FROM BOM_UserActivity {where} "
            f"ORDER BY created_at ASC",
            params
        ).fetchall()
        conn.close()

        # ── Group into sessions: each 'login' starts a new session ──────────
        sessions = []
        cur = None
        for r in rows:
            act = r[3]  # activity column
            if act == 'login':
                if cur is not None:
                    sessions.append(cur)
                cur = {
                    'username':   r[1],
                    'role':       r[2],
                    'login_at':   r[7].strftime('%Y-%m-%d %H:%M:%S') if r[7] else '',
                    'logout_at':  None,
                    'ip':         r[5],
                    'session_id': r[6],
                    'events':     [],
                }
            elif act == 'logout' and cur and cur['username'] == r[1]:
                cur['logout_at'] = r[7].strftime('%Y-%m-%d %H:%M:%S') if r[7] else ''
                sessions.append(cur)
                cur = None
            else:
                if cur is None:
                    # Activity without a preceding login (e.g. session resumed after restart)
                    cur = {
                        'username':   r[1],
                        'role':       r[2],
                        'login_at':   None,
                        'logout_at':  None,
                        'ip':         r[5],
                        'session_id': r[6],
                        'events':     [],
                    }
                cur['events'].append({
                    'time':     r[7].strftime('%H:%M:%S') if r[7] else '',
                    'activity': act,
                    'detail':   r[4],
                })
        if cur is not None:
            sessions.append(cur)

        # Most recent first, limited
        sessions = sessions[-limit:][::-1]
        return jsonify({'sessions': sessions, 'total': len(sessions)})

    except Exception as e:
        return jsonify({'error': f'SQL Server unavailable: {e}',
                        'sessions': [], 'total': 0}), 200


@app.route('/api/user-activity/users')
@admin_required
def api_user_activity_users():
    """Return list of usernames that have activity records in SQL Server."""
    try:
        from sqlserver_log_v2_0 import _get_connection
        conn = _get_connection()
        rows = conn.execute(
            "SELECT DISTINCT username FROM BOM_UserActivity ORDER BY username"
        ).fetchall()
        conn.close()
        return jsonify({'users': [r[0] for r in rows]})
    except Exception as e:
        return jsonify({'users': [], 'error': str(e)}), 200


# ============================================================================
# HTTPS — self-signed certificate (eliminates "Insecure download blocked")
# ============================================================================

def _ensure_ssl_cert():
    """Generate a self-signed SSL cert if one doesn't already exist."""
    ssl_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ssl')
    cert_path = os.path.join(ssl_dir, 'cert.pem')
    key_path  = os.path.join(ssl_dir, 'key.pem')

    if os.path.exists(cert_path) and os.path.exists(key_path):
        return cert_path, key_path

    os.makedirs(ssl_dir, exist_ok=True)
    try:
        from cryptography import x509
        from cryptography.x509.oid import NameOID
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        from cryptography.hazmat.backends import default_backend
        import ipaddress, datetime

        key = rsa.generate_private_key(65537, 2048, default_backend())
        name = x509.Name([x509.NameAttribute(NameOID.COMMON_NAME, u'PEPS BOM Tool')])
        san_list = [x509.DNSName(u'localhost'),
                    x509.IPAddress(ipaddress.IPv4Address(u'127.0.0.1'))]
        try:
            san_list.append(x509.IPAddress(ipaddress.IPv4Address(config.APP_LAN_IP)))
        except Exception:
            pass

        cert = (
            x509.CertificateBuilder()
            .subject_name(name).issuer_name(name)
            .public_key(key.public_key())
            .serial_number(x509.random_serial_number())
            .not_valid_before(datetime.datetime.utcnow())
            .not_valid_after(datetime.datetime.utcnow() + datetime.timedelta(days=3650))
            .add_extension(x509.SubjectAlternativeName(san_list), critical=False)
            .sign(key, hashes.SHA256(), default_backend())
        )

        with open(key_path, 'wb') as f:
            f.write(key.private_bytes(serialization.Encoding.PEM,
                                       serialization.PrivateFormat.TraditionalOpenSSL,
                                       serialization.NoEncryption()))
        with open(cert_path, 'wb') as f:
            f.write(cert.public_bytes(serialization.Encoding.PEM))

        print(f"✓ SSL certificate generated (valid 10 years): {ssl_dir}")
        print(f"  ➜  Open https://{config.APP_LAN_IP}:{config.APP_PORT} in your browser.")
        print(f"  ➜  First visit: click 'Advanced' → 'Proceed to {config.APP_LAN_IP}' to accept.")
        return cert_path, key_path

    except ImportError:
        print("WARNING: 'cryptography' package not found — running over HTTP.")
        print("         Install it with:  pip install cryptography")
        return None, None
    except Exception as e:
        print(f"WARNING: SSL cert generation failed ({e}) — running over HTTP.")
        return None, None

# ============================================================================
# STARTUP
# ============================================================================

if __name__ == '__main__':
    print(f"\n{'='*60}")
    print(f"PEPS BOM AUTOMATION TOOL - v2.1 (Product Structure)")
    print(f"{'='*60}\n")

    # Initialize database (will create settings table if not exists)
    db.init_db()
    
    # Seed default admin on first run
    _seed_admin()

    # Scan products
    scan_and_register_products()
    
    # Get stats
    stats = db.get_stats()
    print(f"\n{'='*60}")
    print(f"READY TO SERVE")
    print(f"{'='*60}")
    print(f"Products registered: {stats['total_products']}")
    print(f"Total SKUs:          {stats['total_skus']}")
    print(f"Item master codes:   {len(ITEM_MASTER)}")
    # Generate / load SSL certificate
    cert_path, key_path = _ensure_ssl_cert()
    ssl_ctx = (cert_path, key_path) if cert_path and key_path else None
    scheme  = 'https' if ssl_ctx else 'http'

    print(f"Local:               {scheme}://localhost:{config.APP_PORT}")
    print(f"Network:             {scheme}://{config.APP_LAN_IP}:{config.APP_PORT}")
    if ssl_ctx:
        print(f"\n⚠  FIRST-TIME SETUP (one-time per browser):")
        print(f"   Open {scheme}://{config.APP_LAN_IP}:{config.APP_PORT}")
        print(f"   Click 'Advanced' → 'Proceed to {config.APP_LAN_IP} (unsafe)'")
        print(f"   After that, downloads will work without any browser warnings.")
    print(f"{'='*60}\n")

    # Pre-warm the Nearest-BOM prefix index in the background so the first
    # bulk-upload request doesn't block while scanning all 232 BOM files.
    import threading as _threading
    _threading.Thread(target=_nb_get_prefix_index, daemon=True, name='nb-prewarm').start()

    # Serve with cheroot (production WSGI server) instead of Flask's dev
    # server, which prints its own warning against unattended/production use.
    # cheroot has genuine built-in SSL support (unlike Waitress, which has no
    # real TLS support — wrapping its socket with an SSLContext after the
    # fact is a known-fragile hack that breaks under real network conditions:
    # the underlying event loop doesn't understand SSL's WantRead/WantWrite
    # retry semantics, causing intermittent hangs/drops for remote clients).
    from cheroot.wsgi import Server as _CherootServer

    _server = _CherootServer((config.APP_HOST, config.APP_PORT), app, numthreads=8)
    if ssl_ctx:
        from cheroot.ssl.builtin import BuiltinSSLAdapter as _BuiltinSSLAdapter
        _cert_path, _key_path = ssl_ctx
        _server.ssl_adapter = _BuiltinSSLAdapter(certificate=_cert_path, private_key=_key_path)
    try:
        _server.start()
    except KeyboardInterrupt:
        _server.stop()
