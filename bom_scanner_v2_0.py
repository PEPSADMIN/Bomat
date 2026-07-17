"""
PEPS BOM Automation Tool - Product Scanner
Version: 2.0
Date: 21 May 2026

Scans xlsm macro files and registers product metadata in database.
"""

import os
import re
import json
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from datetime import datetime
import config_v2_0 as config
from database_v2_1 import db


_SFG_SIZE_VARIETY_CACHE = None
_SFG_SIZE_VARIETY_MTIME = None
_SFG_BOX_RE = re.compile(r'\bBOX\b', re.IGNORECASE)
_SFG_AS_RE = re.compile(r'\bAS\b', re.IGNORECASE)


def _load_sfg_size_variety():
    """
    Load the SFG Size Variety master workbook (BOX + ASSEMBLY sheets) and
    cache the distinct Length/Width/Height lists per sheet. Re-reads the file
    whenever its mtime changes, so edits to the master (e.g. correcting a
    Length/Width) show up on the next request without a server restart.
    """
    global _SFG_SIZE_VARIETY_CACHE, _SFG_SIZE_VARIETY_MTIME
    path = getattr(config, 'SFG_SIZE_VARIETY_PATH', None)
    current_mtime = os.path.getmtime(path) if path and os.path.exists(path) else None
    if _SFG_SIZE_VARIETY_CACHE is not None and current_mtime == _SFG_SIZE_VARIETY_MTIME:
        return _SFG_SIZE_VARIETY_CACHE

    cache = {}
    if path and os.path.exists(path):
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in ('BOX', 'ASSEMBLY'):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            lengths, widths, heights = set(), set(), set()
            pairs = {}            # (l, w) -> 'standard' | 'inbetween' — union across all heights
            by_height = {}        # height -> {(l, w): cat} — the Length/Width grid valid AT that height
            for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                l, w, h, remark = row[0], row[1], row[2], row[3]
                if l is None and w is None and h is None:
                    continue
                if l is not None: lengths.add(l)
                if w is not None: widths.add(w)
                if h is not None: heights.add(h)
                if l is not None and w is not None:
                    r = str(remark or '').strip().lower()
                    if 'standard' in r:
                        cat = 'standard'
                    elif 'inbetween' in r or 'in between' in r:
                        cat = 'inbetween'
                    else:
                        cat = None
                    pairs.setdefault((l, w), cat)
                    if h is not None:
                        by_height.setdefault(h, {})[(l, w)] = cat
            cache[sheet_name] = {
                'lengths': sorted(lengths),
                'widths': sorted(widths),
                'heights': sorted(heights),
                'pairs': [{'l': l, 'w': w, 'cat': cat} for (l, w), cat in sorted(pairs.items())],
                # Some BOX heights use a different L/W grid than others (e.g. a
                # tighter clearance set at one specific height) — keep each
                # height's own valid grid so size pickers stay accurate.
                'grids': {
                    str(h): [{'l': l, 'w': w, 'cat': cat} for (l, w), cat in sorted(grid.items())]
                    for h, grid in sorted(by_height.items())
                },
            }
        wb.close()
    _SFG_SIZE_VARIETY_CACHE = cache
    _SFG_SIZE_VARIETY_MTIME = current_mtime
    return cache


def get_sfg_size_variety():
    """Public accessor for the cached SFG Size Variety master (BOX + ASSEMBLY sheets)."""
    return _load_sfg_size_variety()


def _sfg_size_sheet_for_filename(product_name: str):
    """Pick the BOX or ASSEMBLY master sheet for an SFG product based on its filename."""
    if _SFG_BOX_RE.search(product_name):
        return 'BOX'
    if _SFG_AS_RE.search(product_name):
        return 'ASSEMBLY'
    return None


class BOMScanner:
    """Scans BOM files and registers products"""

    def __init__(self, bom_files_root: str = None):
        """Initialize scanner"""
        self.bom_files_root = bom_files_root or config.BOM_FILES_ROOT
        self.valid_extensions = config.VALID_EXTENSIONS
    
    def scan_and_register_all(self) -> Dict:
        """
        Scan all xlsm files in BOM_FILES_ROOT and register in database.
        Phase 1: read all files (no DB writes) → Phase 2: single atomic transaction.
        This eliminates UNIQUE constraint and database-locked errors.

        Returns:
            Dict with scan statistics
        """
        print(f"\n{'='*60}")
        print(f"SCANNING BOM FILES: {self.bom_files_root}")
        print(f"{'='*60}\n")

        if not os.path.exists(self.bom_files_root):
            print(f"✗ BOM files folder not found: {self.bom_files_root}")
            return {'success': False, 'error': 'Folder not found'}

        # Collect xlsm files, skipping Excel temp lock files (~$...)
        xlsm_files = []
        seen_paths = set()
        for root, dirs, files in os.walk(self.bom_files_root):
            for file in files:
                if file.startswith('~$'):
                    continue
                if any(file.lower().endswith(ext.lower()) for ext in self.valid_extensions):
                    full_path = os.path.join(root, file)
                    if full_path not in seen_paths:
                        seen_paths.add(full_path)
                        xlsm_files.append(full_path)

        print(f"Found {len(xlsm_files)} xlsm files to scan\n")

        # ── Phase 1: read all files (no DB writes) ────────────────────────────
        all_metadata = []
        error_count = 0
        errors = []

        for idx, filepath in enumerate(xlsm_files, start=1):
            try:
                print(f"[{idx}/{len(xlsm_files)}] Scanning: {os.path.basename(filepath)}...", end=' ')
                metadata = self.extract_metadata(filepath)
                metadata['filepath'] = filepath
                all_metadata.append(metadata)
                print(f"OK ({metadata['sku_count']} SKUs, {metadata['component_count']} components)")
            except Exception as e:
                print(f"ERROR: {str(e)}")
                error_count += 1
                errors.append({'file': os.path.basename(filepath), 'error': str(e)})

        # ── Phase 2: single atomic transaction ────────────────────────────────
        # Prefer using the database helper if implemented; otherwise perform
        # an equivalent atomic delete+insert fallback to avoid upgrade/runtime
        # issues where the method might be missing on the `db` object.
        if hasattr(db, 'replace_all_products'):
            db.replace_all_products(all_metadata)
        else:
            # Fallback: open a connection and replace products in one txn
            conn = db.get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute('DELETE FROM products')
                for p in all_metadata:
                    cursor.execute('''
                        INSERT INTO products (
                            name, family, product_group, filepath, category, heights, lengths, widths,
                            colours, departments, sections, wh_codes, item_codes,
                            sku_count, component_count, status, last_scanned, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        p['name'], p['family'], p.get('product_group', ''), p['filepath'], p.get('category', ''),
                        json.dumps(p.get('heights', [])),
                        json.dumps(p.get('lengths', [])),
                        json.dumps(p.get('widths', [])),
                        json.dumps(p.get('colours', [])),
                        json.dumps(p.get('departments', [])),
                        json.dumps(p.get('sections', [])),
                        json.dumps(p.get('wh_codes', [])),
                        json.dumps(p.get('item_codes', [])),
                        p.get('sku_count', 0),
                        p.get('component_count', 0),
                        p.get('status', 'ok'),
                        datetime.now().isoformat(), datetime.now().isoformat()
                    ))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
        success_count = len(all_metadata)

        print(f"\n{'='*60}")
        print(f"SCAN COMPLETE")
        print(f"{'='*60}")
        print(f"Success: {success_count}")
        print(f"Errors:  {error_count}")

        if errors:
            print(f"\nERRORS:")
            for err in errors[:5]:
                print(f"  • {err['file']}: {err['error']}")

        return {
            'success': True,
            'total_files': len(xlsm_files),
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        }
    
    def _extract_metadata_flat_xlsx(self, filepath: str, wb) -> Dict:
        """Read flat SFG xlsx files (Bom Code/PS.No | Issue Item | Qty … format)."""
        filename     = os.path.basename(filepath)
        product_name = os.path.splitext(filename)[0]
        rel_parts    = os.path.relpath(filepath, self.bom_files_root).split(os.sep)
        category     = rel_parts[0] if len(rel_parts) > 1 else 'SFG'
        family       = rel_parts[1] if len(rel_parts) > 2 else product_name
        product_group= rel_parts[2] if len(rel_parts) > 3 else ''

        # Use first non-RULES sheet
        data_sheet = next((s for s in wb.sheetnames if s.upper() != 'RULES'), None)
        if not data_sheet:
            return {'name': product_name, 'family': family, 'product_group': product_group,
                    'category': category, 'heights': [], 'lengths': [], 'widths': [],
                    'colours': [], 'departments': [], 'sections': [], 'wh_codes': [],
                    'item_codes': [], 'sku_count': 0, 'component_count': 0, 'status': 'err'}

        ws  = wb[data_sheet]
        hdr = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(hdr)}

        if 'Bom Code/PS.No' not in col:
            raise ValueError(
                f"Not a flat SFG BOM file — missing 'Bom Code/PS.No' header "
                f"(found: {hdr[:5]})"
            )

        ps_idx    = col.get('Bom Code/PS.No', 0)
        wh_idx    = col.get('Wh Code', 3)
        issue_idx = col.get('Issue Item', 5)
        clr_idx   = col.get('Colour', 10)
        acct_idx  = col.get('Account Group', 9)
        stdwh_idx = col.get('Std Wh Code', 11)

        seen_ps, item_codes, wh_codes, colours, departments = set(), set(), set(), set(), set()
        component_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            ps = str(row[ps_idx] or '').strip() if ps_idx < len(row) else ''
            # Collect item_codes from ALL rows — material rows in BP/TUFTING have
            # empty Bom Code/PS.No but still carry Issue Item codes that must be indexed.
            if issue_idx < len(row) and row[issue_idx]:
                item_codes.add(str(row[issue_idx]).strip())
            if wh_idx < len(row) and row[wh_idx]:
                wh_codes.add(str(row[wh_idx]).strip())
            if stdwh_idx < len(row) and row[stdwh_idx]:
                wh_codes.add(str(row[stdwh_idx]).strip())
            if clr_idx < len(row) and row[clr_idx]:
                colours.add(str(row[clr_idx]).strip())
            if acct_idx < len(row) and row[acct_idx]:
                departments.add(str(row[acct_idx]).strip())
            if not ps:
                continue
            seen_ps.add(ps)
            component_count += 1

        sku_count = len(seen_ps)
        status    = 'ok' if sku_count > 0 else 'err'
        return {
            'name': product_name, 'family': family, 'product_group': product_group,
            'category': category, 'heights': [], 'lengths': [], 'widths': [],
            'colours': sorted(colours), 'departments': sorted(departments), 'sections': [],
            'wh_codes': sorted(wh_codes), 'item_codes': sorted(item_codes),
            'sku_count': sku_count, 'component_count': component_count, 'status': status,
        }

    def extract_metadata(self, filepath: str) -> Dict:
        """
        Extract metadata from a single xlsm file.

        Args:
            filepath: Path to xlsm file

        Returns:
            Dict with product metadata
        """
        wb = load_workbook(filepath, read_only=True, data_only=True)

        # Flat xlsx SFG files (Bom Code/PS.No format) — no REF/DATA sheets
        if filepath.lower().endswith('.xlsx') and \
                'REF' not in wb.sheetnames and 'DATA' not in wb.sheetnames:
            result = self._extract_metadata_flat_xlsx(filepath, wb)
            wb.close()
            return result
        
        # Extract product name from filename
        filename = os.path.basename(filepath)
        product_name = os.path.splitext(filename)[0]

        # Folder structure: <root>/<brand>/<family>/<product_group>/.../<file>
        # e.g. Source Macro/FG/Bonnel Macro/Caspio/file.xlsm
        #   → brand = "FG", family = "Bonnel Macro", product_group = "Caspio"
        rel_parts     = os.path.relpath(filepath, self.bom_files_root).split(os.sep)
        category      = rel_parts[0] if len(rel_parts) > 1 else 'FG'          # brand: FG / SFG
        family        = rel_parts[1] if len(rel_parts) > 2 else product_name  # e.g. Bonnel Macro
        product_group = rel_parts[2] if len(rel_parts) > 3 else ''            # e.g. Caspio
        
        # Count components and extract dept/section/wh_codes from DATA sheet
        component_count = 0
        departments = set()
        sections = set()
        wh_codes = set()
        item_codes = set()   # all unique item codes in this file
        if 'DATA' in wb.sheetnames:
            data_sheet = wb['DATA']
            for row in data_sheet.iter_rows(min_row=2, max_col=10, values_only=True):
                if row[0] is None:
                    break
                if row[4]:
                    ic = str(row[4]).strip()
                    component_count += 1
                    item_codes.add(ic)
                if row[1]: departments.add(str(row[1]).strip())
                if row[2]: sections.add(str(row[2]).strip())
                if row[8]: wh_codes.add(str(row[8]).strip())
        
        # Extract permutations from REF sheet
        heights = set()
        lengths = set()
        widths = set()
        colours = set()
        sku_count = 0

        if 'REF' in wb.sheetnames:
            ref_sheet = wb['REF']
            hdr = [str(c or '').strip().upper()
                   for c in next(ref_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

            PREBUILT_MARKERS = {'MATTRESS CODE', 'ITEMCODE', 'ITEM CODE', 'OTY', 'QTY'}
            is_prebuilt = bool(PREBUILT_MARKERS & set(hdr))

            if is_prebuilt:
                # Format 2: extract L/W/H/colour from unique parent codes
                mc_idx = next((i for i, h in enumerate(hdr)
                               if h in {'MATTRESS CODE', 'PS CODE', 'PARENT CODE'}), None)
                # This format lists every component for EVERY built SKU (one row
                # per component per height/size), unlike the DATA sheet above
                # which only has a single representative SKU's components. Some
                # item codes (e.g. height-conditional ones) only ever appear for
                # specific SKUs here — collect them too, or Global Replace's
                # candidate index silently misses files that only have the code
                # in a non-default-height row.
                ic_idx = next((i for i, h in enumerate(hdr)
                               if h in {'ITEMCODE', 'ITEM CODE', 'ITEM_CODE'}), None)
                seen = set()
                for row in ref_sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        break
                    if ic_idx is not None and ic_idx < len(row) and row[ic_idx]:
                        item_codes.add(str(row[ic_idx]).strip())
                    if mc_idx is None or mc_idx >= len(row) or not row[mc_idx]:
                        continue
                    ps_no = str(row[mc_idx]).strip()
                    if ps_no in seen:
                        continue
                    seen.add(ps_no)
                    # Extract L×W×H — case-insensitive X, no end-anchor (handles -ECOM suffix)
                    m = re.search(r'(\d+)[Xx](\d+)[Xx](\d+)', ps_no)
                    if m:
                        lengths.add(m.group(1))
                        widths.add(m.group(2))
                        heights.add(m.group(3))
                    # Extract 2-4 letter colour code immediately before the dimensions
                    # e.g. BLOR72X48X08, PK72X48X08, BG72X48X08
                    cm = re.search(r'([A-Z]{2,4})\d+[Xx]\d+[Xx]\d+', ps_no)
                    if cm:
                        colours.add(cm.group(1))
                sku_count = len(seen)
                # component_count: unique rows per parent (use DATA sheet count)
            else:
                # Format 1: L, W, H, Colour are direct columns
                for row in ref_sheet.iter_rows(min_row=2, max_col=4, values_only=True):
                    if row[0] is None:
                        break
                    L = str(row[0]).strip()
                    W = str(row[1]).strip()
                    H = str(row[2]).strip()
                    C = str(row[3] or '').strip()
                    if L:
                        lengths.add(L)
                    if W:
                        widths.add(W)
                    if H:
                        heights.add(H)
                    if C:
                        colours.add(C)
                    sku_count += 1
        
        wb.close()

        # SFG products: the macro's own REF sheet only lists the sizes that
        # happen to be in that file. The actual valid size range for SFG comes
        # from the Size Variety master — BOX-named files use its BOX sheet,
        # AS-named files (assembly) use its ASSEMBLY sheet. Override the
        # REF-derived lists with that master data when the filename matches.
        out_heights = sorted(list(heights))
        out_lengths = sorted([int(x) for x in lengths if x.isdigit()])
        out_widths = sorted([int(x) for x in widths if x.isdigit()])
        if category.upper() == 'SFG':
            sheet_key = _sfg_size_sheet_for_filename(product_name)
            if sheet_key:
                variety = _load_sfg_size_variety().get(sheet_key)
                if variety:
                    out_heights = [str(h) for h in variety['heights']]
                    out_lengths = variety['lengths']
                    out_widths = variety['widths']

        # Determine status
        status = 'ok'
        if component_count == 0:
            status = 'warn'
        if sku_count == 0:
            status = 'err'

        return {
            'name': product_name,
            'family': family,
            'product_group': product_group,
            'category': category,
            'heights': out_heights,
            'lengths': out_lengths,
            'widths': out_widths,
            'colours': sorted(list(colours)),
            'departments': sorted(list(departments)),
            'sections': sorted(list(sections)),
            'wh_codes': sorted(list(wh_codes)),
            'item_codes': sorted(list(item_codes)),   # for description match stats
            'sku_count': sku_count,
            'component_count': component_count,
            'status': status
        }
    
    def get_product_families(self) -> List[str]:
        """Get unique list of product families"""
        products = db.get_all_products()
        families = sorted(list(set(p['family'] for p in products)))
        return families
    
    def get_family_metadata(self, family: str) -> Dict:
        """Get aggregated metadata for a family"""
        products = db.get_products_by_family(family)
        
        if not products:
            return None
        
        # Aggregate dimensions across all products in family
        all_heights = set()
        all_lengths = set()
        all_widths = set()
        all_colours = set()
        total_skus = 0
        
        for prod in products:
            if prod['heights']:
                all_heights.update(json.loads(prod['heights']))
            if prod['lengths']:
                all_lengths.update(json.loads(prod['lengths']))
            if prod['widths']:
                all_widths.update(json.loads(prod['widths']))
            if prod['colours']:
                all_colours.update(json.loads(prod['colours']))
            total_skus += prod['sku_count']
        
        return {
            'family': family,
            'product_count': len(products),
            'heights': sorted(list(all_heights)),
            'lengths': sorted([int(x) for x in all_lengths if str(x).isdigit()]),
            'widths': sorted([int(x) for x in all_widths if str(x).isdigit()]),
            'colours': sorted(list(all_colours)),
            'total_skus': total_skus
        }


# Singleton scanner instance
scanner = BOMScanner()
