"""
PEPS BOM Automation Tool - BOM Engine
Version: 2.0
Date: 21 May 2026

Core BOM generation logic:
- Read xlsm macro files
- Parse component rules and formulas
- Generate CreateProductStructure output (43 columns)
- New Product Wizard support
"""

import os
from typing import List, Dict, Tuple, Optional, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import WriteOnlyCell
import re
import config_v2_0 as config


class BOMEngine:
    """Core BOM generation engine"""
    
    def __init__(self):
        """Initialize BOM engine"""
        self.constants = config.RAMCO_CONSTANTS
    
    # ========================================================================
    # XLSM FILE READING
    # ========================================================================
    
    def _read_flat_xlsx_bom(self, filepath: str, wb) -> List[Dict]:
        """Read flat SFG xlsx (Bom Code/PS.No | Issue Item | Qty …) as prebuilt_rows."""
        data_sheet = next((s for s in wb.sheetnames if s.upper() != 'RULES'), None)
        if not data_sheet:
            return []
        ws  = wb[data_sheet]
        hdr = [str(c or '').strip()
               for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(hdr)}

        if 'Bom Code/PS.No' not in col:
            return []

        ps_idx    = col['Bom Code/PS.No']
        seq_idx   = col.get('PS Seq')
        desc_idx  = col.get('PS Desc.')
        plan_idx  = col.get('Process Plan')   # → section
        wh_idx    = col.get('Wh Code')
        issue_idx = col.get('Issue Item')
        idesc_idx = col.get('Item Desc')
        uom_idx   = col.get('Uom')
        qty_idx   = col.get('Qty')
        acct_idx  = col.get('Account Group')  # → department
        clr_idx   = col.get('Colour')         # → colour
        stdwh_idx = col.get('Std Wh Code')

        def _val(row, idx, default=''):
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return default

        rows = []
        last_ps = ''
        for row in ws.iter_rows(min_row=2, values_only=True):
            ps = _val(row, ps_idx)
            if ps:
                last_ps = ps
            effective_ps = ps or last_ps  # material rows carry parent PS.No
            item = _val(row, issue_idx)
            if not item:
                continue
            try:
                qty = float(_val(row, qty_idx, '0') or 0)
            except (ValueError, TypeError):
                qty = 0.0
            wh = _val(row, wh_idx) or _val(row, stdwh_idx)
            rows.append({
                'ps_no':       effective_ps,
                'ps_seq':      _val(row, seq_idx),
                'ps_desc':     _val(row, desc_idx),
                'item_code':   item,
                'description': _val(row, idesc_idx),
                'qty':         qty,
                'uom':         _val(row, uom_idx),
                'wh_code':     wh,
                'department':  _val(row, acct_idx),
                'section':     _val(row, plan_idx),  # Process Plan
                'colour':      _val(row, clr_idx),   # Colour
            })
        return rows

    def read_bom_file(self, filepath: str) -> Tuple[List[Dict], List[List], List[str], List[Dict]]:
        """
        Read BOM xlsm file and extract components and permutations.

        Two REF-sheet formats are auto-detected:
          Format 1 — permutation inputs: columns L, W, H, Colour  → permutations filled, prebuilt_rows empty
          Format 2 — pre-built BOM:      columns MATTRESS CODE, PS.NO, ITEMCODE, UOM, OTY, WH CODE
                                         → prebuilt_rows filled, permutations empty

        Returns:
            Tuple of:
            - components: List of component dicts with rules (from DATA sheet)
            - permutations: List of [L, W, H, Colour] (Format 1 only)
            - dest_headers: Destination column headers (if present)
            - prebuilt_rows: List of ready BOM row dicts (Format 2 only)
        """
        try:
            wb = load_workbook(filepath, read_only=True, data_only=False)

            # Flat SFG xlsx files (Bom Code/PS.No format) — no REF/DATA sheets
            if (filepath.lower().endswith('.xlsx') and
                    'REF' not in wb.sheetnames and 'DATA' not in wb.sheetnames):
                flat_rows = self._read_flat_xlsx_bom(filepath, wb)
                wb.close()
                return [], [], [], flat_rows

            # Read DATA sheet for component rules
            if 'DATA' not in wb.sheetnames:
                raise ValueError(f"DATA sheet not found in {filepath}")
            
            data_sheet = wb['DATA']
            components = []
            # item_code → {department, section} for Format-2 lookup
            _dept_map: Dict[str, Dict] = {}

            # DATA columns: Seq(0) Dept(1) Section(2) MattressCode(3) ItemCode(4)
            #               UOM(5) ActualQty(6) RoundOff(7) WHCode(8)
            # Build two lookups: by item_code AND by level number.
            # Level-based: DATA Level n matches REF PS.NO n*10 (e.g. Level 1 ↔ PS.NO 10)
            _level_dept_map: Dict[int, Dict] = {}
            _level_counter = 0
            for row in data_sheet.iter_rows(min_row=2, max_col=10, values_only=True):
                if row[0] is None:
                    break
                ic   = str(row[4] or '').strip()
                dept = str(row[1] or '').strip()
                sect = str(row[2] or '').strip()
                _level_counter += 1
                if ic:
                    _dept_map[ic] = {'department': dept, 'section': sect}
                _level_dept_map[_level_counter] = {'department': dept, 'section': sect}
                component = {
                    'seq':        row[0],
                    'item_code':  ic,
                    'description': str(row[5] or '').strip(),
                    'qty_formula': str(row[6] or '').strip(),
                    'uom':        str(row[7] or '').strip(),
                    'wh_code':    str(row[8] or '').strip(),
                    'department': dept,
                    'section':    sect,
                }
                if component['item_code']:
                    components.append(component)
            
            # Read REF sheet — auto-detect Format 1 (permutations) vs Format 2 (pre-built BOM)
            permutations = []
            prebuilt_rows = []
            if 'REF' in wb.sheetnames:
                ref_sheet = wb['REF']
                hdr = [str(c or '').strip().upper()
                       for c in next(ref_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

                # Format 2 detection: REF sheet already contains full BOM rows
                PREBUILT_MARKERS = {'MATTRESS CODE', 'ITEMCODE', 'ITEM CODE', 'OTY', 'QTY'}
                is_prebuilt = bool(PREBUILT_MARKERS & set(hdr))

                def _col(names):
                    for i, h in enumerate(hdr):
                        if h in names:
                            return i
                    return None

                def _get(row, idx, default=''):
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return default

                if is_prebuilt:
                    # Format 2: read pre-built BOM rows directly
                    mc_idx  = _col({'MATTRESS CODE', 'PS CODE', 'PARENT CODE'})
                    seq_idx = _col({'PS.NO', 'PS NO', 'SEQ', 'SEQ NO', 'SEQ.NO', 'SEQUENCE'})
                    ic_idx  = _col({'ITEMCODE', 'ITEM CODE', 'ITEM_CODE'})
                    uom_idx = _col({'UOM', 'UNIT'})
                    qty_idx = _col({'OTY', 'QTY', 'QUANTITY', 'QTY.'})
                    wh_idx  = _col({'WH CODE', 'WH_CODE', 'WAREHOUSE', 'WH'})

                    for row in ref_sheet.iter_rows(min_row=2, values_only=True):
                        if not any(row):
                            break
                        ps_no    = _get(row, mc_idx)
                        ps_seq   = _get(row, seq_idx)
                        item_code = _get(row, ic_idx)
                        uom      = _get(row, uom_idx)
                        qty_str  = _get(row, qty_idx, '0')
                        wh_code  = _get(row, wh_idx)

                        if not ps_no or not item_code:
                            continue
                        try:
                            qty = float(qty_str)
                        except (ValueError, TypeError):
                            qty = 0.0

                        # Lookup dept/section: try exact item_code first,
                        # then level-based (PS.NO/10 for 10-step sequences, PS.NO directly for 1-step)
                        _lkp = _dept_map.get(item_code)
                        if not _lkp:
                            try:
                                seq_int = int(float(ps_seq))
                                # Try /10 first (most common: PS.NO 10,20,30...)
                                _lkp = (_level_dept_map.get(seq_int // 10) or
                                        _level_dept_map.get(seq_int) or {})
                            except (ValueError, TypeError):
                                _lkp = {}
                        prebuilt_rows.append({
                            'ps_no':       ps_no,
                            'ps_seq':      ps_seq,
                            'ps_desc':     '',
                            'item_code':   item_code,
                            'description': '',
                            'qty':         qty,
                            'uom':         uom,
                            'wh_code':     wh_code,
                            'department':  _lkp.get('department', ''),
                            'section':     _lkp.get('section', ''),
                        })

                else:
                    # Format 1: permutation inputs (L, W, H, Colour)
                    L_NAMES = {'L', 'LENGTH', 'LEN'}
                    W_NAMES = {'W', 'WIDTH', 'WID'}
                    H_NAMES = {'H', 'HEIGHT', 'HGT', 'INCH', 'INCHES'}
                    C_NAMES = {'COLOUR', 'COLOR', 'CLR', 'C', 'COLOUR CODE', 'COLOR CODE'}

                    l_idx = _col(L_NAMES)
                    w_idx = _col(W_NAMES)
                    h_idx = _col(H_NAMES)
                    c_idx = _col(C_NAMES)

                    # If no recognisable headers, probe first data row for numeric L column
                    if None in (l_idx, w_idx, h_idx):
                        first = next(ref_sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
                        l_idx = 0
                        for i, v in enumerate(first):
                            if v is not None:
                                try:
                                    float(str(v))
                                    l_idx = i
                                    break
                                except (ValueError, TypeError):
                                    continue
                        w_idx = l_idx + 1
                        h_idx = l_idx + 2
                        c_idx = l_idx + 3

                    for row in ref_sheet.iter_rows(min_row=2, values_only=True):
                        if not any(row):
                            break
                        L = _get(row, l_idx)
                        W = _get(row, w_idx)
                        H = _get(row, h_idx)
                        C = _get(row, c_idx)
                        try:
                            float(L)
                            permutations.append([L, W, H, C])
                        except (ValueError, TypeError):
                            continue
            
            # Extract destination headers if present
            dest_headers = []
            if 'DEST' in wb.sheetnames:
                dest_sheet = wb['DEST']
                header_row = next(dest_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                dest_headers = [str(h or '').strip() for h in header_row]
            
            wb.close()
            return components, permutations, dest_headers, prebuilt_rows

        except Exception as e:
            raise Exception(f"Error reading BOM file {filepath}: {str(e)}")
    
    # ========================================================================
    # FORMULA EVALUATION
    # ========================================================================
    
    def evaluate_qty_formula(self, formula: str, L: float, W: float, H: str,
                            colour: str = '', components: List[Dict] = None) -> float:
        """
        Evaluate quantity formula with IF conditions and REF lookups.
        
        Args:
            formula: Formula string (e.g., "IF(H='6',0.5,0.75)" or "2.5")
            L, W, H: Dimensions
            colour: Colour code
            components: List of all components for REF lookups
        
        Returns:
            float: Evaluated quantity
        """
        try:
            # If it's a plain number, return it
            try:
                return float(formula)
            except ValueError:
                pass
            
            # Replace Excel references with Python equivalents
            formula = formula.upper()
            formula = formula.replace('L', str(L))
            formula = formula.replace('W', str(W))
            formula = formula.replace('H', f"'{H}'")
            
            # Handle IF statements
            # Pattern: IF(condition, true_value, false_value)
            if_pattern = r'IF\((.*?),(.*?),(.*?)\)'
            while 'IF(' in formula:
                match = re.search(if_pattern, formula)
                if match:
                    condition = match.group(1).strip()
                    true_val = match.group(2).strip()
                    false_val = match.group(3).strip()

                    # Convert Excel = to Python == (but not !=, <=, >=)
                    py_condition = re.sub(r'(?<![!<>=])=(?![=])', '==', condition)

                    try:
                        result = eval(py_condition)
                        formula = formula.replace(match.group(0), true_val if result else false_val)
                    except Exception:
                        formula = formula.replace(match.group(0), false_val)
                else:
                    break
            
            # Evaluate final expression
            result = eval(formula)
            return float(result)
        
        except Exception as e:
            # If evaluation fails, try to extract a number
            numbers = re.findall(r'\d+\.?\d*', formula)
            if numbers:
                return float(numbers[0])
            return 0.0
    
    # ========================================================================
    # BOM GENERATION
    # ========================================================================
    
    def generate_bom(self, components: List[Dict], permutations: List[List]) -> List[Dict]:
        """
        Generate BOM rows for given components and permutations.
        
        Args:
            components: List of component rules
            permutations: List of [L, W, H, Colour] combinations
        
        Returns:
            List of BOM row dictionaries (43 columns format)
        """
        output_rows = []
        
        for perm in permutations:
            L = float(perm[0])
            W = float(perm[1])
            H = str(perm[2])
            colour = str(perm[3] or '').strip()
            
            # Generate parent item code (e.g., COMBG-72X30X6)
            parent_code = f"PARENT-{int(L)}X{int(W)}X{H}"
            
            # Generate component rows for this permutation
            for comp in components:
                # Evaluate quantity
                qty = self.evaluate_qty_formula(
                    comp['qty_formula'], L, W, H, colour, components
                )
                
                # Create output row
                row = {
                    'ps_no':       parent_code,
                    'ps_seq':      comp['seq'],
                    'item_code':   comp['item_code'],
                    'description': comp['description'],
                    'qty':         qty,
                    'uom':         comp['uom'],
                    'wh_code':     comp['wh_code'],
                    'department':  comp.get('department', ''),
                    'section':     comp.get('section', ''),
                    'L': L,
                    'W': W,
                    'H': H,
                    'colour': colour,
                    # RAMCO constants
                    'maintained_by': self.constants['maintained_by'],
                    'ps_usage': self.constants['ps_usage'],
                    'ps_category': self.constants['ps_category'],
                    'default_for': self.constants['default_for'],
                    'io': self.constants['io'],
                    'valid_from': self.constants['valid_from'],
                    'valid_to': self.constants['valid_to'],
                    'min_qty': self.constants['min_qty'],
                    'max_qty': self.constants['max_qty'],
                    'stock_status': self.constants['stock_status'],
                    'critical': self.constants['critical'],
                    'use_up_option': self.constants['use_up_option'],
                    'process_plan_no': self.constants['process_plan_no'],
                }
                
                output_rows.append(row)
        
        return output_rows
    
    # ========================================================================
    # NEW PRODUCT WIZARD - GENERATE BOM
    # ========================================================================
    
    def generate_new_product_bom(self, product_name: str, prefix: str,
                                 wh_code: str, ps_desc: str, sizes: List[Dict],
                                 components: List[Dict], constants: Dict) -> List[Dict]:
        """
        Generate BOM for new product from wizard input.
        
        Args:
            product_name: Product name
            prefix: Item code prefix (e.g., 'COMBG')
            wh_code: Warehouse code for parent
            ps_desc: PS Description (can be auto-generated)
            sizes: List of size dicts with keys: l, w, h, c (colour)
            components: List of component dicts with keys: code, desc, qty, uom, wh
            constants: RAMCO constants dict
        
        Returns:
            List of BOM row dictionaries (43 columns format)
        """
        # Strip the trailing colour placeholder from prefix once (e.g. PEPSSPKBNLNM → PEPSSPKBNL)
        # The last 2-4 uppercase letters at the end of the prefix are treated as the colour code.
        # Each size row substitutes its own colour in that position.
        _pfx_m = re.match(r'^(.{4,})([A-Z]{2,4})$', prefix)
        _base_prefix = _pfx_m.group(1) if _pfx_m else prefix

        output_rows = []

        for size in sizes:
            L = int(size['l'])
            W = int(size['w'])
            H_padded = str(size['h']).zfill(2)   # zero-pad height: 6 → 06
            colour = str(size.get('c', '')).strip()

            # Build parent item code: {base_prefix}{colour}{L}X{W}X{H_padded}  (no hyphen)
            parent_code = f"{_base_prefix}{colour}{L}X{W}X{H_padded}"

            # Generate PS Description if not provided
            H_str = str(size['h'])
            colour_name = str(size.get('cn', colour)).strip()
            if not ps_desc or ps_desc == 'auto':
                if colour_name:
                    ps_desc_auto = f"{product_name} {colour_name} {L}X{W}X{H_str}"
                else:
                    ps_desc_auto = f"{product_name} {L}X{W}X{H_str}"
            else:
                ps_desc_auto = ps_desc.replace('{colour}', colour_name)
                ps_desc_auto = ps_desc_auto.replace('LxWxH', f"{L}X{W}X{H_str}")
                ps_desc_auto = ps_desc_auto.replace('L', str(L))
                ps_desc_auto = ps_desc_auto.replace('W', str(W))
                ps_desc_auto = ps_desc_auto.replace('H', H_str)

            # Use per-SKU components from the macro file when available (they carry
            # the correct quantity for this specific L×W×H×Colour combination).
            # Fall back to the generic wizard component list when not present.
            spec_comps = size.get('spec_comps')
            comps_for_size = spec_comps if spec_comps else components

            # Generate component rows — ps_seq from REF sheet seq when available,
            # otherwise multiples of 10 (10, 20, 30 …)
            for idx, comp in enumerate(comps_for_size, start=1):
                ps_seq = int(comp.get('seq', idx * 10)) if spec_comps else idx * 10
                row = {
                    'ps_no': parent_code,
                    'ps_desc': ps_desc_auto,
                    'ps_seq': ps_seq,
                    'item_code': comp.get('code', ''),
                    'description': comp.get('desc', ''),
                    'qty': float(comp.get('qty', 0)),
                    'uom': comp.get('uom', 'NOS'),
                    'wh_code': comp.get('wh', ''),
                    'parent_wh_code': wh_code,
                    'L': L,
                    'W': W,
                    'H': H_str,
                    'colour': colour,
                    # RAMCO constants
                    'maintained_by': constants['maintained_by'],
                    'ps_usage': constants['ps_usage'],
                    'ps_category': constants['ps_category'],
                    'default_for': constants['default_for'],
                    'io': constants['io'],
                    'valid_from': constants['valid_from'],
                    'valid_to': constants['valid_to'],
                    'min_qty': constants['min_qty'],
                    'max_qty': constants['max_qty'],
                    'stock_status': constants['stock_status'],
                    'critical': constants['critical'],
                    'use_up_option': constants['use_up_option'],
                    'process_plan_no': constants['process_plan_no'],
                }
                
                output_rows.append(row)
        
        return output_rows
    
    # ========================================================================
    # XLSX OUTPUT WRITING
    # ========================================================================
    
    def write_ramco_xlsx(self, rows: List[Dict], dest_headers: List[str] = None,
                        output_path: str = None, parent_override: str = None):
        """
        Write BOM rows to CreateProductStructure xlsx format (43 columns).
        Uses write_only (streaming) mode for fast output regardless of row count.
        """
        HEADERS = [
            'Rec. No.', 'PS No.', 'PS Seq.', 'P.S Desc.',
            'Valid From', 'Valid To', 'Maintained By', 'PS Usage',
            'PS Category', 'Parent WH Code', 'PS Type', 'Process Plan No.',
            'Mandatory', 'Default For', 'Seq. No.', 'Item Code',
            'Description', 'Qty.', 'UOM', 'WH Code',
            'Supply Type', 'Substitute', 'Substitute %', 'Sequence',
            'Co Product', 'Yield', 'Ref. Doc. No.', 'Ref. Doc. Date',
            'I/O', 'Valid From', 'Valid To', 'Min. Qty.',
            'Max. Qty.', 'Wastage %', 'Remarks', 'Stock Status',
            'Critical', 'Use-Up Option', 'Substitute Group', 'Alternate Seq.',
            'Cost Basis', 'Amortisation Qty', 'Customer Code',
        ]

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("CreateProductStructure")

        # Styled header row via WriteOnlyCell
        hdr_font = Font(bold=True)
        hdr_align = Alignment(horizontal='center')
        hdr_cells = []
        for h in HEADERS:
            c = WriteOnlyCell(ws, value=h)
            c.font = hdr_font
            c.alignment = hdr_align
            hdr_cells.append(c)
        ws.append(hdr_cells)

        # Data rows — one append() call per row (10-20× faster than ws.cell())
        for data_row in rows:
            parent_code = parent_override if parent_override else data_row.get('ps_no', '')
            ws.append([
                1,                                                              # 1  Rec. No.
                parent_code,                                                    # 2  PS No.
                '',                                                             # 3  PS Seq. (always blank)
                data_row.get('ps_desc', ''),                                   # 4  P.S Desc.
                data_row.get('valid_from', ''),                                # 5  Valid From
                data_row.get('valid_to', ''),                                  # 6  Valid To
                data_row.get('maintained_by', ''),                             # 7  Maintained By
                data_row.get('ps_usage', ''),                                  # 8  PS Usage
                data_row.get('ps_category', ''),                               # 9  PS Category
                data_row.get('parent_wh_code', data_row.get('wh_code', '')),  # 10 Parent WH Code
                '',                                                             # 11 PS Type
                data_row.get('process_plan_no', ''),                           # 12 Process Plan No.
                '',                                                             # 13 Mandatory
                data_row.get('default_for', ''),                               # 14 Default For
                data_row.get('ps_seq', ''),                                    # 15 Seq. No.
                data_row.get('item_code', ''),                                 # 16 Item Code
                data_row.get('description', ''),                               # 17 Description
                data_row.get('qty', 0),                                        # 18 Qty.
                data_row.get('uom', ''),                                       # 19 UOM
                data_row.get('wh_code', ''),                                   # 20 WH Code
                '', '', '', '', '', '', '', '',                                 # 21-28 (blank)
                data_row.get('io', ''),                                        # 29 I/O
                data_row.get('valid_from', ''),                                # 30 Valid From
                data_row.get('valid_to', ''),                                  # 31 Valid To
                data_row.get('min_qty', ''),                                   # 32 Min. Qty.
                data_row.get('max_qty', ''),                                   # 33 Max. Qty.
                '', '',                                                         # 34-35 (blank)
                data_row.get('stock_status', ''),                              # 36 Stock Status
                data_row.get('critical', ''),                                  # 37 Critical
                data_row.get('use_up_option', ''),                             # 38 Use-Up Option
                '', '', '', '', '',                                             # 39-43 (blank)
            ])

        wb.save(output_path)
        print(f"✓ Wrote {len(rows)} rows to {output_path}")


# Singleton engine instance
bom_engine = BOMEngine()
