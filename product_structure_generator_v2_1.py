"""
PEPS BOM Automation Tool - Product Structure Generator
Version: 2.1
Date: 22 May 2026

Builds CreateProductStructure and EditProductStructure Excel files
from scratch using openpyxl — no external template files required.

Generates paired UserData + MetaData files inside a ZIP.
"""

import io
import os
from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# Settings keys whose values must be stored as specific types in Excel
_INT_SETTINGS  = {'number_of_records', 'excel_start_row'}
_BOOL_SETTINGS = {'enable'}

def _coerce(key: str, val):
    """Convert setting string values to the types the RAMCO template expects."""
    if key in _INT_SETTINGS:
        try:
            return int(val)
        except (ValueError, TypeError):
            return val
    if key in _BOOL_SETTINGS:
        return str(val).strip() in ('1', 'true', 'True', 'TRUE')
    return val


# ---------------------------------------------------------------------------
# Column layouts
# ---------------------------------------------------------------------------

# Column maps verified against actual template sheets (row 4 field names)
_EDIT_COLS = {
    'rec_no':        (1,  'Rec.No'),
    'status_col':    (2,  'Status'),
    'ps_no':         (3,  'PS No.'),
    'ps_desc':       (4,  'PS Desc.'),
    'ps_item_code':  (5,  'Item Code'),      # parent item = same value as ps_no
    'maintained_by': (7,  'Maintained By'),
    'wh_code':       (8,  'Warehouse Code'), # parent WH
    'parent_qty':    (9,  'Parent Qty.'),
    'default_for':   (12, 'Default For'),
    'ps_seq':        (13, 'PS Seq.'),
    'item_code':     (14, 'Item Code'),      # component item
    'description':   (16, 'Desc.'),
    'qty':           (17, 'Qty.'),
    'uom':           (18, 'UOM'),
    'pp_seq':        (20, 'PP Seq.'),
    'wh_code_item':  (22, 'Warehouse Code'), # component WH
    'io':            (26, 'I/O'),
    'valid_from':    (27, 'Valid From'),
    'valid_to':      (28, 'Valid To'),
    'min_qty':       (29, 'Min. Qty.'),
    'max_qty':       (30, 'Max. Qty.'),
    'stock_status':  (33, 'Stock Status'),
    'critical':      (34, 'Critical'),
    'use_up_option': (35, 'Use-Up Option'),
}

_CREATE_COLS = {
    'rec_no':        (1,  'Rec.No'),
    'status_col':    (2,  'Status'),
    'ps_no':         (3,  'PS No.'),
    'ps_desc':       (4,  'PS Desc.'),
    'ps_item_code':  (5,  'Item Code'),      # parent item = same value as ps_no
    'maintained_by': (7,  'Maintained By'),
    'ps_usage':      (8,  'PS Usage'),
    'ps_category':   (9,  'PS Category'),
    'wh_code':       (10, 'Warehouse Code'), # parent WH
    'parent_qty':    (11, 'Parent Qty.'),
    'default_for':   (14, 'Default For'),
    'ps_seq':        (16, 'PS Seq.'),
    'item_code':     (17, 'Item Code'),      # component item
    'description':   (19, 'Desc.'),
    'qty':           (20, 'Qty.'),
    'uom':           (21, 'UOM'),
    'pp_seq':        (23, 'PP Seq.'),
    'wh_code_item':  (25, 'Warehouse Code'), # component WH
    'io':            (29, 'I/O'),
    'valid_from':    (30, 'Valid From'),
    'valid_to':      (31, 'Valid To'),
    'min_qty':       (32, 'Min. Qty.'),
    'max_qty':       (33, 'Max. Qty.'),
    'stock_status':  (36, 'Stock Status'),
    'critical':      (37, 'Critical'),
    'use_up_option': (38, 'Use-Up Option'),
}

_SETTINGS_MAP = [
    (11, 'Url',                   'url'),
    (12, 'User Id',               'user_id'),
    (13, 'Password',              'password'),
    (14, 'Role Name',             'role_name'),
    (15, 'Organisation Unit',     'organisation_unit'),
    (16, 'LanguageId',            'language_id'),
    (17, 'Number of Records',     'number_of_records'),
    (18, 'Bpc',                   'bpc'),
    (19, 'Bpc Name',              'bpc_name'),
    (20, 'Enable',                'enable'),
    (21, 'Version',               'version'),
    (22, 'User Interface Type',   'user_interface_type'),
    (23, 'ExcelData Start RowNo', 'excel_start_row'),
    (24, 'ExcelData Start ColNo', 'excel_start_col'),
    (25, 'Skip Blankvalues',      'skip_blank_values'),
    (26, 'Mdcf Version',          'mdcf_version'),
]


class ProductStructureGenerator:

    def __init__(self, template_dir: str = None):
        self.template_dir = template_dir or os.path.dirname(__file__)

    # ------------------------------------------------------------------ #
    # Internal builders                                                    #
    # ------------------------------------------------------------------ #

    def _wb_to_bytes(self, wb: Workbook) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _add_login_sheet(self, wb: Workbook, settings: Dict) -> None:
        ws = wb.create_sheet('Login')

        ws.merge_cells('B7:C9')
        ws['B7'] = 'RAMCO VIRTUALWORKS UpLoad Master Data Version 2.0'
        ws['B7'].font = Font(bold=True)
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('B10:C10')
        ws['B10'] = 'CONNECTION ATTRIBUTES (To App Server)'
        ws['B10'].font = Font(bold=True)
        ws['B10'].alignment = Alignment(horizontal='center')

        for row_num, label, key in _SETTINGS_MAP:
            ws[f'B{row_num}'] = label
            ws[f'C{row_num}'] = settings.get(key, '')

        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 55

    def _add_data_sheet(self, wb: Workbook, sheet_name: str,
                        col_layout: Dict, bom_rows: List[Dict]) -> None:
        ws = wb.create_sheet(sheet_name)

        bold = Font(bold=True)

        # Row 1 — column header labels
        for field, (col_idx, label) in col_layout.items():
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font = bold

        # Rows 2-4 — RAMCO format placeholders
        ws.cell(row=2, column=1, value='Rec.No')
        ws.cell(row=3, column=1, value='Combo')
        ws.cell(row=4, column=1, value='String')

        # Build a fast field→col lookup
        col_map = {field: col_idx for field, (col_idx, _) in col_layout.items()}

        # Data rows start at Excel row 5
        for idx, row_data in enumerate(bom_rows, start=1):
            excel_row = 4 + idx
            ws.cell(row=excel_row, column=1, value=row_data.get('rec_no', idx))
            for field, col_idx in col_map.items():
                if field == 'rec_no':
                    continue  # already written above
                val = row_data.get(field)
                if val is not None:
                    ws.cell(row=excel_row, column=col_idx, value=val)
            if not row_data.get('status_col'):
                ws.cell(row=excel_row, column=col_map.get('status_col', 39), value='ACTIVE')

    def _read_metadata_template(self, filename: str) -> bytes:
        path = os.path.join(self.template_dir, 'Excel Templates', filename)
        with open(path, 'rb') as f:
            return f.read()

    def _generate_from_template(self, template_filename: str, data_sheet_name: str,
                                 col_layout: Dict, bom_rows: List[Dict],
                                 settings: Dict) -> bytes:
        """Load template, populate Login column-C values, clear+rewrite data rows."""
        path = os.path.join(self.template_dir, 'Excel Templates', template_filename)
        wb = load_workbook(path)

        # --- Login sheet: only write column C (value column), keep everything else ---
        if 'Login' in wb.sheetnames:
            ws_login = wb['Login']
            for row_num, _label, key in _SETTINGS_MAP:
                ws_login.cell(row=row_num, column=3).value = _coerce(key, settings.get(key, ''))

        # --- Data sheet: clear rows 5+ then write new BOM rows ---
        if data_sheet_name in wb.sheetnames:
            ws_data = wb[data_sheet_name]
            # Blank out every cell from row 5 to the current last row
            if ws_data.max_row >= 5:
                for row in ws_data.iter_rows(min_row=5, max_row=ws_data.max_row):
                    for cell in row:
                        cell.value = None
        else:
            # Sheet absent in template (fallback): create with headers
            ws_data = wb.create_sheet(data_sheet_name)
            bold = Font(bold=True)
            for field, (col_idx, label) in col_layout.items():
                ws_data.cell(row=1, column=col_idx, value=label).font = bold
            ws_data.cell(row=2, column=1, value='Rec.No')
            ws_data.cell(row=3, column=1, value='Combo')
            ws_data.cell(row=4, column=1, value='String')

        col_map = {field: col_idx for field, (col_idx, _) in col_layout.items()}
        for idx, row_data in enumerate(bom_rows, start=1):
            excel_row = 4 + idx
            ws_data.cell(row=excel_row, column=1, value=row_data.get('rec_no', idx))
            for field, col_idx in col_map.items():
                if field == 'rec_no':
                    continue
                val = row_data.get(field)
                if val is not None:
                    ws_data.cell(row=excel_row, column=col_idx, value=val)
            if not row_data.get('status_col'):
                ws_data.cell(row=excel_row, column=col_map.get('status_col', 39), value='ACTIVE')

        return self._wb_to_bytes(wb)

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    def generate_edit_structure(self, product_name: str, bom_rows: List[Dict],
                                settings: Dict) -> Tuple[bytes, bytes, str, str]:
        """Generate EditProductStructure UserData + MetaData bytes."""
        userdata_bytes = self._generate_from_template(
            'EditProdcutStructureUserData_Conso_SFG__02_05_26.xlsx',
            'EditProdcutStructure', _EDIT_COLS, bom_rows, settings,
        )
        metadata_bytes = self._read_metadata_template('EditProdcutStructureMetaData.xlsx')
        date_str  = datetime.now().strftime('%d_%m_%Y')
        safe_name = product_name.replace(' ', '_').replace('/', '_')
        return (
            userdata_bytes,
            metadata_bytes,
            f'EditProductStructure-UserData-{safe_name}-{date_str}.xlsx',
            'EditProductStructure-MetaData.xlsx',
        )

    def generate_create_structure(self, product_name: str, bom_rows: List[Dict],
                                   settings: Dict) -> Tuple[bytes, bytes, str, str]:
        """Generate CreateProductStructure UserData + MetaData bytes."""
        userdata_bytes = self._generate_from_template(
            'CreateProductStructure_NewUserData__HCFSFG__13_05_2026__1.xlsx',
            'CreateProductStructure_New', _CREATE_COLS, bom_rows, settings,
        )
        metadata_bytes = self._read_metadata_template('CreateProductStructure_NewMetaData.xlsx')
        date_str  = datetime.now().strftime('%d_%m_%Y')
        safe_name = product_name.replace(' ', '_').replace('/', '_')
        return (
            userdata_bytes,
            metadata_bytes,
            f'CreateProductStructure-UserData-{safe_name}-{date_str}.xlsx',
            'CreateProductStructure-MetaData.xlsx',
        )

    def convert_bom_to_rows(self, bom_data: List[Dict], product_name: str,
                            flow_type: str = 'edit') -> List[Dict]:
        """Convert BOM engine output to CreateProductStructure / EditProductStructure format.

        Rec No is assigned by unique item_code (first-occurrence order) so every row
        sharing the same component gets the same Rec No.  PP Seq is always 10.
        Both warehouse code fields are always CBESFG.
        """
        # Build ps_no → rec_no mapping in first-occurrence order.
        # All rows sharing the same PS No get the same Rec No.
        ps_to_rec: Dict[str, int] = {}
        next_rec = 1
        for bom_row in bom_data:
            ps = bom_row.get('ps_no') or product_name
            if ps not in ps_to_rec:
                ps_to_rec[ps] = next_rec
                next_rec += 1

        output_rows = []
        for idx, bom_row in enumerate(bom_data, start=1):
            ps_no  = bom_row.get('ps_no') or product_name
            ps_seq = bom_row.get('ps_seq', idx)
            ic     = bom_row.get('item_code', '')
            row_dict = {
                'rec_no':        ps_to_rec[ps_no],
                'status_col':    'ACTIVE',
                'ps_no':         ps_no,
                'ps_desc':       bom_row.get('ps_desc', ''),
                'ps_item_code':  ps_no,
                'maintained_by': 'PEPS',
                'wh_code':       'CBEFG',
                'parent_qty':    1,
                'default_for':   'MRP&Rollup',
                'ps_seq':        ps_seq,
                'item_code':     ic,
                'description':   bom_row.get('description', ''),
                'qty':           bom_row.get('qty', 1),
                'uom':           bom_row.get('uom', 'NOS'),
                'pp_seq':        10,
                'wh_code_item':  bom_row.get('wh_code', ''),
                'io':            'INPUT',
                'valid_from':    '2019-04-01',
                'valid_to':      '2065-03-31',
                'min_qty':       1,
                'max_qty':       100000000,
                'stock_status':  'Accepted',
                'critical':      'No',
                'use_up_option': 'No',
            }
            if flow_type == 'create':
                row_dict.update({
                    'ps_usage':    'Standard',
                    'ps_category': 'Production',
                })
            output_rows.append(row_dict)

        # Sort by PS No so all rows for the same mattress code are grouped together.
        # ps_seq may arrive as a string from the bom_engine (REF sheet values are
        # returned via _get() which always returns str). String ordering breaks for
        # multi-digit sequences ("100" < "20"), so coerce to float for the key.
        def _seq_num(v):
            try: return float(v)
            except (TypeError, ValueError): return 0.0

        output_rows.sort(key=lambda r: (r['rec_no'], _seq_num(r.get('ps_seq', 0))))

        # Normalise ps_seq to int so Excel receives a number, not a string.
        for r in output_rows:
            try: r['ps_seq'] = int(float(r['ps_seq']))
            except (TypeError, ValueError): pass

        return output_rows

    # kept for backward compat — no longer used internally
    def populate_login_sheet(self, ws, settings: Dict):
        for row_num, label, key in _SETTINGS_MAP:
            ws[f'B{row_num}'] = label
            ws[f'C{row_num}'] = settings.get(key, '')


# Singleton
generator = ProductStructureGenerator()
